#!/usr/bin/env python3
"""
Scansione completa del file per capire la struttura reale
"""

import openpyxl

def scan_sheet_content(ws, sheet_name, max_rows=200, max_cols=20):
    """Scansiona il contenuto di un foglio"""
    
    print(f"\n🔍 SCANSIONE COMPLETA FOGLIO: {sheet_name}")
    print("=" * 60)
    
    # Informazioni generali
    print(f"📊 Dimensioni massime: {ws.max_row} righe x {ws.max_column} colonne")
    
    # Scansiono tutte le celle con contenuto
    content_cells = []
    
    scan_rows = min(max_rows, ws.max_row)
    scan_cols = min(max_cols, ws.max_column) 
    
    print(f"🔎 Scansiono area: {scan_rows} righe x {scan_cols} colonne")
    
    for row in range(1, scan_rows + 1):
        for col in range(1, scan_cols + 1):
            try:
                cell = ws.cell(row, col)
                value = cell.value
                
                if value is not None:
                    # Converto il valore in stringa per l'analisi
                    str_value = str(value).strip()
                    
                    if str_value:  # Se non è vuoto
                        col_letter = openpyxl.utils.get_column_letter(col)
                        content_cells.append({
                            'row': row,
                            'col': col,
                            'col_letter': col_letter,
                            'address': f"{col_letter}{row}",
                            'value': str_value,
                            'type': type(value).__name__
                        })
            except Exception as e:
                continue
    
    print(f"✅ Trovate {len(content_cells)} celle con contenuto")
    
    # Cerco celle con keywords NPL/Default
    npl_keywords = ['npl', 'default', 'recovery', 'stage 3', 'non performing', 'coverage', 'rettifiche', 'gbv', 'nbv', 'performing']
    
    npl_cells = []
    for cell in content_cells:
        cell_lower = cell['value'].lower()
        for keyword in npl_keywords:
            if keyword in cell_lower:
                npl_cells.append({**cell, 'keyword': keyword})
                break
    
    if npl_cells:
        print(f"\n📋 CELLE CON KEYWORDS NPL TROVATE ({len(npl_cells)}):")
        for cell in npl_cells[:20]:  # Prime 20
            print(f"   {cell['address']}: '{cell['value'][:50]}' (keyword: {cell['keyword']})")
        
        if len(npl_cells) > 20:
            print(f"   ... e altre {len(npl_cells) - 20} celle")
    
    # Cerco pattern di anni (per identificare header temporali)
    year_cells = []
    for cell in content_cells:
        if cell['type'] in ['int', 'float']:
            try:
                year_val = float(cell['value'])
                if 2020 <= year_val <= 2035:
                    year_cells.append(cell)
            except:
                continue
    
    if year_cells:
        print(f"\n📅 CELLE CON ANNI TROVATE ({len(year_cells)}):")
        
        # Raggruppo per riga per identificare header temporali
        years_by_row = {}
        for cell in year_cells:
            row = cell['row']
            if row not in years_by_row:
                years_by_row[row] = []
            years_by_row[row].append(cell)
        
        for row, cells in sorted(years_by_row.items())[:10]:
            years_str = " | ".join([f"{cell['col_letter']}({cell['value']})" for cell in sorted(cells, key=lambda x: x['col'])[:8]])
            print(f"   Riga {row}: {years_str}")
            if len(cells) > 8:
                print(f"              ... e altri {len(cells) - 8} anni")
    
    # Mostro sample di formule
    formula_cells = []
    for row in range(1, min(100, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            try:
                cell = ws.cell(row, col)
                if hasattr(cell, 'formula') and cell.formula:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    formula_cells.append({
                        'address': f"{col_letter}{row}",
                        'formula': cell.formula
                    })
            except:
                continue
    
    if formula_cells:
        print(f"\n📐 SAMPLE FORMULE TROVATE ({len(formula_cells)} totali):")
        for cell in formula_cells[:10]:
            print(f"   {cell['address']}: {cell['formula'][:60]}...")
    
    return {
        'content_cells': content_cells,
        'npl_cells': npl_cells,
        'year_cells': year_cells,
        'formula_cells': formula_cells
    }

def analyze_npl_areas_detailed(npl_cells, ws):
    """Analizza in dettaglio le aree con NPL"""
    
    if not npl_cells:
        return
    
    print(f"\n🎯 ANALISI DETTAGLIATA AREE NPL")
    print("=" * 40)
    
    # Raggruppo per aree (righe vicine)
    npl_areas = []
    current_area = []
    
    sorted_cells = sorted(npl_cells, key=lambda x: x['row'])
    
    for cell in sorted_cells:
        if not current_area or cell['row'] - current_area[-1]['row'] <= 5:
            current_area.append(cell)
        else:
            if current_area:
                npl_areas.append(current_area)
            current_area = [cell]
    
    if current_area:
        npl_areas.append(current_area)
    
    # Analizzo ogni area
    for i, area in enumerate(npl_areas, 1):
        start_row = area[0]['row']
        end_row = area[-1]['row']
        
        print(f"\n📊 AREA NPL #{i} (righe {start_row}-{end_row}):")
        
        for cell in area:
            print(f"   {cell['address']}: {cell['value']}")
        
        # Esploro l'area circostante
        print(f"\n   🔍 Esplorazione area circostante (righe {start_row-2}-{end_row+10}):")
        
        context_rows = []
        for row in range(max(1, start_row - 2), min(end_row + 11, ws.max_row + 1)):
            try:
                label = ws.cell(row, 1).value
                if label and isinstance(label, str) and label.strip():
                    
                    # Cerco valori nelle colonne successive  
                    row_values = []
                    for col in range(2, min(10, ws.max_column + 1)):
                        value = ws.cell(row, col).value
                        if value is not None:
                            col_letter = openpyxl.utils.get_column_letter(col)
                            row_values.append(f"{col_letter}:{value}")
                    
                    context_rows.append({
                        'row': row,
                        'label': label.strip(),
                        'values': row_values[:5]  # Prime 5
                    })
            except:
                continue
        
        for row_info in context_rows:
            print(f"      {row_info['row']:3d}: {row_info['label'][:40]}")
            if row_info['values']:
                print(f"           Valori: {' | '.join(row_info['values'])}")

def main():
    file_path = '/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello_bancario_completo.xlsx'
    
    try:
        print("🏦 SCANSIONE COMPLETA MODELLO BANCARIO EXCEL")
        print("=" * 50)
        
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        all_results = {}
        
        # Scansiono ogni foglio
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            results = scan_sheet_content(ws, sheet_name)
            all_results[sheet_name] = results
            
            # Analisi dettagliata delle aree NPL
            if results['npl_cells']:
                analyze_npl_areas_detailed(results['npl_cells'], ws)
        
        # Riepilogo finale
        print(f"\n\n📊 RIEPILOGO FINALE")
        print("=" * 30)
        
        for sheet_name, results in all_results.items():
            npl_count = len(results['npl_cells'])
            year_count = len(results['year_cells'])
            formula_count = len(results['formula_cells'])
            
            print(f"\n📋 {sheet_name}:")
            print(f"   🔸 {npl_count} celle NPL")
            print(f"   📅 {year_count} celle con anni")  
            print(f"   📐 {formula_count} celle con formule")
        
    except Exception as e:
        print(f"❌ ERRORE: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
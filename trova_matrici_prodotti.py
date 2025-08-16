#!/usr/bin/env python3
"""
Script per trovare le matrici specifiche di ogni prodotto nel foglio Calcoli
Basandoci sulla riga 49 "Matrice 2: RIMBORSI CAPITALE", cerchiamo le matrici per prodotto
"""

import openpyxl
from openpyxl.utils import get_column_letter

def trova_matrici_prodotti():
    """Trova le matrici per ogni prodotto"""
    
    try:
        wb = openpyxl.load_workbook('modello.xlsx', data_only=False)
        ws_calcoli = wb['Calcoli']
        
        print("🎯 RICERCA MATRICI PRODOTTO SPECIFICHE")
        print("=" * 50)
        
        # Dall'analisi precedente sappiamo che:
        # - Riga 49: "Matrice 2: RIMBORSI CAPITALE" 
        # - Righe 99-148: Le formule di calcolo
        
        # Le formule riferiscono alle righe 53-93 che dovrebbero contenere i rimborsi per prodotto
        print("🔍 Analizzando le righe 50-95 per trovare i rimborsi per prodotto...")
        print()
        
        matrici_trovate = []
        
        # Cerca nelle righe intorno al blocco rimborsi
        for row in range(50, 96):
            row_data = []
            ha_formule = False
            ha_valori = False
            
            # Analizza la riga
            for col in range(1, 15):  # A-N
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.value is not None:
                    row_data.append({
                        'cella': f"{get_column_letter(col)}{row}",
                        'valore': cell.value,
                        'colonna': col
                    })
                    
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        ha_formule = True
                    elif isinstance(cell.value, (int, float)):
                        ha_valori = True
            
            # Se la riga ha molte formule o valori, potrebbe essere una matrice prodotto
            if len(row_data) >= 8 and (ha_formule or ha_valori):
                matrici_trovate.append({
                    'riga': row,
                    'celle': row_data,
                    'ha_formule': ha_formule,
                    'ha_valori': ha_valori
                })
        
        print(f"📊 MATRICI POTENZIALI TROVATE: {len(matrici_trovate)}")
        print()
        
        # Analizza ogni matrice trovata
        for i, matrice in enumerate(matrici_trovate):
            print(f"🎯 MATRICE {i+1} - RIGA {matrice['riga']}:")
            
            # Mostra le prime 8 celle
            for j, cella_data in enumerate(matrice['celle'][:8]):
                valore_short = str(cella_data['valore'])[:50] + "..." if len(str(cella_data['valore'])) > 50 else str(cella_data['valore'])
                print(f"   {cella_data['cella']}: {valore_short}")
            
            if len(matrice['celle']) > 8:
                print(f"   ... e altre {len(matrice['celle']) - 8} celle")
            
            print(f"   Tipo: {'Formule' if matrice['ha_formule'] else 'Valori numerici'}")
            print()
        
        # Ora analizza il pattern delle formule nelle righe 99-118 (prime 20 righe di calcolo)
        print("🔍 ANALISI PATTERN FORMULE (righe 99-118):")
        print("=" * 50)
        
        for row in range(99, 119):  # Prime 20 righe di prodotto
            prodotto_num = row - 98  # Prodotto 1 = riga 99, ecc.
            
            # Analizza la formula nella colonna B
            cell_b = ws_calcoli.cell(row=row, column=2)  # Colonna B
            
            if cell_b.value and isinstance(cell_b.value, str) and cell_b.value.startswith('='):
                formula = cell_b.value
                
                # Estrai i riferimenti dalla formula
                import re
                # Cerca pattern come B7, B53, ecc.
                refs = re.findall(r'[A-Z]\d+', formula)
                
                print(f"📌 PRODOTTO {prodotto_num} (Riga {row}):")
                print(f"   Cella: B{row}")
                print(f"   Formula: {formula[:80]}...")
                print(f"   Riferimenti trovati: {list(set(refs))}")
                
                # Il riferimento che dovrebbe cambiare per ogni prodotto
                # Dalle formule vedo pattern come B7-B53, quindi B53 è quello che conta
                riga_rimborso = None
                for ref in refs:
                    if ref.startswith('B') and '5' in ref and len(ref) <= 4:
                        riga_rimborso = ref
                        break
                
                if riga_rimborso:
                    print(f"   🎯 RIGA RIMBORSO IDENTIFICATA: {riga_rimborso}")
                    
                    # Suggerisci la modifica per collegarsi agli input
                    # Prodotto 1 = colonna D, Prodotto 2 = colonna E, ecc.
                    col_input = get_column_letter(3 + prodotto_num)  # D=4, E=5, ecc.
                    
                    print(f"   📝 SUGGERIMENTO: Cambiare {riga_rimborso} con Input.{col_input}67")
                
                print()
        
        # Ora proviamo a identificare la PRIMA CELLA di ogni matrice prodotto
        print("🎯 IDENTIFICAZIONE PRIMA CELLA OGNI MATRICE PRODOTTO:")
        print("=" * 60)
        
        # Basandoci sull'analisi, sembra che:
        # - Le matrici rimborsi inizino intorno alla riga 53
        # - Ogni prodotto ha la sua riga specifica
        
        for prodotto in range(1, 21):  # 20 prodotti
            riga_matrice = 52 + prodotto  # Prodotto 1 = riga 53, Prodotto 2 = riga 54, ecc.
            
            # Verifica se la riga esiste e ha contenuto
            cell_b = ws_calcoli.cell(row=riga_matrice, column=2)  # Colonna B
            
            col_input = get_column_letter(3 + prodotto)  # D=4, E=5, F=6, ecc.
            
            print(f"📌 PRODOTTO {prodotto}:")
            print(f"   Prima cella matrice stimata: B{riga_matrice}")
            print(f"   Colonna Input corrispondente: {col_input}")
            
            if cell_b.value is not None:
                valore_current = str(cell_b.value)[:50] + "..." if len(str(cell_b.value)) > 50 else str(cell_b.value)
                print(f"   Valore attuale: {valore_current}")
                
                # Suggerisci la formula corretta
                if isinstance(cell_b.value, str) and cell_b.value.startswith('='):
                    print(f"   🔧 FORMULA CORRETTA SUGGERITA: =Input.{col_input}67")
                else:
                    print(f"   🔧 SOSTITUIRE CON: =Input.{col_input}67")
            else:
                print(f"   ⚠️ Cella vuota - inserire: =Input.{col_input}67")
            
            print()
        
    except Exception as e:
        print(f"ERRORE: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    trova_matrici_prodotti()
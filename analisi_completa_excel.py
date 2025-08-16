#!/usr/bin/env python3
"""
Analisi completa e dettagliata del file modello.xlsx
Analizza ogni foglio, cella, formula e struttura dati
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import TYPE_FORMULA, TYPE_NUMERIC, TYPE_STRING, TYPE_BOOL, TYPE_ERROR
from collections import defaultdict
import re
from typing import Dict, List, Tuple, Any, Optional

class ExcelAnalyzer:
    def __init__(self, file_path: str):
        """Inizializza l'analyzer con il file Excel"""
        self.file_path = file_path
        self.workbook = None
        self.analysis_report = {}
        
    def load_workbook(self) -> bool:
        """Carica il workbook Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            return True
        except Exception as e:
            print(f"❌ ERRORE nel caricamento del file: {e}")
            return False
    
    def get_cell_type_description(self, cell) -> str:
        """Restituisce una descrizione del tipo di cella"""
        if cell.data_type == TYPE_FORMULA:
            return "FORMULA"
        elif cell.data_type == TYPE_NUMERIC:
            return "NUMERO"
        elif cell.data_type == TYPE_STRING:
            return "TESTO"
        elif cell.data_type == TYPE_BOOL:
            return "BOOLEAN"
        elif cell.data_type == TYPE_ERROR:
            return "ERRORE"
        else:
            return "SCONOSCIUTO"
    
    def analyze_formula(self, formula: str) -> Dict[str, Any]:
        """Analizza una formula Excel"""
        analysis = {
            'original': formula,
            'functions': [],
            'references': [],
            'external_references': [],
            'complexity': 'SEMPLICE'
        }
        
        # Estrai funzioni Excel
        function_pattern = r'([A-Z][A-Z0-9_]*)\s*\('
        functions = re.findall(function_pattern, formula)
        analysis['functions'] = list(set(functions))
        
        # Estrai riferimenti a celle
        cell_ref_pattern = r'([A-Z]+[0-9]+)|([A-Z]+:[A-Z]+)|([A-Z]+[0-9]+:[A-Z]+[0-9]+)'
        references = re.findall(cell_ref_pattern, formula)
        # Flatten e rimuovi stringhe vuote
        refs = [ref for group in references for ref in group if ref]
        analysis['references'] = list(set(refs))
        
        # Estrai riferimenti esterni (altri fogli)
        external_ref_pattern = r"'?([^'!]+)'?!([A-Z]+[0-9]+|[A-Z]+:[A-Z]+|[A-Z]+[0-9]+:[A-Z]+[0-9]+)"
        external_refs = re.findall(external_ref_pattern, formula)
        analysis['external_references'] = external_refs
        
        # Determina complessità
        if len(analysis['functions']) > 3 or len(analysis['references']) > 10:
            analysis['complexity'] = 'COMPLESSA'
        elif len(analysis['functions']) > 1 or len(analysis['references']) > 3:
            analysis['complexity'] = 'MEDIA'
        
        return analysis
    
    def scan_worksheet_cells(self, worksheet) -> Dict[str, Any]:
        """Scansiona tutte le celle di un foglio di lavoro"""
        cells_data = {
            'all_cells': {},
            'formulas': {},
            'numeric_values': {},
            'text_values': {},
            'empty_cells_in_used_range': [],
            'sections': [],
            'max_row': 0,
            'max_col': 0
        }
        
        # Determina il range utilizzato reale
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        cells_data['max_row'] = max_row
        cells_data['max_col'] = max_col
        
        print(f"    🔍 Scansione range: A1:{get_column_letter(max_col)}{max_row}")
        
        # Scansiona ogni cella nel range utilizzato
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"
                
                if cell.value is not None:
                    cell_info = {
                        'address': cell_address,
                        'value': cell.value,
                        'type': self.get_cell_type_description(cell),
                        'data_type': cell.data_type,
                        'number_format': cell.number_format,
                        'row': row,
                        'col': col
                    }
                    
                    cells_data['all_cells'][cell_address] = cell_info
                    
                    # Categorizza per tipo
                    if cell.data_type == TYPE_FORMULA:
                        formula_analysis = self.analyze_formula(str(cell.value))
                        cell_info['formula_analysis'] = formula_analysis
                        cells_data['formulas'][cell_address] = cell_info
                    elif cell.data_type == TYPE_NUMERIC:
                        cells_data['numeric_values'][cell_address] = cell_info
                    elif cell.data_type == TYPE_STRING:
                        cells_data['text_values'][cell_address] = cell_info
                elif row <= max_row and col <= max_col:
                    # Cella vuota nel range utilizzato
                    cells_data['empty_cells_in_used_range'].append(cell_address)
        
        return cells_data
    
    def identify_sections(self, cells_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Identifica sezioni e tabelle nel foglio"""
        sections = []
        text_cells = cells_data['text_values']
        
        # Cerca intestazioni potenziali (celle di testo in grassetto o in posizioni chiave)
        potential_headers = {}
        for address, cell_info in text_cells.items():
            value = str(cell_info['value']).strip()
            if len(value) > 2 and not value.isdigit():
                potential_headers[address] = {
                    'text': value,
                    'row': cell_info['row'],
                    'col': cell_info['col']
                }
        
        # Raggruppa per righe per identificare sezioni
        rows_with_headers = defaultdict(list)
        for address, header_info in potential_headers.items():
            rows_with_headers[header_info['row']].append(header_info)
        
        # Identifica sezioni basate su pattern
        for row, headers in rows_with_headers.items():
            if len(headers) >= 2:  # Almeno 2 intestazioni nella stessa riga
                section = {
                    'type': 'TABLE_HEADER',
                    'row': row,
                    'headers': [h['text'] for h in headers],
                    'col_range': (min(h['col'] for h in headers), max(h['col'] for h in headers))
                }
                sections.append(section)
        
        return sections
    
    def analyze_temporal_structure(self, cells_data: Dict[str, Any]) -> Dict[str, Any]:
        """Analizza la struttura temporale (anni, trimestri, mesi)"""
        temporal_info = {
            'years': [],
            'quarters': [],
            'months': [],
            'temporal_headers': {},
            'temporal_pattern': None
        }
        
        text_cells = cells_data['text_values']
        
        # Pattern per identificare anni, trimestri, mesi
        year_pattern = r'20\d{2}'
        quarter_pattern = r'[Qq][1-4]|T[1-4]|Q[1-4]'
        month_pattern = r'(Gen|Feb|Mar|Apr|Mag|Giu|Lug|Ago|Set|Ott|Nov|Dic|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
        
        for address, cell_info in text_cells.items():
            value = str(cell_info['value']).strip()
            
            # Cerca anni
            years = re.findall(year_pattern, value)
            if years:
                temporal_info['years'].extend(years)
                temporal_info['temporal_headers'][address] = {'type': 'YEAR', 'values': years}
            
            # Cerca trimestri
            quarters = re.findall(quarter_pattern, value, re.IGNORECASE)
            if quarters:
                temporal_info['quarters'].extend(quarters)
                temporal_info['temporal_headers'][address] = {'type': 'QUARTER', 'values': quarters}
            
            # Cerca mesi
            months = re.findall(month_pattern, value, re.IGNORECASE)
            if months:
                temporal_info['months'].extend(months)
                temporal_info['temporal_headers'][address] = {'type': 'MONTH', 'values': months}
        
        # Rimuovi duplicati e ordina
        temporal_info['years'] = sorted(list(set(temporal_info['years'])))
        temporal_info['quarters'] = list(set(temporal_info['quarters']))
        temporal_info['months'] = list(set(temporal_info['months']))
        
        # Determina il pattern temporale predominante
        if temporal_info['years']:
            if temporal_info['quarters']:
                temporal_info['temporal_pattern'] = 'QUARTERLY'
            elif temporal_info['months']:
                temporal_info['temporal_pattern'] = 'MONTHLY'
            else:
                temporal_info['temporal_pattern'] = 'YEARLY'
        
        return temporal_info
    
    def analyze_worksheet(self, sheet_name: str) -> Dict[str, Any]:
        """Analizza completamente un singolo foglio di lavoro"""
        print(f"\n📊 ANALISI FOGLIO: {sheet_name}")
        print("=" * 50)
        
        worksheet = self.workbook[sheet_name]
        
        # Scansiona tutte le celle
        print("  🔍 Scansione celle...")
        cells_data = self.scan_worksheet_cells(worksheet)
        
        # Identifica sezioni
        print("  📋 Identificazione sezioni...")
        sections = self.identify_sections(cells_data)
        cells_data['sections'] = sections
        
        # Analizza struttura temporale
        print("  📅 Analisi struttura temporale...")
        temporal_info = self.analyze_temporal_structure(cells_data)
        
        # Statistiche generali
        stats = {
            'total_cells_with_data': len(cells_data['all_cells']),
            'formulas_count': len(cells_data['formulas']),
            'numeric_values_count': len(cells_data['numeric_values']),
            'text_values_count': len(cells_data['text_values']),
            'sections_count': len(sections),
            'max_row': cells_data['max_row'],
            'max_col': cells_data['max_col']
        }
        
        return {
            'sheet_name': sheet_name,
            'cells_data': cells_data,
            'temporal_info': temporal_info,
            'stats': stats
        }
    
    def analyze_inter_sheet_connections(self) -> Dict[str, Any]:
        """Analizza le connessioni tra fogli"""
        connections = {
            'sheet_references': defaultdict(list),
            'reference_graph': defaultdict(set)
        }
        
        for sheet_name in self.workbook.sheetnames:
            sheet_analysis = self.analysis_report.get(sheet_name, {})
            formulas = sheet_analysis.get('cells_data', {}).get('formulas', {})
            
            for cell_address, cell_info in formulas.items():
                formula_analysis = cell_info.get('formula_analysis', {})
                external_refs = formula_analysis.get('external_references', [])
                
                for sheet_ref, cell_ref in external_refs:
                    connections['sheet_references'][sheet_name].append({
                        'from_cell': cell_address,
                        'to_sheet': sheet_ref,
                        'to_cell': cell_ref,
                        'formula': cell_info['value']
                    })
                    connections['reference_graph'][sheet_name].add(sheet_ref)
        
        return connections
    
    def run_complete_analysis(self) -> bool:
        """Esegue l'analisi completa del file Excel"""
        print("🚀 AVVIO ANALISI COMPLETA DEL FILE EXCEL")
        print("=" * 60)
        
        if not self.load_workbook():
            return False
        
        print(f"📁 File: {self.file_path}")
        print(f"📑 Fogli trovati: {self.workbook.sheetnames}")
        
        # Analizza ogni foglio
        for sheet_name in self.workbook.sheetnames:
            self.analysis_report[sheet_name] = self.analyze_worksheet(sheet_name)
        
        # Analizza connessioni tra fogli
        print(f"\n🔗 ANALISI CONNESSIONI TRA FOGLI")
        print("=" * 40)
        self.analysis_report['inter_sheet_connections'] = self.analyze_inter_sheet_connections()
        
        return True
    
    def print_detailed_report(self):
        """Stampa il report dettagliato dell'analisi"""
        print("\n\n" + "=" * 80)
        print("📋 REPORT DETTAGLIATO ANALISI EXCEL")
        print("=" * 80)
        
        # Report per ogni foglio
        for sheet_name, analysis in self.analysis_report.items():
            if sheet_name == 'inter_sheet_connections':
                continue
                
            print(f"\n\n🔷 FOGLIO: {sheet_name}")
            print("=" * 60)
            
            stats = analysis['stats']
            print(f"📊 STATISTICHE GENERALI:")
            print(f"   • Dimensioni: {stats['max_col']} colonne × {stats['max_row']} righe")
            print(f"   • Celle con dati: {stats['total_cells_with_data']}")
            print(f"   • Formule: {stats['formulas_count']}")
            print(f"   • Valori numerici: {stats['numeric_values_count']}")
            print(f"   • Valori di testo: {stats['text_values_count']}")
            print(f"   • Sezioni identificate: {stats['sections_count']}")
            
            # Struttura temporale
            temporal = analysis['temporal_info']
            if temporal['temporal_pattern']:
                print(f"\n📅 STRUTTURA TEMPORALE:")
                print(f"   • Pattern: {temporal['temporal_pattern']}")
                print(f"   • Anni: {temporal['years']}")
                if temporal['quarters']:
                    print(f"   • Trimestri: {temporal['quarters']}")
                if temporal['months']:
                    print(f"   • Mesi: {temporal['months']}")
            
            # Sezioni identificate
            sections = analysis['cells_data']['sections']
            if sections:
                print(f"\n📋 SEZIONI IDENTIFICATE:")
                for i, section in enumerate(sections, 1):
                    print(f"   {i}. Riga {section['row']}: {', '.join(section['headers'])}")
            
            # Mappatura completa delle celle con contenuto
            print(f"\n🗺️  MAPPA COMPLETA DELLE CELLE:")
            all_cells = analysis['cells_data']['all_cells']
            
            # Raggruppa per righe per una visualizzazione ordinata
            rows_data = defaultdict(list)
            for address, cell_info in all_cells.items():
                rows_data[cell_info['row']].append(cell_info)
            
            for row in sorted(rows_data.keys()):
                cells_in_row = sorted(rows_data[row], key=lambda x: x['col'])
                print(f"\n   📍 RIGA {row}:")
                for cell_info in cells_in_row:
                    value_preview = str(cell_info['value'])[:50]
                    if len(str(cell_info['value'])) > 50:
                        value_preview += "..."
                    print(f"      {cell_info['address']}: [{cell_info['type']}] {value_preview}")
            
            # Analisi dettagliata delle formule
            formulas = analysis['cells_data']['formulas']
            if formulas:
                print(f"\n🧮 ANALISI DETTAGLIATA FORMULE:")
                for address, formula_info in formulas.items():
                    formula_analysis = formula_info['formula_analysis']
                    print(f"\n   📍 {address}: {formula_info['value']}")
                    print(f"      • Complessità: {formula_analysis['complexity']}")
                    if formula_analysis['functions']:
                        print(f"      • Funzioni: {', '.join(formula_analysis['functions'])}")
                    if formula_analysis['references']:
                        print(f"      • Riferimenti interni: {', '.join(formula_analysis['references'])}")
                    if formula_analysis['external_references']:
                        ext_refs = [f"{sheet}!{cell}" for sheet, cell in formula_analysis['external_references']]
                        print(f"      • Riferimenti esterni: {', '.join(ext_refs)}")
            
            # Valori numerici significativi
            numeric_values = analysis['cells_data']['numeric_values']
            if numeric_values:
                print(f"\n💰 VALORI NUMERICI SIGNIFICATIVI:")
                # Mostra solo valori non zero e significativi
                significant_values = {addr: info for addr, info in numeric_values.items() 
                                    if info['value'] != 0 and abs(info['value']) >= 0.01}
                
                if significant_values:
                    for address, value_info in list(significant_values.items())[:20]:  # Primi 20
                        formatted_value = f"{value_info['value']:,.2f}" if isinstance(value_info['value'], (int, float)) else str(value_info['value'])
                        print(f"      {address}: {formatted_value}")
                    if len(significant_values) > 20:
                        print(f"      ... e altri {len(significant_values) - 20} valori")
        
        # Report connessioni tra fogli
        connections = self.analysis_report.get('inter_sheet_connections', {})
        if connections and connections['sheet_references']:
            print(f"\n\n🔗 CONNESSIONI TRA FOGLI")
            print("=" * 40)
            
            for source_sheet, refs in connections['sheet_references'].items():
                if refs:
                    print(f"\n📊 Da {source_sheet}:")
                    for ref in refs[:10]:  # Primi 10 riferimenti
                        print(f"   {ref['from_cell']} → {ref['to_sheet']}!{ref['to_cell']}")
                    if len(refs) > 10:
                        print(f"   ... e altri {len(refs) - 10} riferimenti")
        
        # Riepilogo finale
        print(f"\n\n📋 RIEPILOGO GENERALE")
        print("=" * 40)
        total_cells = sum(analysis['stats']['total_cells_with_data'] 
                         for sheet_name, analysis in self.analysis_report.items() 
                         if sheet_name != 'inter_sheet_connections')
        total_formulas = sum(analysis['stats']['formulas_count'] 
                           for sheet_name, analysis in self.analysis_report.items() 
                           if sheet_name != 'inter_sheet_connections')
        
        print(f"• Fogli totali: {len(self.workbook.sheetnames)}")
        print(f"• Celle totali con dati: {total_cells}")
        print(f"• Formule totali: {total_formulas}")
        
        # Stato di completamento per foglio
        print(f"\n📊 STATO COMPLETAMENTO PER FOGLIO:")
        for sheet_name, analysis in self.analysis_report.items():
            if sheet_name == 'inter_sheet_connections':
                continue
            
            stats = analysis['stats']
            completion_score = min(100, (stats['total_cells_with_data'] / (stats['max_row'] * stats['max_col']) * 100) if stats['max_row'] * stats['max_col'] > 0 else 0)
            
            status = "🔴 VUOTO" if completion_score < 5 else "🟡 PARZIALE" if completion_score < 50 else "🟢 COMPLETO"
            print(f"   {sheet_name}: {status} ({completion_score:.1f}% celle utilizzate)")


def main():
    """Funzione principale per eseguire l'analisi"""
    file_path = "/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello.xlsx"
    
    analyzer = ExcelAnalyzer(file_path)
    
    if analyzer.run_complete_analysis():
        analyzer.print_detailed_report()
        print(f"\n✅ ANALISI COMPLETATA CON SUCCESSO!")
    else:
        print(f"\n❌ ERRORE DURANTE L'ANALISI!")


if __name__ == "__main__":
    main()
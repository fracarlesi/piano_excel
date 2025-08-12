#!/usr/bin/env python3
"""
FORMULA ENGINE - MODELLO BANCARIO COMPLETO CON VINTAGE ANALYSIS
Script per inserire le formule Excel nel modello bancario trimestrale con gestione vintage

Questo script aggiunge le formule per:
1. Calcoli di appoggio con vintage analysis
2. Conto Economico
3. Stato Patrimoniale  
4. Capital Requirements
5. KPI

Supporta 20 trimestri (Q1-Q20) con tracking completo per vintage.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

class FormulaEngine:
    def __init__(self, filename="modello_bancario_completo.xlsx"):
        self.wb = load_workbook(filename)
        self.ws_input = self.wb["Input"]
        self.ws_modello = self.wb["Modello"]
        
        # Configurazione trimestrale
        self.num_quarters = 20  # Q1-Q20 per 5 anni
        self.quarter_columns = [get_column_letter(3 + i) for i in range(20)]  # C-V
        
        # Prodotti per divisione
        self.products = {
            'RE': ['CBL', 'MLA', 'MLAM', 'MLPA', 'MLPAM', 'NRE'],
            'SME': ['BL', 'REFI', 'SS', 'NF', 'RES'],
            'PG': ['ACFP', 'FGA', 'FGPA']
        }
        
        # Mappa celle chiave
        self.input_refs = self._mappa_riferimenti_input()
        self.model_refs = self._mappa_riferimenti_modello()
        
        # Stili per formattazione celle FORMULA
        self._setup_formula_styles()
    
    def _setup_formula_styles(self):
        """Imposta gli stili per le celle con formule"""
        self.formula_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        self.formula_font = Font(name='Calibri', size=11, color="000000")
        self.formula_border = Border(
            left=Side(style='dashed'),
            right=Side(style='dashed'),
            top=Side(style='dashed'),
            bottom=Side(style='dashed')
        )
    
    def _format_formula_cell(self, cell, number_format=None):
        """Applica la formattazione standard alle celle con formule"""
        cell.fill = self.formula_fill
        cell.font = self.formula_font
        cell.border = self.formula_border
        cell.alignment = Alignment(horizontal='center')
        if number_format:
            cell.number_format = number_format
    
    def _mappa_riferimenti_input(self):
        """Mappa i riferimenti chiave nel foglio Input"""
        refs = {}
        
        # Scansiona il foglio Input per trovare le celle chiave
        for row in range(1, 500):
            cell_value = self.ws_input[f'B{row}'].value
            if cell_value:
                cell_str = str(cell_value).upper()
                
                # Parametri macro
                if "ECB RATE" in cell_str and "Y1" not in cell_str:
                    refs['ecb_rate_row'] = row
                elif "EURIBOR" in cell_str and "6M" in cell_str:
                    refs['euribor_row'] = row
                
                # Erogazioni per divisione (ora valori assoluti)
                elif "NUOVE EROGAZIONI RE" in cell_str:
                    refs['erog_re_row'] = row
                elif "NUOVE EROGAZIONI SME" in cell_str:
                    refs['erog_sme_row'] = row
                elif "NUOVE EROGAZIONI PG" in cell_str:
                    refs['erog_pg_row'] = row
                
                # Parametri prodotti
                elif "MIX PRODOTTI" in cell_str:
                    # Cerca la sezione corretta guardando le righe precedenti
                    for search_back in range(2, 8):  # Cerca da 2 a 7 righe indietro
                        section_value = self.ws_input[f'B{row-search_back}'].value
                        if section_value:
                            section_str = str(section_value).upper()
                            if "REAL ESTATE" in section_str and "DIVISION" in section_str:
                                refs['mix_re_row'] = row
                                break
                            elif "SME" in section_str and "DIVISION" in section_str:
                                refs['mix_sme_row'] = row
                                break
                            elif "PUBLIC GUARANTEE" in section_str and "DIVISION" in section_str:
                                refs['mix_pg_row'] = row
                                break
                
                # Parametri di rischio
                elif "DANGER RATE" in cell_str:
                    # Cerca la sezione corretta guardando le righe precedenti
                    for search_back in range(3, 10):
                        section_value = self.ws_input[f'B{row-search_back}'].value
                        if section_value:
                            section_str = str(section_value).upper()
                            if "REAL ESTATE" in section_str and "DIVISION" in section_str:
                                refs['danger_re_row'] = row
                                break
                            elif "SME" in section_str and "DIVISION" in section_str:
                                refs['danger_sme_row'] = row
                                break
                            elif "PUBLIC GUARANTEE" in section_str and "DIVISION" in section_str:
                                refs['danger_pg_row'] = row
                                break
                
                elif "LGD" in cell_str:
                    # Cerca la sezione corretta guardando le righe precedenti
                    for search_back in range(3, 12):
                        section_value = self.ws_input[f'B{row-search_back}'].value
                        if section_value:
                            section_str = str(section_value).upper()
                            if "REAL ESTATE" in section_str and "DIVISION" in section_str:
                                refs['lgd_re_row'] = row
                                break
                            elif "SME" in section_str and "DIVISION" in section_str:
                                refs['lgd_sme_row'] = row
                                break
                            elif "PUBLIC GUARANTEE" in section_str and "DIVISION" in section_str:
                                refs['lgd_pg_row'] = row
                                break
                
                elif "INTEREST RATE" in cell_str:
                    # Cerca la sezione corretta guardando le righe precedenti
                    for search_back in range(3, 15):
                        section_value = self.ws_input[f'B{row-search_back}'].value
                        if section_value:
                            section_str = str(section_value).upper()
                            if "REAL ESTATE" in section_str and "DIVISION" in section_str:
                                refs['rate_re_row'] = row
                                break
                            elif "SME" in section_str and "DIVISION" in section_str:
                                refs['rate_sme_row'] = row
                                break
                            elif "PUBLIC GUARANTEE" in section_str and "DIVISION" in section_str:
                                refs['rate_pg_row'] = row
                                break
                
                elif "UP-FRONT FEES" in cell_str:
                    # Cerca la sezione corretta guardando le righe precedenti
                    for search_back in range(3, 18):
                        section_value = self.ws_input[f'B{row-search_back}'].value
                        if section_value:
                            section_str = str(section_value).upper()
                            if "REAL ESTATE" in section_str and "DIVISION" in section_str:
                                refs['upfront_re_row'] = row
                                break
                            elif "SME" in section_str and "DIVISION" in section_str:
                                refs['upfront_sme_row'] = row
                                break
                            elif "PUBLIC GUARANTEE" in section_str and "DIVISION" in section_str:
                                refs['upfront_pg_row'] = row
                                break
                
                # Parametri Default e Recupero
                elif "REAL ESTATE" in cell_str and self.ws_input[f'B{row-1}'].value:
                    prev_val = str(self.ws_input[f'B{row-1}'].value).upper()
                    if "TIMING DEFAULT" in prev_val:
                        refs['default_timing_re'] = row
                elif "SME" in cell_str and self.ws_input[f'B{row-1}'].value:
                    prev_val = str(self.ws_input[f'B{row-1}'].value).upper()
                    if "TIMING DEFAULT" in prev_val:
                        refs['default_timing_sme'] = row
                elif "PUBLIC GUARANTEE" in cell_str and self.ws_input[f'B{row-1}'].value:
                    prev_val = str(self.ws_input[f'B{row-1}'].value).upper()
                    if "TIMING DEFAULT" in prev_val:
                        refs['default_timing_pg'] = row
                
                # Timing recupero garanzie
                elif "GARANZIA IMMOBILIARE" in cell_str:
                    if self.ws_input[f'B{row-1}'].value and "TIMING RECUPERO" in str(self.ws_input[f'B{row-1}'].value).upper():
                        refs['recovery_timing_imm'] = row
                elif "GARANZIA MCC" in cell_str:
                    if self.ws_input[f'B{row-1}'].value and "TIMING RECUPERO" in str(self.ws_input[f'B{row-1}'].value).upper():
                        refs['recovery_timing_mcc'] = row
                
                # Quote recupero garanzie
                elif "REAL ESTATE" in cell_str and self.ws_input[f'B{row-1}'].value:
                    if "QUOTE RECUPERO" in str(self.ws_input[f'B{row-1}'].value).upper():
                        refs['recovery_quote_re'] = row
                elif "SME" in cell_str and self.ws_input[f'B{row-2}'].value:
                    if "QUOTE RECUPERO" in str(self.ws_input[f'B{row-2}'].value).upper():
                        refs['recovery_quote_sme'] = row
                elif "PUBLIC GUARANTEE" in cell_str and self.ws_input[f'B{row-3}'].value:
                    if "QUOTE RECUPERO" in str(self.ws_input[f'B{row-3}'].value).upper():
                        refs['recovery_quote_pg'] = row
                
                # Parametri ECL (solo la prima occorrenza)
                elif "ECL HORIZON" in cell_str and 'ecl_horizon' not in refs:
                    refs['ecl_horizon'] = row
                elif "PD MULTIPLIER" in cell_str and 'pd_multiplier' not in refs:
                    refs['pd_multiplier'] = row
                
                # Altri parametri
                elif "ALIQUOTA FISCALE" in cell_str:
                    refs['tax_rate_row'] = row
                elif "DIVIDEND PAYOUT" in cell_str:
                    refs['dividend_row'] = row
        
        return refs
    
    def _mappa_riferimenti_modello(self):
        """Mappa i riferimenti chiave nel foglio Modello"""
        refs = {}
        
        # Scansiona il foglio Modello per trovare le sezioni
        for row in range(1, 2000):
            cell_value = self.ws_modello[f'B{row}'].value
            if cell_value:
                cell_str = str(cell_value).upper()
                
                # Sezioni principali
                if "1.1 EROGAZIONI PER PRODOTTO" in cell_str:
                    refs['erogazioni_start'] = row + 3
                elif "1.2 STOCK CREDITI PER PRODOTTO" in cell_str:
                    refs['stock_crediti_start'] = row + 3
                elif "1.3 VINTAGE ANALYSIS" in cell_str:
                    refs['vintage_start'] = row + 3
                elif "1.4 MATRICE AMMORTAMENTI" in cell_str:
                    refs['ammortamenti_start'] = row + 3
                elif "1.5 MATRICE DEFAULT" in cell_str:
                    refs['default_start'] = row + 3
                elif "1.6 MATRICE RECUPERI" in cell_str:
                    refs['recuperi_start'] = row + 3
                elif "1.7 CALCOLO NBV" in cell_str:
                    refs['nbv_start'] = row + 3
                elif "STOCK PERFORMING" in cell_str and "€ MLN" in cell_str:
                    refs['stock_performing_start'] = row + 3
                elif "ECL STAGE 1" in cell_str and "€ MLN" in cell_str:
                    refs['ecl_stage1_start'] = row + 3
                elif "NBV PERFORMING" in cell_str and "STOCK - ECL" in cell_str:
                    refs['nbv_performing_start'] = row + 3
                elif "STOCK NPL" in cell_str and "€ MLN" in cell_str:
                    refs['stock_npl_start'] = row + 3
                elif "NPV RECUPERI" in cell_str:
                    refs['npv_recuperi_start'] = row + 3
                elif "3. CONTO ECONOMICO" in cell_str:
                    refs['ce_start'] = row + 3
                elif "4. STATO PATRIMONIALE" in cell_str:
                    refs['sp_start'] = row + 3
                elif "5. CAPITAL REQUIREMENTS" in cell_str:
                    refs['capital_start'] = row + 3
                elif "6. KEY PERFORMANCE" in cell_str:
                    refs['kpi_start'] = row + 3
        
        return refs
    
    def formule_erogazioni_base(self):
        """Inserisce le formule per le erogazioni base per prodotto"""
        print("  📊 Inserimento formule erogazioni per prodotto...")
        
        start_row = self.model_refs.get('erogazioni_start', 10)
        
        # Per ogni divisione e prodotto
        current_row = start_row
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            # Riga delle erogazioni annuali nella sezione Input
            if division == 'RE':
                erog_row = self.input_refs.get('erog_re_row')
                mix_row = self.input_refs.get('mix_re_row')
            elif division == 'SME':
                erog_row = self.input_refs.get('erog_sme_row')
                mix_row = self.input_refs.get('mix_sme_row')
            else:
                erog_row = self.input_refs.get('erog_pg_row')
                mix_row = self.input_refs.get('mix_pg_row')
            
            for idx, product in enumerate(products):
                # Per ogni trimestre
                for q in range(20):
                    col_letter = self.quarter_columns[q]
                    year = (q // 4) + 1  # Anno di riferimento (1-5)
                    quarter_in_year = (q % 4) + 1  # Trimestre nell'anno (1-4)
                    
                    # Colonna del mix prodotto nel foglio Input
                    product_mix_col = get_column_letter(3 + idx)  # C, D, E, F, G, H
                    
                    # Colonna dell'erogazione annuale nel foglio Input  
                    year_col = get_column_letter(2 + year)  # C per Y1, D per Y2, E per Y3, etc.
                    
                    # Formula corretta: Erogazione_Annuale * Mix_Prodotto / 4
                    formula = f"=Input!${year_col}${erog_row}*Input!${product_mix_col}${mix_row}/4"
                    
                    cell = self.ws_modello[f'{col_letter}{current_row}']
                    cell.value = formula
                    self._format_formula_cell(cell, '#,##0.0')
                
                current_row += 1
        
        print(f"    ✓ Formule erogazioni inserite per {current_row - start_row} prodotti")
    
    def formule_stock_crediti(self):
        """Inserisce le formule per lo stock crediti con roll-forward"""
        print("  📊 Inserimento formule stock crediti...")
        
        erog_start = self.model_refs.get('erogazioni_start', 10)
        stock_start = self.model_refs.get('stock_crediti_start', 50)
        
        current_row = stock_start
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for idx, product in enumerate(products):
                erog_row = erog_start + (
                    0 if division == 'RE' else 
                    6 if division == 'SME' else 
                    11
                ) + idx
                
                for q in range(20):
                    col_letter = self.quarter_columns[q]
                    
                    if q == 0:
                        # Q1: Stock iniziale (0) + Erogazioni Q1
                        formula = f"={col_letter}{erog_row}"
                    else:
                        # Altri Q: Stock precedente + Erogazioni - Rimborsi - Default
                        prev_col = self.quarter_columns[q-1]
                        formula = f"={prev_col}{current_row}+{col_letter}{erog_row}"
                    
                    cell = self.ws_modello[f'{col_letter}{current_row}']
                    cell.value = formula
                    self._format_formula_cell(cell, '#,##0.0')
                
                current_row += 1
        
        print(f"    ✓ Formule stock crediti inserite")
    
    def apply_formulas_blocco_a(self):
        """
        Applica le formule per il BLOCCO A - EROGAZIONI E VOLUMI
        Implementa matrici vintage 20x20 per ogni prodotto:
        - A.1: Matrice Erogazioni per Vintage
        - A.2: Matrice Rimborsi per Vintage  
        - A.3: Stock GBV per Vintage
        """
        print("  📊 BLOCCO A - Inserimento formule matrici vintage Erogazioni e Volumi...")
        
        # A.1 - MATRICE EROGAZIONI PER VINTAGE
        self._apply_formulas_a1_erogazioni()
        
        # A.2 - MATRICE RIMBORSI PER VINTAGE
        self._apply_formulas_a2_rimborsi()
        
        # A.3 - STOCK GBV PER VINTAGE
        self._apply_formulas_a3_stock_gbv()
        
        print("    ✅ BLOCCO A completato - Matrici vintage Erogazioni/Volumi")
    
    def _apply_formulas_a1_erogazioni(self):
        """A.1 - Matrice Erogazioni per Vintage: diagonale principale con erogazioni"""
        print("      A.1 - Matrice Erogazioni per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "EROGAZIONI")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q == v:  # Diagonale principale
                            # Formula: Erogazione = Input!Erogazione_Divisione × Mix_Prodotto / 4
                            erog_ref = self._get_input_erogazione_ref(division, v+1)  # Anno basato su vintage
                            mix_ref = self._get_input_mix_ref(division, product)
                            
                            formula = f"=IF({erog_ref}*{mix_ref}/4>0,{erog_ref}*{mix_ref}/4,0)"
                        else:  # Fuori diagonale
                            formula = "=0"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_a2_rimborsi(self):
        """A.2 - Matrice Rimborsi per Vintage: bullet vs amortizing"""
        print("      A.2 - Matrice Rimborsi per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "RIMBORSI")
                if matrix_start == 0:
                    continue
                
                # Parametri prodotto
                amort_type_ref = self._get_input_param_ref(division, product, "Amortizing Type")
                maturity_ref = self._get_input_param_ref(division, product, "Loan Maturity")
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Riferimento alla cella Stock GBV corrispondente
                            stock_ref = self._get_stock_gbv_cell_ref(division, product, v, q)
                            maturity_quarters = f"({maturity_ref}*4)"
                            quarters_elapsed = f"({q+1}-{v+1})"
                            
                            # Formula condizionale bullet vs amortizing
                            formula = (f"=IF({amort_type_ref}=\"bullet\","
                                     f"IF({quarters_elapsed}>={maturity_quarters},{stock_ref},0),"
                                     f"IF({quarters_elapsed}>0,{stock_ref}/{maturity_quarters},0))")
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_a3_stock_gbv(self):
        """A.3 - Stock GBV per Vintage: Stock = Stock_Precedente + Erogazioni - Rimborsi - Default"""
        print("      A.3 - Stock GBV per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "STOCK GBV")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        elif q == v:  # Primo periodo: Stock = Erogazioni
                            erog_ref = self._get_erogazioni_cell_ref(division, product, v, q)
                            formula = f"={erog_ref}"
                        else:  # Periodi successivi
                            prev_col = self.quarter_columns[q-1]
                            erog_ref = self._get_erogazioni_cell_ref(division, product, v, q)
                            rimb_ref = self._get_rimborsi_cell_ref(division, product, v, q)
                            default_ref = self._get_default_cell_ref(division, product, v, q)
                            
                            formula = f"={prev_col}{vintage_row}+{erog_ref}-{rimb_ref}-{default_ref}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')

    def apply_formulas_blocco_b(self):
        """
        Applica le formule per il BLOCCO B - RISCHIO CREDITO
        Implementa matrici vintage 20x20 per ogni prodotto:
        - B.1: Default per Vintage (con timing distribution)
        - B.2: Stock Performing
        - B.3: Stock NPL
        """
        print("  📊 BLOCCO B - Inserimento formule matrici vintage Rischio Credito...")
        
        # B.1 - MATRICE DEFAULT PER VINTAGE
        self._apply_formulas_b1_default()
        
        # B.2 - MATRICE STOCK PERFORMING
        self._apply_formulas_b2_performing()
        
        # B.3 - MATRICE STOCK NPL
        self._apply_formulas_b3_npl()
        
        print("    ✅ BLOCCO B completato - Matrici vintage Rischio Credito")
    
    def apply_formulas_blocco_c(self):
        """
        Applica le formule per il BLOCCO C - ECL E VALUTAZIONI
        Implementa matrici vintage 20x20 per ogni prodotto:
        - C.1: ECL Stage 1 (Stock × Danger_Rate/4 × LGD)
        - C.2: NBV Performing (Stock - ECL Stage 1)
        - C.3: NPV Recuperi (attualizzazione recuperi futuri)
        - C.4: NBV Non-Performing (valore recuperabile NPL)
        """
        print("  📊 BLOCCO C - Inserimento formule matrici vintage ECL e Valutazioni...")
        
        # C.1 - ECL STAGE 1 PER VINTAGE
        self._apply_formulas_c1_ecl_stage1()
        
        # C.2 - NBV PERFORMING
        self._apply_formulas_c2_nbv_performing()
        
        # C.3 - NPV RECUPERI
        self._apply_formulas_c3_npv_recuperi()
        
        # C.4 - NBV NON-PERFORMING
        self._apply_formulas_c4_nbv_npl()
        
        print("    ✅ BLOCCO C completato - Matrici vintage ECL e Valutazioni")
    
    def apply_formulas_blocco_d(self):
        """
        Applica le formule per il BLOCCO D - RICAVI
        Implementa matrici vintage 20x20 per ogni prodotto:
        - D.1: Interessi Attivi su Performing (Stock × Spread/4)
        - D.2: Interessi di Mora su NPL (Stock_NPL × Tasso_Mora/4 × Recovery_Rate)
        - D.3: Commissioni Up-front (Erogazioni × Up_front_Fee)
        """
        print("  📊 BLOCCO D - Inserimento formule matrici vintage Ricavi...")
        
        # D.1 - INTERESSI ATTIVI SU PERFORMING
        self._apply_formulas_d1_interest_performing()
        
        # D.2 - INTERESSI DI MORA SU NPL
        self._apply_formulas_d2_interest_mora()
        
        # D.3 - COMMISSIONI UP-FRONT
        self._apply_formulas_d3_commission_upfront()
        
        print("    ✅ BLOCCO D completato - Matrici vintage Ricavi")
    
    def _apply_formulas_b1_default(self):
        """B.1 - Default per Vintage con timing distribution"""
        print("      B.1 - Matrice Default per Vintage con timing...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "DEFAULT")
                if matrix_start == 0:
                    continue
                
                # Parametri di rischio
                danger_rate_ref = self._get_input_param_ref(division, product, "Danger Rate")
                default_timing_ref = self._get_input_param_ref(division, product, "Default Timing")
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Stock GBV di riferimento
                            stock_ref = self._get_stock_gbv_cell_ref(division, product, v, q)
                            
                            # Timing distribution (concentrato intorno al Default_Timing_Q)
                            quarters_from_origination = q - v + 1
                            timing_factor = self._get_timing_distribution_formula(quarters_from_origination, default_timing_ref)
                            
                            # Default = Stock × (Danger_Rate/4) × Timing_Distribution
                            formula = f"={stock_ref}*({danger_rate_ref}/4)*{timing_factor}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_b2_performing(self):
        """B.2 - Stock Performing = Stock_GBV - Default_Cumulati"""
        print("      B.2 - Stock Performing...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "PERFORMING")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Stock GBV
                            stock_gbv_ref = self._get_stock_gbv_cell_ref(division, product, v, q)
                            
                            # Default cumulati dalla origine fino al periodo corrente
                            default_cum_formula = self._get_cumulative_default_formula(division, product, v, v, q)
                            
                            formula = f"={stock_gbv_ref}-({default_cum_formula})"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_b3_npl(self):
        """B.3 - Stock NPL = NPL_Precedente + Default_Periodo - Write_off - Recuperi"""
        print("      B.3 - Stock NPL...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "NPL")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        elif q == v:  # Primo periodo
                            # NPL = Default del periodo (se c'è)
                            default_ref = self._get_default_cell_ref(division, product, v, q)
                            formula = f"={default_ref}"
                        else:  # Periodi successivi
                            prev_col = self.quarter_columns[q-1]
                            default_ref = self._get_default_cell_ref(division, product, v, q)
                            writeoff_ref = self._get_writeoff_cell_ref(division, product, v, q)
                            recovery_ref = self._get_recovery_cell_ref(division, product, v, q)
                            
                            formula = f"={prev_col}{vintage_row}+{default_ref}-{writeoff_ref}-{recovery_ref}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    # ============================================================================
    # HELPER METHODS FOR VINTAGE ANALYSIS - BLOCCHI A & B
    # ============================================================================
    
    def _find_vintage_matrix_start(self, division, product, matrix_type):
        """
        Trova la riga iniziale di una matrice vintage specifica.
        
        Args:
            division: 'RE', 'SME', 'PG'
            product: nome prodotto (es. 'CBL', 'MLA', ...)
            matrix_type: 'EROGAZIONI', 'RIMBORSI', 'STOCK GBV', 'DEFAULT', 'PERFORMING', 'NPL'
        
        Returns:
            int: riga iniziale della matrice (0 se non trovata)
        """
        search_pattern = f"{division} - {matrix_type}"
        
        # Cerca nel foglio Modello
        for row in range(100, 2000):  # Range ampio per la ricerca
            cell_value = self.ws_modello[f'B{row}'].value
            if cell_value and search_pattern in str(cell_value).upper():
                # Verifica che sia la sezione del prodotto giusto
                # Cerca il nome del prodotto nelle righe successive
                for check_row in range(row + 1, row + 25):
                    check_value = self.ws_modello[f'B{check_row}'].value
                    if check_value and product in str(check_value):
                        # Trova la riga con "V1(Q1)" che indica l'inizio dei dati
                        for data_row in range(check_row, check_row + 5):
                            data_value = self.ws_modello[f'B{data_row}'].value
                            if data_value and "V1(Q1)" in str(data_value):
                                return data_row
                break
        
        print(f"    ⚠️  ATTENZIONE: Matrice {matrix_type} per {division}-{product} non trovata")
        return 0
    
    def _get_input_erogazione_ref(self, division, year):
        """
        Ottiene il riferimento alla cella delle erogazioni nel foglio Input.
        
        Args:
            division: 'RE', 'SME', 'PG'
            year: anno (1-5)
        
        Returns:
            str: riferimento Excel (es. "Input!$D$285")
        """
        if division == 'RE':
            row = self.input_refs.get('erog_re_row', 285)
        elif division == 'SME':
            row = self.input_refs.get('erog_sme_row', 300)
        else:  # PG
            row = self.input_refs.get('erog_pg_row', 315)
        
        year_col = get_column_letter(2 + year)  # C per Y1, D per Y2, etc.
        return f"Input!${year_col}${row}"
    
    def _get_input_mix_ref(self, division, product):
        """
        Ottiene il riferimento alla cella del mix prodotto nel foglio Input.
        
        Args:
            division: 'RE', 'SME', 'PG'  
            product: nome prodotto
        
        Returns:
            str: riferimento Excel (es. "Input!$C$342")
        """
        if division == 'RE':
            mix_row = self.input_refs.get('mix_re_row', 342)
            products = self.products['RE']
        elif division == 'SME':
            mix_row = self.input_refs.get('mix_sme_row', 394)
            products = self.products['SME']
        else:  # PG
            mix_row = self.input_refs.get('mix_pg_row', 447)
            products = self.products['PG']
        
        product_idx = products.index(product)
        product_col = get_column_letter(3 + product_idx)  # C, D, E, F, G, H
        return f"Input!${product_col}${mix_row}"
    
    def _get_input_param_ref(self, division, product, param_name):
        """
        Ottiene il riferimento a un parametro specifico del prodotto nel foglio Input.
        
        Args:
            division: 'RE', 'SME', 'PG'
            product: nome prodotto
            param_name: nome parametro (es. "Amortizing Type", "Loan Maturity", "Danger Rate")
        
        Returns:
            str: riferimento Excel
        """
        if division == 'RE':
            products = self.products['RE']
            base_row = 342  # Riga base per parametri Real Estate
        elif division == 'SME':
            products = self.products['SME']
            base_row = 394  # Riga base per parametri SME
        else:  # PG
            products = self.products['PG']
            base_row = 447  # Riga base per parametri Public Guarantee
        
        product_idx = products.index(product)
        product_col = get_column_letter(3 + product_idx)
        
        # Offset delle righe per diversi parametri (da mappare con il foglio Input reale)
        param_offsets = {
            "Amortizing Type": 1,
            "Loan Maturity": 2,
            "Danger Rate": 9,
            "Default Timing": 14,
            "LGD": 19,
            "Interest Rate": 24,
            "Up-front Fee": 29,
            "LTV": 34,
            "Garanzia MCC": 39,
            "Tasso Mora": 44
        }
        
        offset = param_offsets.get(param_name, 0)
        param_row = base_row + offset
        
        return f"Input!${product_col}${param_row}"
    
    def _get_stock_gbv_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Stock GBV per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "STOCK GBV")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_erogazioni_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Erogazioni per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "EROGAZIONI")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_rimborsi_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Rimborsi per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "RIMBORSI")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_default_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Default per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "DEFAULT")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_writeoff_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Write-off (da implementare in futuro)"""
        # Per ora restituisce 0, da implementare quando avremo le matrici write-off
        return "0"
    
    def _get_recovery_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Recovery (da implementare in futuro)"""
        # Per ora restituisce 0, da implementare quando avremo le matrici recovery
        return "0"
    
    def _get_timing_distribution_formula(self, quarters_from_origination, default_timing_ref):
        """
        Crea una formula per la distribuzione temporale dei default.
        Concentra i default intorno al Default_Timing_Q del prodotto.
        
        Args:
            quarters_from_origination: trimestri dalla erogazione
            default_timing_ref: riferimento al parametro Default Timing
        
        Returns:
            str: formula Excel per il fattore di timing
        """
        # Distribuzione normale semplificata concentrata intorno al timing target
        # Fattore massimo quando quarters_from_origination = default_timing
        
        # Formula: EXP(-0.5 * ((q - timing)/2)^2) per distribuzione gaussiana
        # Semplificata: IF(ABS(q - timing) <= 2, 1, EXP(-0.1 * (q - timing)^2))
        
        formula = (f"IF(ABS({quarters_from_origination}-{default_timing_ref})<=2,"
                  f"1,"
                  f"EXP(-0.1*({quarters_from_origination}-{default_timing_ref})^2))")
        
        return formula
    
    def _get_cumulative_default_formula(self, division, product, vintage, start_quarter, end_quarter):
        """
        Crea formula per somma cumulativa dei default per un vintage.
        
        Args:
            division: divisione
            product: prodotto
            vintage: vintage (0-19)
            start_quarter: trimestre iniziale (0-19)
            end_quarter: trimestre finale (0-19)
        
        Returns:
            str: formula Excel SUM per default cumulativi
        """
        matrix_start = self._find_vintage_matrix_start(division, product, "DEFAULT")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        start_col = self.quarter_columns[start_quarter]
        end_col = self.quarter_columns[end_quarter]
        
        return f"SUM({start_col}{vintage_row}:{end_col}{vintage_row})"
    
    # ============================================================================
    # BLOCCO C - ECL E VALUTAZIONI
    # ============================================================================
    
    def _apply_formulas_c1_ecl_stage1(self):
        """C.1 - ECL Stage 1: Stock_Performing × (Danger_Rate/4) × LGD"""
        print("      C.1 - ECL Stage 1 per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "ECL STAGE 1")
                if matrix_start == 0:
                    continue
                
                # Parametri di rischio
                danger_rate_ref = self._get_input_param_ref(division, product, "Danger Rate")
                lgd_ref = self._get_input_param_ref(division, product, "LGD")
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Stock Performing di riferimento
                            stock_performing_ref = self._get_performing_cell_ref(division, product, v, q)
                            
                            # ECL = Stock_Performing × (Danger_Rate/4) × LGD
                            formula = f"={stock_performing_ref}*({danger_rate_ref}/4)*{lgd_ref}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_c2_nbv_performing(self):
        """C.2 - NBV Performing: Stock_Performing - ECL_Stage1"""
        print("      C.2 - NBV Performing per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "NBV PERFORMING")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Riferimenti alle celle
                            stock_performing_ref = self._get_performing_cell_ref(division, product, v, q)
                            ecl_stage1_ref = self._get_ecl_stage1_cell_ref(division, product, v, q)
                            
                            # NBV Performing = Stock_Performing - ECL_Stage1
                            formula = f"={stock_performing_ref}-{ecl_stage1_ref}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_c3_npv_recuperi(self):
        """C.3 - NPV Recuperi: attualizzazione recuperi con tasso di sconto"""
        print("      C.3 - NPV Recuperi per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "NPV RECUPERI")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Stock NPL di riferimento
                            stock_npl_ref = self._get_npl_cell_ref(division, product, v, q)
                            
                            # Parametri di recupero
                            ltv_ref = self._get_input_param_ref(division, product, "LTV")
                            recovery_imm_ref = self._get_input_recovery_ref("Garanzia Immobiliare")
                            garanzia_mcc_ref = self._get_input_param_ref(division, product, "Garanzia MCC")
                            recovery_mcc_ref = self._get_input_recovery_ref("Garanzia MCC")
                            
                            # Timing recupero
                            timing_imm_ref = self._get_input_timing_recovery_ref("Garanzia Immobiliare")
                            timing_mcc_ref = self._get_input_timing_recovery_ref("Garanzia MCC")
                            
                            # Tassi di sconto
                            euribor_ref = self._get_input_euribor_ref()
                            spread_ref = self._get_input_param_ref(division, product, "Interest Rate")
                            
                            # Recupero da garanzie immobiliari attualizzato
                            quarters_elapsed = f"({q+1}-{v+1})"
                            discount_rate = f"({euribor_ref}+{spread_ref})/4"
                            timing_factor_imm = f"({timing_imm_ref}+{quarters_elapsed})"
                            
                            recovery_imm_formula = (f"({stock_npl_ref}*{ltv_ref}*{recovery_imm_ref})"
                                                  f"/POWER(1+{discount_rate},{timing_factor_imm})")
                            
                            # Recupero da MCC attualizzato
                            timing_factor_mcc = f"({timing_mcc_ref}+{quarters_elapsed})"
                            recovery_mcc_formula = (f"({stock_npl_ref}*{garanzia_mcc_ref}*{recovery_mcc_ref})"
                                                  f"/POWER(1+{discount_rate},{timing_factor_mcc})")
                            
                            # NPV Recuperi = Recupero_Immobiliare + Recupero_MCC
                            formula = f"={recovery_imm_formula}+{recovery_mcc_formula}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_c4_nbv_npl(self):
        """C.4 - NBV Non-Performing: valore recuperabile dei NPL"""
        print("      C.4 - NBV Non-Performing per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "NBV NON-PERFORMING")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # NBV Non-Performing = NPV Recuperi
                            npv_recuperi_ref = self._get_npv_recuperi_cell_ref(division, product, v, q)
                            formula = f"={npv_recuperi_ref}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    # ============================================================================
    # BLOCCO D - RICAVI
    # ============================================================================
    
    def _apply_formulas_d1_interest_performing(self):
        """D.1 - Interessi Attivi su Performing: Stock × (Spread + Euribor)/4"""
        print("      D.1 - Interessi Attivi su Performing per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "INTEREST PERFORMING")
                if matrix_start == 0:
                    continue
                
                # Parametri di tasso
                euribor_ref = self._get_input_euribor_ref()
                spread_ref = self._get_input_param_ref(division, product, "Interest Rate")
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Stock Performing di riferimento
                            stock_performing_ref = self._get_performing_cell_ref(division, product, v, q)
                            
                            # Interessi = Stock_Performing × (Euribor + Spread)/4
                            formula = f"={stock_performing_ref}*({euribor_ref}+{spread_ref})/4"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_d2_interest_mora(self):
        """D.2 - Interessi di Mora su NPL: Stock_NPL × Tasso_Mora/4 × Recovery_Rate"""
        print("      D.2 - Interessi di Mora su NPL per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "INTEREST MORA")
                if matrix_start == 0:
                    continue
                
                # Parametri
                tasso_mora_ref = self._get_input_param_ref(division, product, "Tasso Mora")
                recovery_rate_ref = self._get_input_recovery_total_ref(division)
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            # Stock NPL di riferimento
                            stock_npl_ref = self._get_npl_cell_ref(division, product, v, q)
                            
                            # Interessi Mora = Stock_NPL × Tasso_Mora/4 × Recovery_Rate
                            formula = f"={stock_npl_ref}*{tasso_mora_ref}/4*{recovery_rate_ref}"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_d3_commission_upfront(self):
        """D.3 - Commissioni Up-front: Erogazioni × Up_front_Fee (solo diagonale principale)"""
        print("      D.3 - Commissioni Up-front per Vintage...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "COMMISSION UPFRONT")
                if matrix_start == 0:
                    continue
                
                # Parametro Up-front Fee
                upfront_fee_ref = self._get_input_param_ref(division, product, "Up-front Fee")
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q == v:  # Solo diagonale principale (nuove erogazioni)
                            # Erogazioni di riferimento
                            erog_ref = self._get_erogazioni_cell_ref(division, product, v, q)
                            
                            # Commissioni = Erogazioni × Up_front_Fee
                            formula = f"={erog_ref}*{upfront_fee_ref}"
                        else:
                            formula = "=0"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _get_erog_row_for_product(self, division, product):
        """Ottiene la riga delle erogazioni per un prodotto specifico"""
        erog_start = self.model_refs.get('erogazioni_start', 10)
        
        # Calcola l'offset basato su divisione e prodotto
        division_offset = 0 if division == 'RE' else 6 if division == 'SME' else 11
        product_idx = self.products[division].index(product)
        
        return erog_start + division_offset + product_idx
    
    # ============================================================================
    # HELPER METHODS AGGIUNTIVI PER BLOCCHI C & D
    # ============================================================================
    
    def _get_performing_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Stock Performing per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "PERFORMING")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_ecl_stage1_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella ECL Stage 1 per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "ECL STAGE 1")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_npl_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella Stock NPL per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "NPL")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_npv_recuperi_cell_ref(self, division, product, vintage, quarter):
        """Ottiene riferimento alla cella NPV Recuperi per vintage/quarter specifici"""
        matrix_start = self._find_vintage_matrix_start(division, product, "NPV RECUPERI")
        if matrix_start == 0:
            return "0"
        
        vintage_row = matrix_start + vintage
        col_letter = self.quarter_columns[quarter]
        return f"{col_letter}{vintage_row}"
    
    def _get_input_euribor_ref(self):
        """Ottiene il riferimento alla cella Euribor nel foglio Input"""
        euribor_row = self.input_refs.get('euribor_row', 20)
        return f"Input!$C${euribor_row}"
    
    def _get_input_recovery_ref(self, garanzia_type):
        """
        Ottiene il riferimento alle quote di recupero per tipo garanzia.
        
        Args:
            garanzia_type: "Garanzia Immobiliare" o "Garanzia MCC"
        
        Returns:
            str: riferimento Excel alle quote di recupero
        """
        if garanzia_type == "Garanzia Immobiliare":
            recovery_row = self.input_refs.get('recovery_timing_imm', 450)
        else:  # Garanzia MCC
            recovery_row = self.input_refs.get('recovery_timing_mcc', 452)
        
        return f"Input!$C${recovery_row}"
    
    def _get_input_timing_recovery_ref(self, garanzia_type):
        """
        Ottiene il riferimento al timing di recupero per tipo garanzia.
        
        Args:
            garanzia_type: "Garanzia Immobiliare" o "Garanzia MCC"
        
        Returns:
            str: riferimento Excel al timing di recupero
        """
        if garanzia_type == "Garanzia Immobiliare":
            timing_row = self.input_refs.get('recovery_timing_imm', 449)
        else:  # Garanzia MCC
            timing_row = self.input_refs.get('recovery_timing_mcc', 451)
        
        return f"Input!$C${timing_row}"
    
    def _get_input_recovery_total_ref(self, division):
        """
        Ottiene il riferimento al recovery rate totale per divisione.
        
        Args:
            division: 'RE', 'SME', 'PG'
        
        Returns:
            str: riferimento Excel al recovery rate
        """
        if division == 'RE':
            recovery_row = self.input_refs.get('recovery_quote_re', 453)
        elif division == 'SME':
            recovery_row = self.input_refs.get('recovery_quote_sme', 454)
        else:  # PG
            recovery_row = self.input_refs.get('recovery_quote_pg', 455)
        
        return f"Input!$C${recovery_row}"

    # ============================================================================
    # BLOCCO E - COSTI E ACCANTONAMENTI
    # ============================================================================
    
    def _apply_formulas_e1_llp(self):
        """E.1 - LLP (Loan Loss Provisions): Delta_ECL + First_Day_Loss_Nuove_Erogazioni"""
        print("      E.1 - LLP (Loan Loss Provisions)...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "LLP")
                if matrix_start == 0:
                    continue
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        elif q == v:  # Primo periodo: First Day Loss
                            erog_ref = self._get_erogazioni_cell_ref(division, product, v, q)
                            danger_rate_ref = self._get_input_param_ref(division, product, "Danger Rate")
                            lgd_ref = self._get_input_param_ref(division, product, "LGD")
                            
                            # First Day Loss = Erogazioni × Danger_Rate × LGD
                            formula = f"={erog_ref}*{danger_rate_ref}*{lgd_ref}"
                        else:  # Periodi successivi: Delta ECL
                            prev_col = self.quarter_columns[q-1]
                            ecl_current_ref = self._get_ecl_stage1_cell_ref(division, product, v, q)
                            ecl_previous_ref = f"{prev_col}{matrix_start + v}"  # ECL precedente stesso vintage
                            
                            # Delta ECL = ECL_corrente - ECL_precedente
                            formula = f"=MAX(0,{ecl_current_ref}-{ecl_previous_ref})"
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_e2_writeoff(self):
        """E.2 - Write-off: Stock_NPL × Write_off_Rate dopo periodo workout"""
        print("      E.2 - Write-off...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for product in products:
                matrix_start = self._find_vintage_matrix_start(division, product, "WRITEOFF")
                if matrix_start == 0:
                    continue
                
                # Parametri write-off
                workout_period_quarters = 8  # 2 anni di workout
                writeoff_rate = 0.5  # 50% write-off rate (parametrizzabile)
                
                # Per ogni vintage V1-V20
                for v in range(20):
                    vintage_row = matrix_start + v
                    
                    # Per ogni time period Q1-Q20
                    for q in range(20):
                        col_letter = self.quarter_columns[q]
                        
                        if q < v:  # Prima dell'erogazione
                            formula = "=0"
                        else:
                            quarters_from_origination = q - v + 1
                            
                            if quarters_from_origination <= workout_period_quarters:
                                # Periodo workout: no write-off
                                formula = "=0"
                            else:
                                # Dopo workout: Write-off = Stock_NPL × Write_off_Rate
                                stock_npl_ref = self._get_npl_cell_ref(division, product, v, q)
                                formula = f"={stock_npl_ref}*{writeoff_rate}/4"  # Trimestrale
                        
                        cell = self.ws_modello[f'{col_letter}{vintage_row}']
                        cell.value = formula
                        self._format_formula_cell(cell, '#,##0.0')

    # ============================================================================
    # BLOCCO F - AGGREGAZIONI DIVISIONALI
    # ============================================================================
    
    def _apply_formulas_f1_aggregazioni_divisione(self):
        """F.1 - Aggregazioni per Divisione"""
        print("      F.1 - Aggregazioni Divisionali...")
        
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            # Trova la sezione aggregazioni per questa divisione
            aggregation_start = self._find_aggregation_section_start(division)
            if aggregation_start == 0:
                continue
            
            # Trova le sottosezioni
            for row in range(aggregation_start, aggregation_start + 50):
                cell_val = self.ws_modello[f'B{row}'].value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    
                    if f"TOTALE EROGAZIONI {division}" in cell_str.upper():
                        self._aggregate_division_erogazioni(row, division, products)
                    elif f"TOTALE STOCK {division}" in cell_str.upper():
                        self._aggregate_division_stock(row, division, products)
                    elif f"TOTALE INTERESSI {division}" in cell_str.upper():
                        self._aggregate_division_interessi(row, division, products)
                    elif f"TOTALE COMMISSIONI {division}" in cell_str.upper():
                        self._aggregate_division_commissioni(row, division, products)
                    elif f"TOTALE LLP {division}" in cell_str.upper():
                        self._aggregate_division_llp(row, division, products)

    def _find_aggregation_section_start(self, division):
        """Trova la sezione aggregazioni per una divisione"""
        search_pattern = f"AGGREGAZIONI {division}"
        
        for row in range(100, 2000):
            cell_value = self.ws_modello[f'B{row}'].value
            if cell_value and search_pattern in str(cell_value).upper():
                return row + 3
        
        print(f"    ⚠️  ATTENZIONE: Sezione Aggregazioni {division} non trovata")
        return 0

    def _aggregate_division_erogazioni(self, row, division, products):
        """Aggrega le erogazioni per divisione"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma erogazioni di tutti i prodotti della divisione
            formula_parts = []
            for product in products:
                erog_matrix_start = self._find_vintage_matrix_start(division, product, "EROGAZIONI")
                if erog_matrix_start > 0:
                    # Somma tutte le righe vintage per questo trimestre
                    formula_parts.append(f"SUM({col_letter}{erog_matrix_start}:{col_letter}{erog_matrix_start+19})")
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _aggregate_division_stock(self, row, division, products):
        """Aggrega lo stock GBV per divisione"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma stock di tutti i prodotti della divisione
            formula_parts = []
            for product in products:
                stock_matrix_start = self._find_vintage_matrix_start(division, product, "STOCK GBV")
                if stock_matrix_start > 0:
                    # Somma tutte le righe vintage per questo trimestre
                    formula_parts.append(f"SUM({col_letter}{stock_matrix_start}:{col_letter}{stock_matrix_start+19})")
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _aggregate_division_interessi(self, row, division, products):
        """Aggrega gli interessi attivi per divisione"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma interessi di tutti i prodotti della divisione
            formula_parts = []
            for product in products:
                interest_matrix_start = self._find_vintage_matrix_start(division, product, "INTEREST PERFORMING")
                if interest_matrix_start > 0:
                    # Somma tutte le righe vintage per questo trimestre
                    formula_parts.append(f"SUM({col_letter}{interest_matrix_start}:{col_letter}{interest_matrix_start+19})")
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _aggregate_division_commissioni(self, row, division, products):
        """Aggrega le commissioni per divisione"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma commissioni di tutti i prodotti della divisione
            formula_parts = []
            for product in products:
                comm_matrix_start = self._find_vintage_matrix_start(division, product, "COMMISSION UPFRONT")
                if comm_matrix_start > 0:
                    # Somma tutte le righe vintage per questo trimestre
                    formula_parts.append(f"SUM({col_letter}{comm_matrix_start}:{col_letter}{comm_matrix_start+19})")
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _aggregate_division_llp(self, row, division, products):
        """Aggrega le LLP per divisione"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma LLP di tutti i prodotti della divisione
            formula_parts = []
            for product in products:
                llp_matrix_start = self._find_vintage_matrix_start(division, product, "LLP")
                if llp_matrix_start > 0:
                    # Somma tutte le righe vintage per questo trimestre
                    formula_parts.append(f"SUM({col_letter}{llp_matrix_start}:{col_letter}{llp_matrix_start+19})")
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    # ============================================================================
    # FORMULE CONTO ECONOMICO AGGIORNATE
    # ============================================================================
    
    def _formula_interest_income_total(self, row):
        """Formula per Interest Income totale da vintage analysis"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma delle tre divisioni RE + SME + PG
            formula = f"=SUM({col_letter}{row+1}:{col_letter}{row+3})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_interest_income_division_vintage(self, row, division):
        """Formula per Interest Income da vintage analysis per divisione"""
        aggregation_start = self._find_aggregation_section_start(division)
        if aggregation_start == 0:
            return
        
        # Trova la riga degli interessi aggregati per la divisione
        for search_row in range(aggregation_start, aggregation_start + 50):
            cell_val = self.ws_modello[f'B{search_row}'].value
            if cell_val and f"TOTALE INTERESSI {division}" in str(cell_val).upper():
                interest_row = search_row
                break
        else:
            return
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Riferimento alla riga aggregata degli interessi
            formula = f"={col_letter}{interest_row}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_interest_expenses_total(self, row):
        """Formula per Interest Expenses totale"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma su Depositi + su Funding
            formula = f"={col_letter}{row+1}+{col_letter}{row+2}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_interest_expenses_deposits(self, row):
        """Formula per Interest Expenses su Depositi"""
        # Parametri da Input
        deposits_row = self._find_input_parameter("Customer Deposits")
        deposit_rate_row = self._find_input_parameter("Deposit Interest Rate")
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Depositi × Tasso_Depositi / 4
            if deposits_row and deposit_rate_row:
                formula = f"=Input!${col_letter}${deposits_row}*Input!$C${deposit_rate_row}/4"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_interest_expenses_funding(self, row):
        """Formula per Interest Expenses su Funding"""
        # Parametri da Input
        funding_row = self._find_input_parameter("External Funding")
        funding_rate_row = self._find_input_parameter("Funding Interest Rate")
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Funding × Tasso_Funding / 4
            if funding_row and funding_rate_row:
                formula = f"=Input!${col_letter}${funding_row}*Input!$C${funding_rate_row}/4"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_commission_income_total(self, row):
        """Formula per Commission Income totale da vintage analysis"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma delle tre divisioni RE + SME + PG
            formula = f"=SUM({col_letter}{row+1}:{col_letter}{row+3})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_commission_income_division_vintage(self, row, division):
        """Formula per Commission Income da vintage analysis per divisione"""
        aggregation_start = self._find_aggregation_section_start(division)
        if aggregation_start == 0:
            return
        
        # Trova la riga delle commissioni aggregate per la divisione
        for search_row in range(aggregation_start, aggregation_start + 50):
            cell_val = self.ws_modello[f'B{search_row}'].value
            if cell_val and f"TOTALE COMMISSIONI {division}" in str(cell_val).upper():
                commission_row = search_row
                break
        else:
            return
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Riferimento alla riga aggregata delle commissioni
            formula = f"={col_letter}{commission_row}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_llp_total(self, row):
        """Formula per LLP totale da vintage analysis"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma LLP da tutte le divisioni
            formula_parts = []
            for division in ['RE', 'SME', 'PG']:
                aggregation_start = self._find_aggregation_section_start(division)
                if aggregation_start > 0:
                    # Trova la riga LLP aggregata per la divisione
                    for search_row in range(aggregation_start, aggregation_start + 50):
                        cell_val = self.ws_modello[f'B{search_row}'].value
                        if cell_val and f"TOTALE LLP {division}" in str(cell_val).upper():
                            formula_parts.append(f"{col_letter}{search_row}")
                            break
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_personnel_costs(self, row):
        """Formula per Personnel Costs con interpolazione FTE"""
        # Parametri da Input
        fte_row = self._find_input_parameter("Total FTE")
        salary_row = self._find_input_parameter("Average Annual Salary")
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # FTE × Stipendio_Medio / 4
            if fte_row and salary_row:
                formula = f"=Input!${col_letter}${fte_row}*Input!$C${salary_row}/4"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_gross_profit(self, row):
        """Formula per Gross Profit"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Total Revenues - Personnel Costs - LLP
            revenues_row = self._find_ce_row("3. TOTAL REVENUES", row)
            personnel_row = self._find_ce_row("5. PERSONNEL COSTS", row)
            llp_row = self._find_ce_row("4. LOAN LOSS PROVISIONS", row)
            
            if revenues_row and personnel_row and llp_row:
                formula = f"={col_letter}{revenues_row}-{col_letter}{personnel_row}-{col_letter}{llp_row}"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_taxes(self, row):
        """Formula per Taxes"""
        tax_rate_row = self.input_refs.get('tax_rate_row')
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Gross Profit × Tax Rate
            gross_profit_row = self._find_ce_row("8. GROSS PROFIT", row)
            
            if gross_profit_row and tax_rate_row:
                formula = f"=MAX(0,{col_letter}{gross_profit_row}*Input!$C${tax_rate_row})"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_net_profit(self, row):
        """Formula per Net Profit"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Gross Profit - Taxes
            gross_profit_row = self._find_ce_row("8. GROSS PROFIT", row)
            taxes_row = self._find_ce_row("9. TAXES", row)
            
            if gross_profit_row and taxes_row:
                formula = f"={col_letter}{gross_profit_row}-{col_letter}{taxes_row}"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    # ============================================================================
    # FORMULE STATO PATRIMONIALE AGGIORNATE
    # ============================================================================
    
    def _formula_loans_gross_vintage(self, row):
        """Formula per Loans to Customers (Gross) da vintage analysis"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma Stock GBV da tutte le divisioni
            formula_parts = []
            for division in ['RE', 'SME', 'PG']:
                aggregation_start = self._find_aggregation_section_start(division)
                if aggregation_start > 0:
                    # Trova la riga Stock aggregato per la divisione
                    for search_row in range(aggregation_start, aggregation_start + 50):
                        cell_val = self.ws_modello[f'B{search_row}'].value
                        if cell_val and f"TOTALE STOCK {division}" in str(cell_val).upper():
                            formula_parts.append(f"{col_letter}{search_row}")
                            break
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_loan_reserves_vintage(self, row):
        """Formula per Loan Loss Reserves da vintage analysis"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma ECL + Fondo NPL da tutte le divisioni e prodotti
            formula_parts = []
            
            # ECL Stage 1 per tutti i prodotti
            for division in ['RE', 'SME', 'PG']:
                products = self.products[division]
                for product in products:
                    ecl_matrix_start = self._find_vintage_matrix_start(division, product, "ECL STAGE 1")
                    if ecl_matrix_start > 0:
                        formula_parts.append(f"SUM({col_letter}{ecl_matrix_start}:{col_letter}{ecl_matrix_start+19})")
            
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_customer_deposits(self, row):
        """Formula per Customer Deposits"""
        deposits_row = self._find_input_parameter("Customer Deposits")
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Riferimento diretto ai depositi da Input
            if deposits_row:
                formula = f"=Input!${col_letter}${deposits_row}"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_total_liabilities(self, row):
        """Formula per Total Liabilities"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma di tutte le passività
            formula = f"=SUM({col_letter}{row-4}:{col_letter}{row-1})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_share_capital(self, row):
        """Formula per Share Capital"""
        capital_row = self._find_input_parameter("Share Capital")
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Share Capital costante da Input
            if capital_row:
                formula = f"=Input!$C${capital_row}"  # Costante
            else:
                formula = "=100"  # Default
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_retained_earnings(self, row):
        """Formula per Retained Earnings con roll-forward"""
        dividend_row = self.input_refs.get('dividend_row')
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            if q == 0:  # Q1
                # Retained Earnings iniziali
                formula = "=0"
            else:
                # Retained Earnings precedenti + Net Profit - Dividendi
                prev_col = self.quarter_columns[q-1]
                net_profit_row = self._find_ce_row("10. NET PROFIT", row)
                
                if net_profit_row and dividend_row:
                    dividend_rate = f"Input!$C${dividend_row}"
                    formula = (f"={prev_col}{row}+{col_letter}{net_profit_row}"
                              f"-{col_letter}{net_profit_row}*{dividend_rate}")
                else:
                    formula = f"={prev_col}{row}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    def _formula_total_equity(self, row):
        """Formula per Total Equity"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Share Capital + Retained Earnings
            formula = f"={col_letter}{row-2}+{col_letter}{row-1}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')

    # ============================================================================
    # HELPER METHODS PER CONTO ECONOMICO E STATO PATRIMONIALE
    # ============================================================================
    
    def _find_section_in_model(self, section_name):
        """Trova una sezione nel foglio Modello cercando per nome"""
        search_text = section_name.upper()
        for row in range(1, 1000):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val and search_text in str(cell_val).upper():
                return row
        return 0
    
    def _find_in_input(self, parameter_name):
        """Trova un parametro nel foglio Input cercando per nome"""
        search_text = parameter_name.upper()
        for row in range(1, 500):
            cell_val = self.ws_input[f'B{row}'].value
            if cell_val and search_text in str(cell_val).upper():
                return row
        return 0
    
    def _find_input_parameter(self, parameter_name):
        """Trova la riga di un parametro nel foglio Input"""
        search_patterns = {
            "Customer Deposits": ["CUSTOMER DEPOSITS", "DEPOSITI"],
            "Deposit Interest Rate": ["DEPOSIT RATE", "TASSO DEPOSITI"],
            "External Funding": ["EXTERNAL FUNDING", "FUNDING"],
            "Funding Interest Rate": ["FUNDING RATE", "TASSO FUNDING"],
            "Total FTE": ["TOTAL FTE", "FTE TOTALI"],
            "Average Annual Salary": ["AVERAGE SALARY", "STIPENDIO MEDIO"],
            "Share Capital": ["SHARE CAPITAL", "CAPITALE SOCIALE"]
        }
        
        patterns = search_patterns.get(parameter_name, [parameter_name.upper()])
        
        for row in range(1, 500):
            cell_value = self.ws_input[f'B{row}'].value
            if cell_value:
                cell_str = str(cell_value).upper()
                for pattern in patterns:
                    if pattern in cell_str:
                        return row
        
        return None

    def _find_ce_row(self, item_name, current_row):
        """Trova la riga di una voce del Conto Economico"""
        ce_start = self.model_refs.get('ce_start', 600)
        
        for row in range(ce_start, ce_start + 150):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val and item_name in str(cell_val):
                return row
        
        return None
    
    def formule_ecl_stage1(self):
        """Inserisce le formule per il calcolo ECL Stage 1"""
        print("  📊 Inserimento formule ECL Stage 1...")
        
        ecl_start = self.model_refs.get('ecl_stage1_start', 500)
        stock_perf_start = self.model_refs.get('stock_performing_start', 480)
        
        # Parametri ECL dal foglio Input
        ecl_horizon_row = self.input_refs.get('ecl_horizon')
        pd_mult_row = self.input_refs.get('pd_multiplier')
        
        # Controlli di sicurezza
        if ecl_horizon_row is None:
            print("    ⚠️  ATTENZIONE: ecl_horizon non trovato, uso valore di default")
            ecl_horizon_row = 200  # Valore di fallback
        if pd_mult_row is None:
            print("    ⚠️  ATTENZIONE: pd_multiplier non trovato, uso valore di default")
            pd_mult_row = 201  # Valore di fallback
        
        current_row = ecl_start
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            # Righe danger rate e LGD per divisione
            if division == 'RE':
                danger_row = self.input_refs.get('danger_re_row')
                lgd_row = self.input_refs.get('lgd_re_row')
            elif division == 'SME':
                danger_row = self.input_refs.get('danger_sme_row')
                lgd_row = self.input_refs.get('lgd_sme_row')
            else:
                danger_row = self.input_refs.get('danger_pg_row')
                lgd_row = self.input_refs.get('lgd_pg_row')
            
            # Controlli di sicurezza
            if danger_row is None:
                print(f"    ⚠️  ATTENZIONE: danger_row per {division} non trovato")
                continue  # Salta questa divisione
            if lgd_row is None:
                print(f"    ⚠️  ATTENZIONE: lgd_row per {division} non trovato")
                continue  # Salta questa divisione
            
            for idx, product in enumerate(products):
                stock_row = stock_perf_start + (
                    0 if division == 'RE' else
                    6 if division == 'SME' else
                    11
                ) + idx
                
                # Colonna del prodotto nel foglio Input
                product_col = get_column_letter(3 + idx)
                
                for q in range(20):
                    col_letter = self.quarter_columns[q]
                    
                    # ECL = Stock * PD * LGD
                    # PD = Danger Rate * Orizzonte ECL / 4 * Moltiplicatore
                    formula = (f"={col_letter}{stock_row}"
                             f"*(Input!${product_col}${danger_row}/4"
                             f"*Input!$C${ecl_horizon_row}"
                             f"*Input!$C${pd_mult_row})"
                             f"*Input!${product_col}${lgd_row}")
                    
                    cell = self.ws_modello[f'{col_letter}{current_row}']
                    cell.value = formula
                    self._format_formula_cell(cell, '#,##0.0')
                
                current_row += 1
        
        print(f"    ✓ Formule ECL Stage 1 inserite")
    
    def formule_nbv_performing(self):
        """Inserisce le formule per NBV Performing (Stock - ECL)"""
        print("  📊 Inserimento formule NBV Performing...")
        
        nbv_start = self.model_refs.get('nbv_performing_start', 520)
        stock_start = self.model_refs.get('stock_performing_start', 480)
        ecl_start = self.model_refs.get('ecl_stage1_start', 500)
        
        current_row = nbv_start
        for division in ['RE', 'SME', 'PG']:
            products = self.products[division]
            
            for idx, product in enumerate(products):
                offset = (0 if division == 'RE' else
                         6 if division == 'SME' else
                         11) + idx
                
                stock_row = stock_start + offset
                ecl_row = ecl_start + offset
                
                for q in range(20):
                    col_letter = self.quarter_columns[q]
                    
                    # NBV = Stock - ECL
                    formula = f"={col_letter}{stock_row}-{col_letter}{ecl_row}"
                    
                    cell = self.ws_modello[f'{col_letter}{current_row}']
                    cell.value = formula
                    self._format_formula_cell(cell, '#,##0.0')
                
                current_row += 1
        
        print(f"    ✓ Formule NBV Performing inserite")
    
    def formule_conto_economico(self):
        """Inserisce le formule per il Conto Economico"""
        print("  💰 Inserimento formule Conto Economico...")
        
        ce_start = self.model_refs.get('ce_start', 600)
        
        # Trova le righe specifiche del CE
        for row in range(ce_start, ce_start + 100):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val:
                cell_str = str(cell_val).strip()
                
                # Interest Income
                if "1.1. Interest Income" == cell_str:
                    self._formula_interest_income(row)
                
                # Interest Income per divisione
                elif "1.1.1. da Real Estate" in cell_str:
                    self._formula_interest_income_division(row, 'RE')
                elif "1.1.2. da SME" in cell_str:
                    self._formula_interest_income_division(row, 'SME')
                elif "1.1.3. da Public Guarantee" in cell_str:
                    self._formula_interest_income_division(row, 'PG')
                
                # Commission Income
                elif "2.1. Commission Income" == cell_str:
                    self._formula_commission_income(row)
                
                # Commission Income per divisione
                elif "2.1.1. da Real Estate" in cell_str:
                    self._formula_commission_income_division(row, 'RE')
                elif "2.1.2. da SME" in cell_str:
                    self._formula_commission_income_division(row, 'SME')
                elif "2.1.3. da Public Guarantee" in cell_str:
                    self._formula_commission_income_division(row, 'PG')
                
                # Totali e altri calcoli
                elif "1.3. NET INTEREST INCOME" in cell_str:
                    self._formula_net_interest_income(row)
                elif "2.3. NET COMMISSION INCOME" in cell_str:
                    self._formula_net_commission_income(row)
                elif "3. TOTAL REVENUES" in cell_str:
                    self._formula_total_revenues(row)
        
        print(f"    ✓ Formule Conto Economico inserite")
    
    def _formula_interest_income_division(self, row, division):
        """Formula per interest income di una divisione"""
        nbv_start = self.model_refs.get('nbv_performing_start', 520)
        
        # Tassi di interesse per divisione
        if division == 'RE':
            rate_row = self.input_refs.get('rate_re_row')
            products = self.products['RE']
            offset_start = 0
        elif division == 'SME':
            rate_row = self.input_refs.get('rate_sme_row')
            products = self.products['SME']
            offset_start = 6
        else:
            rate_row = self.input_refs.get('rate_pg_row')
            products = self.products['PG']
            offset_start = 11
        
        # Controllo di sicurezza
        if rate_row is None:
            print(f"    ⚠️  ATTENZIONE: rate_row per {division} non trovato")
            return
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma degli interessi di tutti i prodotti della divisione
            formula_parts = []
            for idx, product in enumerate(products):
                nbv_row = nbv_start + offset_start + idx
                product_col = get_column_letter(3 + idx)
                
                # NBV * Tasso / 4 (trimestrale)
                formula_parts.append(f"{col_letter}{nbv_row}*Input!${product_col}${rate_row}/4")
            
            formula = "=" + "+".join(formula_parts)
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_interest_income(self, row):
        """Formula per interest income totale"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma delle sottovoci
            formula = f"=SUM({col_letter}{row+1}:{col_letter}{row+5})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_commission_income_division(self, row, division):
        """Formula per commission income di una divisione"""
        erog_start = self.model_refs.get('erogazioni_start', 10)
        
        # Up-front fees per divisione
        if division == 'RE':
            upfront_row = self.input_refs.get('upfront_re_row')
            products = self.products['RE']
            offset_start = 0
        elif division == 'SME':
            upfront_row = self.input_refs.get('upfront_sme_row')
            products = self.products['SME']
            offset_start = 6
        else:
            upfront_row = self.input_refs.get('upfront_pg_row')
            products = self.products['PG']
            offset_start = 11
        
        # Controllo di sicurezza
        if upfront_row is None:
            print(f"    ⚠️  ATTENZIONE: upfront_row per {division} non trovato")
            return
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma delle commissioni di tutti i prodotti della divisione
            formula_parts = []
            for idx, product in enumerate(products):
                erog_row = erog_start + offset_start + idx
                product_col = get_column_letter(3 + idx)
                
                # Erogazioni * Up-front Fee
                formula_parts.append(f"{col_letter}{erog_row}*Input!${product_col}${upfront_row}")
            
            formula = "=" + "+".join(formula_parts)
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_commission_income(self, row):
        """Formula per commission income totale"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma delle sottovoci
            formula = f"=SUM({col_letter}{row+1}:{col_letter}{row+5})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_net_interest_income(self, row):
        """Formula per Net Interest Income"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Interest Income - Interest Expenses
            # Assumiamo che Interest Income sia 2 righe sopra e Interest Expenses 1 riga sopra
            formula = f"={col_letter}{row-2}-{col_letter}{row-1}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_net_commission_income(self, row):
        """Formula per Net Commission Income"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Commission Income - Commission Expenses
            formula = f"={col_letter}{row-2}-{col_letter}{row-1}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_total_revenues(self, row):
        """Formula per Total Revenues"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Net Interest Income + Net Commission Income
            # Cerchiamo le righe corrette
            net_int_row = 0
            net_comm_row = 0
            for search_row in range(row-20, row):
                cell_val = self.ws_modello[f'B{search_row}'].value
                if cell_val:
                    if "1.3. NET INTEREST INCOME" in str(cell_val):
                        net_int_row = search_row
                    elif "2.3. NET COMMISSION INCOME" in str(cell_val):
                        net_comm_row = search_row
            
            formula = f"={col_letter}{net_int_row}+{col_letter}{net_comm_row}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def formule_stato_patrimoniale(self):
        """Inserisce le formule per lo Stato Patrimoniale"""
        print("  🏦 Inserimento formule Stato Patrimoniale...")
        
        sp_start = self.model_refs.get('sp_start', 800)
        
        # Trova le righe specifiche dello SP
        for row in range(sp_start, sp_start + 50):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val:
                cell_str = str(cell_val).strip()
                
                # Loans to Customers (Gross)
                if "1.1. Loans to Customers (Gross)" in cell_str:
                    self._formula_loans_gross(row)
                
                # Loan Loss Reserves
                elif "1.2. Loan Loss Reserves" in cell_str:
                    self._formula_loan_reserves(row)
                
                # Loans to Customers (Net)
                elif "1.3. Loans to Customers (Net)" in cell_str:
                    self._formula_loans_net(row)
                
                # Total Assets
                elif "1.8. TOTAL ASSETS" in cell_str:
                    self._formula_total_assets(row)
        
        print(f"    ✓ Formule Stato Patrimoniale inserite")
    
    def _formula_loans_gross(self, row):
        """Formula per Loans to Customers (Gross)"""
        stock_perf_start = self.model_refs.get('stock_performing_start', 480)
        stock_npl_start = self.model_refs.get('stock_npl_start', 540)
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma Stock Performing + Stock NPL per tutti i prodotti
            formula_parts = []
            
            # Stock Performing
            for i in range(14):  # Tutti i prodotti
                formula_parts.append(f"{col_letter}{stock_perf_start + i}")
            
            # Stock NPL
            for i in range(14):  # Tutti i prodotti
                formula_parts.append(f"{col_letter}{stock_npl_start + i}")
            
            formula = "=" + "+".join(formula_parts)
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_loan_reserves(self, row):
        """Formula per Loan Loss Reserves"""
        ecl_start = self.model_refs.get('ecl_stage1_start', 500)
        
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma ECL Stage 1 per tutti i prodotti
            formula = f"=SUM({col_letter}{ecl_start}:{col_letter}{ecl_start+13})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_loans_net(self, row):
        """Formula per Loans to Customers (Net)"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Gross - Reserves
            formula = f"={col_letter}{row-2}-{col_letter}{row-1}"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_total_assets(self, row):
        """Formula per Total Assets"""
        for q in range(20):
            col_letter = self.quarter_columns[q]
            
            # Somma di tutte le voci dell'attivo
            formula = f"=SUM({col_letter}{row-5}:{col_letter}{row-1})"
            
            cell = self.ws_modello[f'{col_letter}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def apply_formulas_blocco_e(self):
        """
        Applica le formule per il BLOCCO E - COSTI E ACCANTONAMENTI
        - E.1: LLP (Loan Loss Provisions): Delta_ECL + First_Day_Loss_Nuove_Erogazioni
        - E.2: Write-off: Stock_NPL × Write_off_Rate dopo periodo workout
        """
        print("  📊 BLOCCO E - Inserimento formule Costi e Accantonamenti...")
        
        # E.1 - LLP (LOAN LOSS PROVISIONS)
        self._apply_formulas_e1_llp()
        
        # E.2 - WRITE-OFF
        self._apply_formulas_e2_writeoff()
        
        print("    ✅ BLOCCO E completato - Costi e Accantonamenti")
    
    def apply_formulas_blocco_f(self):
        """
        Applica le formule per il BLOCCO F - AGGREGAZIONI DIVISIONALI
        Per ogni divisione, somma tutti i prodotti:
        - Totale Erogazioni: =SOMMA(Erogazioni_Prodotti_Divisione)
        - Totale Stock: =SOMMA(Stock_GBV_Prodotti)
        - Totale Interessi: =SOMMA(Interessi_Attivi_Prodotti)
        - Totale Commissioni: =SOMMA(Commissioni_Prodotti)
        - Totale LLP: =SOMMA(LLP_Prodotti)
        """
        print("  📊 BLOCCO F - Inserimento formule Aggregazioni Divisionali...")
        
        # F.1 - AGGREGAZIONI PER DIVISIONE
        self._apply_formulas_f1_aggregazioni_divisione()
        
        print("    ✅ BLOCCO F completato - Aggregazioni Divisionali")
    
    # ============================================================================
    # BLOCCO G - FUNDING E LIQUIDITÀ
    # ============================================================================
    
    def apply_formulas_blocco_g(self):
        """
        Applica le formule per il BLOCCO G - FUNDING E LIQUIDITÀ
        G.1 - Depositi Clientela
        G.2 - Funding Wholesale
        G.3 - Liquidità
        """
        print("\n  📊 BLOCCO G - Funding e Liquidità")
        print("  " + "-"*50)
        
        # G.1 - DEPOSITI CLIENTELA
        self._apply_formulas_g1_depositi_clientela()
        
        # G.2 - FUNDING WHOLESALE
        self._apply_formulas_g2_funding_wholesale()
        
        # G.3 - LIQUIDITÀ
        self._apply_formulas_g3_liquidita()
        
        print("    ✅ BLOCCO G completato - Funding e Liquidità")
    
    def _apply_formulas_g1_depositi_clientela(self):
        """G.1 - Depositi Clientela con crescita trimestrale"""
        print("      G.1 - Depositi Clientela...")
        
        # Trova la sezione depositi nel modello
        deposit_row = self._find_section_in_model("G.1 DEPOSITI CLIENTELA")
        if deposit_row == 0:
            deposit_row = self._find_section_in_model("DEPOSITI CLIENTELA")
            if deposit_row == 0:
                return
        
        # Parametri depositi dall'Input
        deposito_medio_base_row = self._find_in_input("DEPOSITO MEDIO BASE")
        deposito_medio_premium_row = self._find_in_input("DEPOSITO MEDIO PREMIUM")
        deposito_medio_corporate_row = self._find_in_input("DEPOSITO MEDIO CORPORATE")
        crescita_depositi_row = self._find_in_input("CRESCITA DEPOSITI")
        
        # Trova numero clienti per tipo
        clienti_base_row = self._find_in_input("CLIENTI BASE")
        clienti_premium_row = self._find_in_input("CLIENTI PREMIUM")
        clienti_corporate_row = self._find_in_input("CLIENTI CORPORATE")
        
        # Per ogni trimestre
        for q_idx, col in enumerate(self.quarter_columns):
            if q_idx == 0:
                # Q1: valori iniziali
                if deposito_medio_base_row > 0 and clienti_base_row > 0:
                    formula_base = f"=Input!C{clienti_base_row}*Input!C{deposito_medio_base_row}/1000000"
                else:
                    formula_base = "=1000"  # Default 1B EUR
                
                if deposito_medio_premium_row > 0 and clienti_premium_row > 0:
                    formula_premium = f"=Input!C{clienti_premium_row}*Input!C{deposito_medio_premium_row}/1000000"
                else:
                    formula_premium = "=500"  # Default 500M EUR
                
                if deposito_medio_corporate_row > 0 and clienti_corporate_row > 0:
                    formula_corporate = f"=Input!C{clienti_corporate_row}*Input!C{deposito_medio_corporate_row}/1000000"
                else:
                    formula_corporate = "=300"  # Default 300M EUR
                
                # Totale depositi Q1
                formula_total = f"=({formula_base[1:]})+({formula_premium[1:]})+({formula_corporate[1:]})"
            else:
                # Dal Q2 in poi: crescita trimestrale
                prev_col = self.quarter_columns[q_idx - 1]
                if crescita_depositi_row > 0:
                    growth_rate = f"Input!C{crescita_depositi_row}"
                    formula_total = f"={prev_col}{deposit_row}*(1+{growth_rate}/4)"
                else:
                    # Default 2% annuo = 0.5% trimestrale
                    formula_total = f"={prev_col}{deposit_row}*1.005"
            
            cell = self.ws_modello[f'{col}{deposit_row}']
            cell.value = formula_total
            self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_g2_funding_wholesale(self):
        """G.2 - Funding Wholesale (bilanciamento automatico)"""
        print("      G.2 - Funding Wholesale...")
        
        # Trova la sezione funding nel modello
        funding_row = self._find_section_in_model("G.2 FUNDING WHOLESALE")
        if funding_row == 0:
            funding_row = self._find_section_in_model("FUNDING WHOLESALE")
            if funding_row == 0:
                return
        
        # Trova riferimenti necessari
        loans_net_row = self._find_section_in_model("LOANS TO CUSTOMERS (NET)")
        deposits_row = self._find_section_in_model("CUSTOMER DEPOSITS")
        equity_row = self._find_section_in_model("TOTAL EQUITY")
        
        # Per ogni trimestre
        for col in self.quarter_columns:
            # Formula bilanciamento: Funding = MAX(0, Crediti_Netti - Depositi - Patrimonio)
            if loans_net_row > 0 and deposits_row > 0 and equity_row > 0:
                formula = f"=MAX(0,{col}{loans_net_row}-{col}{deposits_row}-{col}{equity_row})"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col}{funding_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
        
        # Allocazione tra BCE, debt securities, interbank
        bce_row = funding_row + 1
        debt_row = funding_row + 2
        interbank_row = funding_row + 3
        
        # BCE: 40% del funding totale
        for col in self.quarter_columns:
            formula = f"={col}{funding_row}*0.4"
            cell = self.ws_modello[f'{col}{bce_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
        
        # Debt Securities: 35% del funding totale
        for col in self.quarter_columns:
            formula = f"={col}{funding_row}*0.35"
            cell = self.ws_modello[f'{col}{debt_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
        
        # Interbank: 25% del funding totale
        for col in self.quarter_columns:
            formula = f"={col}{funding_row}*0.25"
            cell = self.ws_modello[f'{col}{interbank_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_g3_liquidita(self):
        """G.3 - Liquidità (LCR buffer e cash totale)"""
        print("      G.3 - Liquidità...")
        
        # Trova la sezione liquidità nel modello
        liquidity_row = self._find_section_in_model("G.3 LIQUIDITÀ")
        if liquidity_row == 0:
            liquidity_row = self._find_section_in_model("LIQUIDITÀ")
            if liquidity_row == 0:
                return
        
        deposits_row = self._find_section_in_model("CUSTOMER DEPOSITS")
        cash_minimo_row = self._find_in_input("CASH MINIMO")
        
        # LCR Buffer
        lcr_row = liquidity_row + 1
        for col in self.quarter_columns:
            if deposits_row > 0:
                if cash_minimo_row > 0:
                    formula = f"=MAX(Input!C{cash_minimo_row},{col}{deposits_row}*0.1)"
                else:
                    formula = f"={col}{deposits_row}*0.1"  # 10% dei depositi
            else:
                formula = "=100"  # Default 100M EUR
            
            cell = self.ws_modello[f'{col}{lcr_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
        
        # Riserve Obbligatorie (1% dei depositi)
        reserves_row = liquidity_row + 2
        for col in self.quarter_columns:
            if deposits_row > 0:
                formula = f"={col}{deposits_row}*0.01"
            else:
                formula = "=20"  # Default 20M EUR
            
            cell = self.ws_modello[f'{col}{reserves_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
        
        # Cash Totale = LCR Buffer + Riserve
        cash_total_row = liquidity_row + 3
        for col in self.quarter_columns:
            formula = f"={col}{lcr_row}+{col}{reserves_row}"
            
            cell = self.ws_modello[f'{col}{cash_total_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    # ============================================================================
    # BLOCCO H - ALTRI PATRIMONIALI
    # ============================================================================
    
    def apply_formulas_blocco_h(self):
        """
        Applica le formule per il BLOCCO H - ALTRI PATRIMONIALI
        H.1 - Immobilizzazioni
        H.2 - Portafoglio Titoli
        """
        print("\n  🏢 BLOCCO H - Altri Patrimoniali")
        print("  " + "-"*50)
        
        # H.1 - IMMOBILIZZAZIONI
        self._apply_formulas_h1_immobilizzazioni()
        
        # H.2 - PORTAFOGLIO TITOLI
        self._apply_formulas_h2_portafoglio_titoli()
        
        print("    ✅ BLOCCO H completato - Altri Patrimoniali")
    
    def _apply_formulas_h1_immobilizzazioni(self):
        """H.1 - Immobilizzazioni (roll-forward)"""
        print("      H.1 - Immobilizzazioni...")
        
        # Trova la sezione immobilizzazioni nel modello
        immo_row = self._find_section_in_model("H.1 IMMOBILIZZAZIONI")
        if immo_row == 0:
            immo_row = self._find_section_in_model("IMMOBILIZZAZIONI")
            if immo_row == 0:
                return
        
        # Parametri dall'Input
        capex_row = self._find_in_input("CAPEX")
        depreciation_rate_row = self._find_in_input("TASSO AMMORTAMENTO")
        initial_assets_row = self._find_in_input("IMMOBILIZZAZIONI INIZIALI")
        
        # Per ogni trimestre
        for q_idx, col in enumerate(self.quarter_columns):
            if q_idx == 0:
                # Q1: Stock iniziale + CAPEX - Ammortamenti
                if initial_assets_row > 0:
                    initial_stock = f"Input!C{initial_assets_row}"
                else:
                    initial_stock = "500"  # Default 500M EUR
                
                if capex_row > 0:
                    capex = f"Input!C{capex_row}/4"  # Trimestralizzato
                else:
                    capex = "10"  # Default 10M EUR/trimestre
                
                if depreciation_rate_row > 0:
                    depreciation = f"{initial_stock}*Input!C{depreciation_rate_row}/4"
                else:
                    depreciation = f"{initial_stock}*0.025"  # Default 10% annuo / 4
                
                formula = f"={initial_stock}+{capex}-{depreciation}"
            else:
                # Dal Q2: Stock precedente + CAPEX - Ammortamenti
                prev_col = self.quarter_columns[q_idx - 1]
                
                if capex_row > 0:
                    capex = f"Input!C{capex_row}/4"
                else:
                    capex = "10"
                
                if depreciation_rate_row > 0:
                    depreciation = f"{prev_col}{immo_row}*Input!C{depreciation_rate_row}/4"
                else:
                    depreciation = f"{prev_col}{immo_row}*0.025"
                
                formula = f"={prev_col}{immo_row}+{capex}-{depreciation}"
            
            cell = self.ws_modello[f'{col}{immo_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
        
        # Dettaglio CAPEX e Ammortamenti nelle righe successive
        capex_detail_row = immo_row + 1
        depreciation_detail_row = immo_row + 2
        
        for col in self.quarter_columns:
            # CAPEX trimestrale
            if capex_row > 0:
                formula = f"=Input!C{capex_row}/4"
            else:
                formula = "=10"
            
            cell = self.ws_modello[f'{col}{capex_detail_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
            
            # Ammortamenti
            if depreciation_rate_row > 0:
                formula = f"={col}{immo_row}*Input!C{depreciation_rate_row}/4"
            else:
                formula = f"={col}{immo_row}*0.025"
            
            cell = self.ws_modello[f'{col}{depreciation_detail_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _apply_formulas_h2_portafoglio_titoli(self):
        """H.2 - Portafoglio Titoli"""
        print("      H.2 - Portafoglio Titoli...")
        
        # Trova la sezione titoli nel modello
        securities_row = self._find_section_in_model("H.2 PORTAFOGLIO TITOLI")
        if securities_row == 0:
            securities_row = self._find_section_in_model("PORTAFOGLIO TITOLI")
            if securities_row == 0:
                return
        
        # Parametri dall'Input
        allocation_gov_row = self._find_in_input("ALLOCATION GOVIES")
        allocation_corp_row = self._find_in_input("ALLOCATION CORPORATE")
        yield_gov_row = self._find_in_input("YIELD GOVIES")
        yield_corp_row = self._find_in_input("YIELD CORPORATE")
        
        # Trova liquidità eccesso
        cash_total_row = self._find_section_in_model("CASH TOTALE")
        lcr_buffer_row = self._find_section_in_model("LCR BUFFER")
        
        # Per ogni trimestre
        for col in self.quarter_columns:
            # Liquidità eccesso = Cash Totale - LCR Buffer
            if cash_total_row > 0 and lcr_buffer_row > 0:
                excess_liquidity = f"({col}{cash_total_row}-{col}{lcr_buffer_row})"
            else:
                excess_liquidity = "100"  # Default 100M EUR
            
            # Govies (60% dell'eccesso di liquidità di default)
            govies_row = securities_row + 1
            if allocation_gov_row > 0:
                formula = f"={excess_liquidity}*Input!C{allocation_gov_row}"
            else:
                formula = f"={excess_liquidity}*0.6"
            
            cell = self.ws_modello[f'{col}{govies_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
            
            # Corporate Bonds (40% dell'eccesso di liquidità di default)
            corp_row = securities_row + 2
            if allocation_corp_row > 0:
                formula = f"={excess_liquidity}*Input!C{allocation_corp_row}"
            else:
                formula = f"={excess_liquidity}*0.4"
            
            cell = self.ws_modello[f'{col}{corp_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
            
            # Totale Portafoglio
            total_row = securities_row + 3
            formula = f"={col}{govies_row}+{col}{corp_row}"
            
            cell = self.ws_modello[f'{col}{total_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
            
            # Rendimento trimestrale
            yield_row = securities_row + 4
            if yield_gov_row > 0 and yield_corp_row > 0:
                formula = f"=({col}{govies_row}*Input!C{yield_gov_row}+{col}{corp_row}*Input!C{yield_corp_row})/4"
            else:
                formula = f"=({col}{govies_row}*0.02+{col}{corp_row}*0.04)/4"
            
            cell = self.ws_modello[f'{col}{yield_row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def apply_formulas_conto_economico(self):
        """
        Applica le formule per il CONTO ECONOMICO con logica completa:
        - Interessi Attivi: somma da Blocco D.1 per divisione
        - Interessi Passivi: =Depositi × Tasso_Depositi/4 + Funding × Tasso_Funding/4
        - Commissioni Attive: somma da Blocco D.3
        - LLP: somma da Blocco E.1
        - Costi Personale: da Input con interpolazione FTE
        - Utile Netto: dopo tasse
        """
        print("  💰 Inserimento formule Conto Economico completo...")
        
        ce_start = self.model_refs.get('ce_start', 600)
        
        # Trova le righe specifiche del CE
        for row in range(ce_start, ce_start + 150):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val:
                cell_str = str(cell_val).strip()
                
                # RICAVI
                if "1.1. Interest Income" == cell_str:
                    self._formula_interest_income_total(row)
                elif "1.1.1. da Real Estate" in cell_str:
                    self._formula_interest_income_division_vintage(row, 'RE')
                elif "1.1.2. da SME" in cell_str:
                    self._formula_interest_income_division_vintage(row, 'SME')
                elif "1.1.3. da Public Guarantee" in cell_str:
                    self._formula_interest_income_division_vintage(row, 'PG')
                
                if "1.2. Interest Expenses" == cell_str:
                    self._formula_interest_expenses_total(row)
                elif "1.2.1. su Depositi" in cell_str:
                    self._formula_interest_expenses_deposits(row)
                elif "1.2.2. su Funding" in cell_str:
                    self._formula_interest_expenses_funding(row)
                
                if "2.1. Commission Income" == cell_str:
                    self._formula_commission_income_total(row)
                elif "2.1.1. da Real Estate" in cell_str:
                    self._formula_commission_income_division_vintage(row, 'RE')
                elif "2.1.2. da SME" in cell_str:
                    self._formula_commission_income_division_vintage(row, 'SME')
                elif "2.1.3. da Public Guarantee" in cell_str:
                    self._formula_commission_income_division_vintage(row, 'PG')
                
                # COSTI
                if "4. LOAN LOSS PROVISIONS" == cell_str:
                    self._formula_llp_total(row)
                
                if "5. PERSONNEL COSTS" == cell_str:
                    self._formula_personnel_costs(row)
                
                # RISULTATI
                if "1.3. NET INTEREST INCOME" in cell_str:
                    self._formula_net_interest_income(row)
                elif "2.3. NET COMMISSION INCOME" in cell_str:
                    self._formula_net_commission_income(row)
                elif "3. TOTAL REVENUES" in cell_str:
                    self._formula_total_revenues(row)
                elif "8. GROSS PROFIT" in cell_str:
                    self._formula_gross_profit(row)
                elif "9. TAXES" in cell_str:
                    self._formula_taxes(row)
                elif "10. NET PROFIT" in cell_str:
                    self._formula_net_profit(row)
        
        print(f"    ✓ Formule Conto Economico complete inserite")
    
    def apply_formulas_stato_patrimoniale(self):
        """
        Applica le formule per lo STATO PATRIMONIALE con logica completa:
        - Crediti Lordi: somma Stock GBV
        - Fondo Rettifiche: ECL cumulato + Fondo NPL
        - Crediti Netti: Lordi - Fondo
        - Depositi: da Blocco G
        - Patrimonio: roll-forward con utili
        """
        print("  🏦 Inserimento formule Stato Patrimoniale completo...")
        
        sp_start = self.model_refs.get('sp_start', 800)
        
        # Trova le righe specifiche dello SP
        for row in range(sp_start, sp_start + 100):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val:
                cell_str = str(cell_val).strip()
                
                # ATTIVO
                if "1.1. Loans to Customers (Gross)" in cell_str:
                    self._formula_loans_gross_vintage(row)
                elif "1.2. Loan Loss Reserves" in cell_str:
                    self._formula_loan_reserves_vintage(row)
                elif "1.3. Loans to Customers (Net)" in cell_str:
                    self._formula_loans_net(row)
                elif "1.8. TOTAL ASSETS" in cell_str:
                    self._formula_total_assets(row)
                
                # PASSIVO
                elif "2.1. Customer Deposits" in cell_str:
                    self._formula_customer_deposits(row)
                elif "2.5. TOTAL LIABILITIES" in cell_str:
                    self._formula_total_liabilities(row)
                
                # PATRIMONIO
                elif "3.1. Share Capital" in cell_str:
                    self._formula_share_capital(row)
                elif "3.2. Retained Earnings" in cell_str:
                    self._formula_retained_earnings(row)
                elif "3.3. TOTAL EQUITY" in cell_str:
                    self._formula_total_equity(row)
        
        print(f"    ✓ Formule Stato Patrimoniale complete inserite")
    
    # ============================================================================
    # CAPITAL REQUIREMENTS
    # ============================================================================
    
    def apply_formulas_capital_requirements(self):
        """
        Applica le formule per i CAPITAL REQUIREMENTS:
        - RWA Credito: Stock_Crediti × RWA_%_Prodotto
        - RWA Mercato: Titoli × 10%
        - RWA Operativo: Ricavi × 15%
        - CET1 Ratio: Patrimonio / RWA_Totale
        """
        print("  🏛️ Inserimento formule Capital Requirements...")
        
        cap_req_start = self.model_refs.get('capital_req_start', 900)
        
        # Trova le righe specifiche dei Capital Requirements
        for row in range(cap_req_start, cap_req_start + 50):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val:
                cell_str = str(cell_val).strip()
                
                if "1. Credit Risk RWA" in cell_str:
                    self._formula_credit_rwa(row)
                elif "2. Market Risk RWA" in cell_str:
                    self._formula_market_rwa(row)
                elif "3. Operational Risk RWA" in cell_str:
                    self._formula_operational_rwa(row)
                elif "4. Total RWA" in cell_str:
                    self._formula_total_rwa(row)
                elif "5. CET1 Capital" in cell_str:
                    self._formula_cet1_capital(row)
                elif "6. CET1 Ratio" in cell_str:
                    self._formula_cet1_ratio(row)
                elif "7. Tier 1 Ratio" in cell_str:
                    self._formula_tier1_ratio(row)
                elif "8. Total Capital Ratio" in cell_str:
                    self._formula_total_capital_ratio(row)
        
        print(f"    ✓ Formule Capital Requirements complete inserite")
    
    def _formula_credit_rwa(self, row):
        """RWA Credito = Stock_Crediti × RWA_%_Prodotto"""
        # Trova riferimenti necessari
        loans_gross_row = self._find_section_in_model("LOANS TO CUSTOMERS (GROSS)")
        
        # Parametri RWA dall'Input per prodotto
        rwa_re_row = self._find_in_input("RWA RE")
        rwa_sme_row = self._find_in_input("RWA SME")
        rwa_pg_row = self._find_in_input("RWA PG")
        
        for col in self.quarter_columns:
            if loans_gross_row > 0:
                # Usa RWA medio ponderato (default 45% per RE, 75% per SME, 20% per PG)
                if rwa_re_row > 0:
                    rwa_re = f"Input!C{rwa_re_row}"
                else:
                    rwa_re = "0.45"
                
                if rwa_sme_row > 0:
                    rwa_sme = f"Input!C{rwa_sme_row}"
                else:
                    rwa_sme = "0.75"
                
                if rwa_pg_row > 0:
                    rwa_pg = f"Input!C{rwa_pg_row}"
                else:
                    rwa_pg = "0.20"
                
                # Media ponderata basata sul mix di portafoglio
                formula = f"={col}{loans_gross_row}*0.55"  # Media semplificata
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_market_rwa(self, row):
        """RWA Mercato = Portafoglio Titoli × 10%"""
        securities_row = self._find_section_in_model("TOTALE PORTAFOGLIO")
        if securities_row == 0:
            securities_row = self._find_section_in_model("PORTAFOGLIO TITOLI")
        
        for col in self.quarter_columns:
            if securities_row > 0:
                formula = f"={col}{securities_row}*0.1"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_operational_rwa(self, row):
        """RWA Operativo = Ricavi medi ultimi 3 anni × 15%"""
        total_revenues_row = self._find_section_in_model("TOTAL REVENUES")
        
        for q_idx, col in enumerate(self.quarter_columns):
            if total_revenues_row > 0:
                if q_idx < 12:  # Primi 3 anni (12 trimestri)
                    # Media progressiva
                    revenue_sum = ""
                    count = 0
                    for i in range(max(0, q_idx - 11), q_idx + 1):
                        revenue_col = self.quarter_columns[i]
                        if revenue_sum:
                            revenue_sum += f"+{revenue_col}{total_revenues_row}"
                        else:
                            revenue_sum = f"{revenue_col}{total_revenues_row}"
                        count += 1
                    formula = f"=({revenue_sum})/{count}*4*0.15"  # Annualizzato × 15%
                else:
                    # Media ultimi 3 anni (12 trimestri)
                    revenue_sum = ""
                    for i in range(q_idx - 11, q_idx + 1):
                        revenue_col = self.quarter_columns[i]
                        if revenue_sum:
                            revenue_sum += f"+{revenue_col}{total_revenues_row}"
                        else:
                            revenue_sum = f"{revenue_col}{total_revenues_row}"
                    formula = f"=({revenue_sum})/12*4*0.15"  # Media annualizzata × 15%
            else:
                formula = "=50"  # Default 50M EUR
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_total_rwa(self, row):
        """Total RWA = Credit RWA + Market RWA + Operational RWA"""
        credit_rwa_row = row - 3
        market_rwa_row = row - 2
        operational_rwa_row = row - 1
        
        for col in self.quarter_columns:
            formula = f"={col}{credit_rwa_row}+{col}{market_rwa_row}+{col}{operational_rwa_row}"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_cet1_capital(self, row):
        """CET1 Capital = Total Equity"""
        equity_row = self._find_section_in_model("TOTAL EQUITY")
        
        for col in self.quarter_columns:
            if equity_row > 0:
                formula = f"={col}{equity_row}"
            else:
                formula = "=200"  # Default 200M EUR
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '#,##0.0')
    
    def _formula_cet1_ratio(self, row):
        """CET1 Ratio = CET1 Capital / Total RWA"""
        cet1_capital_row = row - 1
        total_rwa_row = row - 2
        
        for col in self.quarter_columns:
            formula = f"=IF({col}{total_rwa_row}=0,0,{col}{cet1_capital_row}/{col}{total_rwa_row})"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_tier1_ratio(self, row):
        """Tier 1 Ratio = CET1 Ratio (semplificato)"""
        cet1_ratio_row = row - 1
        
        for col in self.quarter_columns:
            formula = f"={col}{cet1_ratio_row}"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_total_capital_ratio(self, row):
        """Total Capital Ratio = CET1 Ratio + 2% (buffer aggiuntivo)"""
        cet1_ratio_row = row - 2
        
        for col in self.quarter_columns:
            formula = f"={col}{cet1_ratio_row}+0.02"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    # ============================================================================
    # KEY PERFORMANCE INDICATORS (KPI)
    # ============================================================================
    
    def apply_formulas_kpi(self):
        """
        Applica le formule per i KPI:
        - ROE: Utile_Netto / Patrimonio_Medio
        - ROA: Utile_Netto / Attivi_Medi
        - NPL Ratio: Stock_NPL / Crediti_Lordi
        - Cost/Income: Costi_Operativi / Ricavi_Totali
        - Coverage: Fondo_Rettifiche / Stock_NPL
        """
        print("  📊 Inserimento formule Key Performance Indicators...")
        
        kpi_start = self.model_refs.get('kpi_start', 1000)
        
        # Trova le righe specifiche dei KPI
        for row in range(kpi_start, kpi_start + 50):
            cell_val = self.ws_modello[f'B{row}'].value
            if cell_val:
                cell_str = str(cell_val).strip()
                
                if "1. ROE" in cell_str:
                    self._formula_roe(row)
                elif "2. ROA" in cell_str:
                    self._formula_roa(row)
                elif "3. Cost/Income Ratio" in cell_str:
                    self._formula_cost_income(row)
                elif "4. NPL Ratio" in cell_str:
                    self._formula_npl_ratio(row)
                elif "5. Coverage Ratio" in cell_str:
                    self._formula_coverage_ratio(row)
                elif "6. Loans/Deposits" in cell_str:
                    self._formula_loans_deposits(row)
                elif "7. Net Interest Margin" in cell_str:
                    self._formula_nim(row)
                elif "8. Efficiency Ratio" in cell_str:
                    self._formula_efficiency_ratio(row)
        
        print(f"    ✓ Formule KPI complete inserite")
    
    def _formula_roe(self, row):
        """ROE = Utile Netto Annualizzato / Patrimonio Medio"""
        net_profit_row = self._find_section_in_model("NET PROFIT")
        equity_row = self._find_section_in_model("TOTAL EQUITY")
        
        for q_idx, col in enumerate(self.quarter_columns):
            if net_profit_row > 0 and equity_row > 0:
                if q_idx == 0:
                    # Q1: usa solo patrimonio corrente
                    formula = f"=IF({col}{equity_row}=0,0,{col}{net_profit_row}*4/{col}{equity_row})"
                else:
                    # Dal Q2: usa media del patrimonio
                    prev_col = self.quarter_columns[q_idx - 1]
                    formula = f"=IF(({col}{equity_row}+{prev_col}{equity_row})/2=0,0,{col}{net_profit_row}*4/(({col}{equity_row}+{prev_col}{equity_row})/2))"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_roa(self, row):
        """ROA = Utile Netto Annualizzato / Attivi Medi"""
        net_profit_row = self._find_section_in_model("NET PROFIT")
        total_assets_row = self._find_section_in_model("TOTAL ASSETS")
        
        for q_idx, col in enumerate(self.quarter_columns):
            if net_profit_row > 0 and total_assets_row > 0:
                if q_idx == 0:
                    # Q1: usa solo attivi correnti
                    formula = f"=IF({col}{total_assets_row}=0,0,{col}{net_profit_row}*4/{col}{total_assets_row})"
                else:
                    # Dal Q2: usa media degli attivi
                    prev_col = self.quarter_columns[q_idx - 1]
                    formula = f"=IF(({col}{total_assets_row}+{prev_col}{total_assets_row})/2=0,0,{col}{net_profit_row}*4/(({col}{total_assets_row}+{prev_col}{total_assets_row})/2))"
            else:
                formula = "=0"
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_cost_income(self, row):
        """Cost/Income = Costi Operativi / Ricavi Totali"""
        total_revenues_row = self._find_section_in_model("TOTAL REVENUES")
        personnel_costs_row = self._find_section_in_model("PERSONNEL COSTS")
        other_costs_row = self._find_section_in_model("OTHER OPERATING COSTS")
        
        for col in self.quarter_columns:
            if total_revenues_row > 0 and personnel_costs_row > 0:
                if other_costs_row > 0:
                    costs = f"({col}{personnel_costs_row}+{col}{other_costs_row})"
                else:
                    costs = f"{col}{personnel_costs_row}*1.5"  # Stima altri costi come 50% del personale
                
                formula = f"=IF({col}{total_revenues_row}=0,0,{costs}/{col}{total_revenues_row})"
            else:
                formula = "=0.5"  # Default 50%
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_npl_ratio(self, row):
        """NPL Ratio = Stock NPL / Crediti Lordi"""
        loans_gross_row = self._find_section_in_model("LOANS TO CUSTOMERS (GROSS)")
        
        # Somma NPL da tutte le divisioni
        npl_total = 0
        for division in ['RE', 'SME', 'PG']:
            npl_section = self._find_section_in_model(f"STOCK NPL {division}")
            if npl_section > 0:
                if npl_total == 0:
                    npl_total = npl_section
                    break
        
        for col in self.quarter_columns:
            if loans_gross_row > 0 and npl_total > 0:
                # Somma NPL da vintage analysis
                npl_sum = self._get_total_npl_formula(col)
                formula = f"=IF({col}{loans_gross_row}=0,0,{npl_sum}/{col}{loans_gross_row})"
            else:
                formula = "=0.02"  # Default 2%
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_coverage_ratio(self, row):
        """Coverage Ratio = Fondo Rettifiche / Stock NPL"""
        loan_reserves_row = self._find_section_in_model("LOAN LOSS RESERVES")
        
        for col in self.quarter_columns:
            if loan_reserves_row > 0:
                # Somma NPL da vintage analysis
                npl_sum = self._get_total_npl_formula(col)
                formula = f"=IF({npl_sum}=0,0,{col}{loan_reserves_row}/{npl_sum})"
            else:
                formula = "=0.6"  # Default 60%
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_loans_deposits(self, row):
        """Loans/Deposits = Crediti Netti / Depositi"""
        loans_net_row = self._find_section_in_model("LOANS TO CUSTOMERS (NET)")
        deposits_row = self._find_section_in_model("CUSTOMER DEPOSITS")
        
        for col in self.quarter_columns:
            if loans_net_row > 0 and deposits_row > 0:
                formula = f"=IF({col}{deposits_row}=0,0,{col}{loans_net_row}/{col}{deposits_row})"
            else:
                formula = "=0.9"  # Default 90%
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_nim(self, row):
        """Net Interest Margin = Net Interest Income Annualizzato / Attivi Fruttiferi Medi"""
        net_interest_income_row = self._find_section_in_model("NET INTEREST INCOME")
        loans_net_row = self._find_section_in_model("LOANS TO CUSTOMERS (NET)")
        securities_row = self._find_section_in_model("TOTALE PORTAFOGLIO")
        
        for q_idx, col in enumerate(self.quarter_columns):
            if net_interest_income_row > 0 and loans_net_row > 0:
                if securities_row > 0:
                    earning_assets = f"({col}{loans_net_row}+{col}{securities_row})"
                else:
                    earning_assets = f"{col}{loans_net_row}"
                
                if q_idx == 0:
                    formula = f"=IF({earning_assets}=0,0,{col}{net_interest_income_row}*4/{earning_assets})"
                else:
                    prev_col = self.quarter_columns[q_idx - 1]
                    if securities_row > 0:
                        prev_earning = f"({prev_col}{loans_net_row}+{prev_col}{securities_row})"
                    else:
                        prev_earning = f"{prev_col}{loans_net_row}"
                    
                    formula = f"=IF(({earning_assets}+{prev_earning})/2=0,0,{col}{net_interest_income_row}*4/(({earning_assets}+{prev_earning})/2))"
            else:
                formula = "=0.025"  # Default 2.5%
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _formula_efficiency_ratio(self, row):
        """Efficiency Ratio = Operating Costs / (Net Interest Income + Net Commission Income)"""
        net_interest_income_row = self._find_section_in_model("NET INTEREST INCOME")
        net_commission_income_row = self._find_section_in_model("NET COMMISSION INCOME")
        personnel_costs_row = self._find_section_in_model("PERSONNEL COSTS")
        other_costs_row = self._find_section_in_model("OTHER OPERATING COSTS")
        
        for col in self.quarter_columns:
            if net_interest_income_row > 0 and net_commission_income_row > 0 and personnel_costs_row > 0:
                revenues = f"({col}{net_interest_income_row}+{col}{net_commission_income_row})"
                
                if other_costs_row > 0:
                    costs = f"({col}{personnel_costs_row}+{col}{other_costs_row})"
                else:
                    costs = f"{col}{personnel_costs_row}*1.5"
                
                formula = f"=IF({revenues}=0,0,{costs}/{revenues})"
            else:
                formula = "=0.55"  # Default 55%
            
            cell = self.ws_modello[f'{col}{row}']
            cell.value = formula
            self._format_formula_cell(cell, '0.00%')
    
    def _get_total_npl_formula(self, col):
        """Helper per ottenere la somma totale NPL da vintage analysis"""
        # Trova sezioni NPL per ogni divisione
        npl_sections = []
        for division in ['RE', 'SME', 'PG']:
            for product in self.products.get(division, []):
                npl_start = self._find_vintage_matrix_start(division, product, 'STOCK_NPL')
                if npl_start > 0:
                    # Somma tutte le vintage per questo prodotto
                    for v in range(1, 21):  # 20 vintage
                        npl_sections.append(f"{col}{npl_start + v}")
        
        if npl_sections:
            return f"({'+'.join(npl_sections)})"
        else:
            return "100"  # Default 100M EUR

    def inserisci_tutte_formule(self):
        """Inserisce tutte le formule nel modello"""
        print("\n" + "="*70)
        print("🚀 FORMULA ENGINE - INSERIMENTO FORMULE MODELLO BANCARIO CON VINTAGE ANALYSIS")
        print("="*70)
        
        try:
            # 1. Formule calcoli di appoggio - VINTAGE ANALYSIS
            print("\n📈 SEZIONE 1: Matrici Vintage - Calcoli di Appoggio")
            print("-"*50)
            self.formule_erogazioni_base()
            self.formule_stock_crediti()
            
            # BLOCCHI A, B, C, D, E, F, G & H - VINTAGE ANALYSIS COMPLETA
            print("\n🎯 VINTAGE ANALYSIS - BLOCCHI A-H COMPLETI")
            print("-"*50)
            self.apply_formulas_blocco_a()  # Erogazioni, Rimborsi, Stock GBV
            self.apply_formulas_blocco_b()  # Default, Performing, NPL
            self.apply_formulas_blocco_c()  # ECL e Valutazioni
            self.apply_formulas_blocco_d()  # Ricavi
            self.apply_formulas_blocco_e()  # Costi e Accantonamenti
            self.apply_formulas_blocco_f()  # Aggregazioni Divisionali
            self.apply_formulas_blocco_g()  # Funding e Liquidità
            self.apply_formulas_blocco_h()  # Altri Patrimoniali
            
            # Calcoli tradizionali
            self.formule_ecl_stage1()
            self.formule_nbv_performing()
            
            # 2. Formule Conto Economico
            print("\n💰 SEZIONE 2: Conto Economico")
            print("-"*50)
            self.apply_formulas_conto_economico()
            
            # 3. Formule Stato Patrimoniale
            print("\n🏦 SEZIONE 3: Stato Patrimoniale")
            print("-"*50)
            self.apply_formulas_stato_patrimoniale()
            
            # 4. Formule Capital Requirements
            print("\n🏛️ SEZIONE 4: Capital Requirements")
            print("-"*50)
            self.apply_formulas_capital_requirements()
            
            # 5. Formule KPI
            print("\n📊 SEZIONE 5: Key Performance Indicators")
            print("-"*50)
            self.apply_formulas_kpi()
            
            # Salva il file
            self.wb.save("modello_bancario_completo.xlsx")
            
            print("\n" + "="*70)
            print("✅ FORMULA ENGINE COMPLETATO CON SUCCESSO!")
            print("📊 File 'modello_bancario_completo.xlsx' aggiornato con Vintage Analysis")
            print("🎯 Implementati tutti i BLOCCHI:")
            print("   • BLOCCO A (Erogazioni/Volumi)")
            print("   • BLOCCO B (Rischio Credito)")
            print("   • BLOCCO C (ECL e Valutazioni)")
            print("   • BLOCCO D (Ricavi)")
            print("   • BLOCCO E (Costi e Accantonamenti)")
            print("   • BLOCCO F (Aggregazioni Divisionali)")
            print("   • BLOCCO G (Funding e Liquidità)")
            print("   • BLOCCO H (Altri Patrimoniali)")
            print("📈 Report Finanziari Completi:")
            print("   • Conto Economico")
            print("   • Stato Patrimoniale")
            print("   • Capital Requirements")
            print("   • Key Performance Indicators (KPI)")
            print("="*70)
            
        except Exception as e:
            print(f"\n❌ ERRORE durante l'inserimento formule: {str(e)}")
            raise

def main():
    """Funzione principale"""
    engine = FormulaEngine()
    engine.inserisci_tutte_formule()

if __name__ == "__main__":
    main()
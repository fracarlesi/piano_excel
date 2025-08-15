#!/usr/bin/env python3
"""
Script pulito per inserire le formule dei rimborsi
"""

import openpyxl
from openpyxl.utils import get_column_letter

def main():
    # Apri il file
    wb = openpyxl.load_workbook("modello.xlsx")
    ws_calc = wb["Calcoli"]
    
    print("=== INSERIMENTO FORMULE RIMBORSI ===\n")
    
    # Per ogni prodotto
    for prodotto in range(20):
        base_col_calc = 2 + (prodotto * 43)
        col_input = 4 + prodotto
        col_letter_input = get_column_letter(col_input)
        
        # Matrice rimborsi (seconda matrice)
        base_row_rimborsi = 49  # Inizio seconda matrice
        header_row = base_row_rimborsi + 3  # Riga intestazioni
        data_start_row = header_row + 1  # Prima riga dati (53)
        
        if prodotto == 0:
            print(f"Prodotto 1 (Input colonna {col_letter_input}):")
            print("Inserimento formule dinamiche...\n")
        
        # Popola la matrice
        for erog_row in range(40):
            for trim_col in range(40):
                row = data_start_row + erog_row
                col = base_col_calc + trim_col
                
                if trim_col >= erog_row:
                    # Calcola riferimenti
                    erog_row_ref = 7 + erog_row  # Riga nella matrice erogazioni
                    erog_col_ref = base_col_calc + erog_row
                    erog_cell_ref = get_column_letter(erog_col_ref) + str(erog_row_ref)
                    
                    trimestri_passati = trim_col - erog_row
                    
                    # Formula completa che gestisce amortizing e bullet
                    tipo_ref = f'Input!${col_letter_input}$67'
                    maturity_ref = f'Input!${col_letter_input}$68' 
                    pre_amort_ref = f'Input!${col_letter_input}$69'
                    
                    # Formula dinamica
                    formula = (
                        f'=SE({erog_cell_ref}=0;0;'
                        f'SE({tipo_ref}="bullet";'
                        f'SE({trimestri_passati}={maturity_ref}-1;{erog_cell_ref};0);'
                        f'SE({trimestri_passati}<{pre_amort_ref};0;'
                        f'SE({trimestri_passati}<{maturity_ref};'
                        f'{erog_cell_ref}/({maturity_ref}-{pre_amort_ref});0))))'
                    )
                    
                    # Semplifica se siamo oltre i 40 trimestri
                    if trimestri_passati > 40:
                        formula = '=0'
                    
                    cell = ws_calc.cell(row=row, column=col)
                    cell.value = formula
                    
                    # Mostra esempi per il primo prodotto
                    if prodotto == 0 and erog_row == 0 and trim_col < 5:
                        cell_ref = get_column_letter(col) + str(row)
                        print(f"  Cella {cell_ref} (T{trim_col+1}):")
                        print(f"    Trimestri dall'erogazione: {trimestri_passati}")
                        if trim_col < 2:
                            print(f"    Formula: {formula[:80]}...")
                        print()
                else:
                    # Prima dell'erogazione, nessun rimborso
                    cell = ws_calc.cell(row=row, column=col)
                    cell.value = '=0'
    
    print("✅ Formule inserite con successo!\n")
    print("Le formule gestiscono dinamicamente:")
    print("- Prodotti BULLET: rimborso totale alla scadenza")
    print("- Prodotti AMORTIZING: rate costanti")
    print("- Pre-ammortamento: periodo iniziale solo interessi")
    print("\nTutto basato sui parametri in Input (righe 67-69)")
    
    # Salva il file
    wb.save("modello.xlsx")
    print("\nFile salvato!")

if __name__ == "__main__":
    main()
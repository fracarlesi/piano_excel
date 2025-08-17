#!/usr/bin/env python3
"""
🏦 CALCOLATORE AVANZATO PER PIANO INDUSTRIALE BANCARIO
Alternativa a Python in Excel - Script esterno per calcoli complessi

Uso:
    python excel_calculator.py
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime

class PianoIndustrialeCalculator:
    """Calcolatore per il piano industriale bancario"""
    
    def __init__(self, file_path="modello.xlsx"):
        self.file_path = file_path
        self.workbook = None
        self.backup_created = False
        
    def load_workbook(self):
        """Carica il workbook Excel"""
        try:
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"File non trovato: {self.file_path}")
            
            self.workbook = openpyxl.load_workbook(self.file_path)
            print(f"✅ Workbook caricato: {self.file_path}")
            return True
            
        except PermissionError:
            print("❌ ERRORE: Il file Excel è aperto in Excel!")
            print("   Chiudi il file Excel e riprova.")
            return False
        except Exception as e:
            print(f"❌ Errore nel caricamento: {e}")
            return False
    
    def create_backup(self):
        """Crea backup del file originale"""
        if not self.backup_created:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"modello_backup_{timestamp}.xlsx"
            
            try:
                import shutil
                shutil.copy2(self.file_path, backup_path)
                print(f"✅ Backup creato: {backup_path}")
                self.backup_created = True
            except Exception as e:
                print(f"⚠️ Warning: Impossibile creare backup: {e}")
    
    def read_input_parameters(self):
        """Legge i parametri dal foglio Input"""
        if not self.workbook:
            return {}
        
        input_sheet = self.workbook['Input']
        parameters = {}
        
        # Leggo parametri dalla struttura attuale
        # Euribor e altri parametri macro (righe 8+)
        for row in range(8, 20):  # Assumo che i parametri siano nelle prime righe
            param_name = input_sheet.cell(row=row, column=2).value  # Colonna B
            if param_name:
                # Leggo i valori per Y1, Y2, etc.
                values = {}
                for col in range(4, 10):  # Colonne D-I per anni
                    value = input_sheet.cell(row=row, column=col).value
                    if value is not None:
                        year_label = input_sheet.cell(row=7, column=col).value or f"Y{col-3}"
                        values[year_label] = value
                
                parameters[param_name] = values
        
        return parameters
    
    def calculate_advanced_metrics(self, parameters):
        """Calcola metriche avanzate per il piano industriale"""
        results = {}
        
        try:
            # Esempio: Calcolo NPV (Net Present Value)
            if 'Euribor 3M' in parameters:
                euribor_values = list(parameters['Euribor 3M'].values())
                if euribor_values:
                    # Calcolo media Euribor
                    avg_euribor = sum(euribor_values) / len(euribor_values)
                    results['avg_euribor'] = avg_euribor
                    
                    # Calcolo discount rate (Euribor + spread)
                    discount_rate = avg_euribor + 0.02  # Assumo 2% di spread
                    results['discount_rate'] = discount_rate
            
            # Aggiungi qui altri calcoli specifici per il tuo piano industriale
            # ROE, ROA, Cost of Risk, etc.
            
        except Exception as e:
            print(f"⚠️ Errore nei calcoli: {e}")
        
        return results
    
    def update_calculations_sheet(self, results):
        """Aggiorna il foglio Calcoli con i risultati"""
        if not self.workbook or not results:
            return False
        
        calc_sheet = self.workbook['Calcoli']
        
        try:
            # Trova una zona libera per inserire i risultati
            # Aggiungo i risultati in una sezione dedicata
            start_row = calc_sheet.max_row + 2
            
            calc_sheet.cell(row=start_row, column=1).value = "CALCOLI PYTHON"
            calc_sheet.cell(row=start_row + 1, column=1).value = "Risultato"
            calc_sheet.cell(row=start_row + 1, column=2).value = "Valore"
            
            row = start_row + 2
            for key, value in results.items():
                calc_sheet.cell(row=row, column=1).value = key
                calc_sheet.cell(row=row, column=2).value = value
                row += 1
            
            print(f"✅ Risultati aggiunti al foglio Calcoli (riga {start_row})")
            return True
            
        except Exception as e:
            print(f"❌ Errore nell'aggiornamento: {e}")
            return False
    
    def save_workbook(self, output_path=None):
        """Salva il workbook"""
        if not self.workbook:
            return False
        
        save_path = output_path or self.file_path
        
        try:
            self.workbook.save(save_path)
            print(f"✅ File salvato: {save_path}")
            return True
        except PermissionError:
            print("❌ ERRORE: Impossibile salvare - file Excel aperto!")
            return False
        except Exception as e:
            print(f"❌ Errore nel salvataggio: {e}")
            return False
    
    def run_analysis(self):
        """Esegue l'analisi completa"""
        print("🏦 AVVIO ANALISI PIANO INDUSTRIALE")
        print("=" * 50)
        
        # Carica workbook
        if not self.load_workbook():
            return False
        
        # Crea backup
        self.create_backup()
        
        # Legge parametri
        print("\n📊 Lettura parametri...")
        parameters = self.read_input_parameters()
        print(f"   Parametri trovati: {len(parameters)}")
        
        for param, values in parameters.items():
            print(f"   - {param}: {values}")
        
        # Calcola metriche
        print("\n🧮 Calcolo metriche avanzate...")
        results = self.calculate_advanced_metrics(parameters)
        print(f"   Risultati calcolati: {len(results)}")
        
        for key, value in results.items():
            print(f"   - {key}: {value:.4f}")
        
        # Aggiorna foglio
        print("\n💾 Aggiornamento foglio Calcoli...")
        if self.update_calculations_sheet(results):
            # Salva (commentato per sicurezza - rimuovi commento per salvare)
            # self.save_workbook()
            print("⚠️  IMPORTANTE: Salvataggio disabilitato per sicurezza")
            print("   Rimuovi commento da save_workbook() per salvare")
        
        print("\n✅ Analisi completata!")
        return True

def main():
    """Funzione principale"""
    calculator = PianoIndustrialeCalculator()
    success = calculator.run_analysis()
    
    if success:
        print("\n🎯 PROSSIMI PASSI:")
        print("1. Verifica i risultati nel foglio Calcoli")
        print("2. Personalizza i calcoli modificando calculate_advanced_metrics()")
        print("3. Abilita il salvataggio rimuovendo il commento da save_workbook()")
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())
#!/usr/bin/env python3
"""
Analisi dettagliata del foglio 'Calcoli' del file modello.xlsx
Focus sull'area intorno alla riga 99 e matrice Stock GBV
"""

import openpyxl
from openpyxl.utils import get_column_letter

def analizza_foglio_calcoli():
    try:
        # Carica il file Excel
        workbook = openpyxl.load_workbook('/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello.xlsx', data_only=False)
        
        if 'Calcoli' not in workbook.sheetnames:
            print("ERRORE: Il foglio 'Calcoli' non esiste nel file")
            print(f"Fogli disponibili: {workbook.sheetnames}")
            return
        
        sheet = workbook['Calcoli']
        print("=" * 80)
        print("ANALISI DETTAGLIATA FOGLIO 'CALCOLI' - AREA RIGA 99")
        print("=" * 80)
        
        # 1. Analisi area intorno alla riga 99 (da riga 95 a 105)
        print("\n1. AREA INTORNO ALLA RIGA 99 (righe 95-105)")
        print("-" * 50)
        
        for row_num in range(95, 106):
            row_data = []
            # Esamina dalle colonne A alla AO (fino alla colonna 41)
            for col_num in range(1, 42):  # A=1, AO=41
                cell = sheet.cell(row=row_num, column=col_num)
                col_letter = get_column_letter(col_num)
                
                # Mostra valori e formule significativi
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.strip():
                        row_data.append(f"{col_letter}{row_num}: '{cell.value}'")
                    elif isinstance(cell.value, (int, float)) and cell.value != 0:
                        row_data.append(f"{col_letter}{row_num}: {cell.value}")
                    elif str(cell.value).startswith('='):
                        row_data.append(f"{col_letter}{row_num}: {cell.value}")
            
            if row_data:
                print(f"Riga {row_num}: {' | '.join(row_data[:10])}")  # Limita output
                if len(row_data) > 10:
                    print(f"         ... e altri {len(row_data)-10} valori")
        
        # 2. Intestazioni colonne da B a AO
        print("\n2. INTESTAZIONI COLONNE (da B a AO)")
        print("-" * 50)
        
        # Cerca intestazioni nelle prime 10 righe
        for header_row in range(1, 11):
            headers = []
            for col_num in range(2, 42):  # B=2, AO=41
                cell = sheet.cell(row=header_row, column=col_num)
                col_letter = get_column_letter(col_num)
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                    headers.append(f"{col_letter}: '{cell.value.strip()}'")
            
            if headers:
                print(f"Riga {header_row}: {' | '.join(headers[:8])}")
                if len(headers) > 8:
                    print(f"          ... e altri {len(headers)-8} headers")
        
        # 3. Ricerca specifica per "Stock GBV", "Erogazioni", "Rimborsi"
        print("\n3. RICERCA TERMINI CHIAVE")
        print("-" * 50)
        
        keywords = ["Stock", "GBV", "Erogazioni", "Rimborsi", "Erogazione", "Rimborso"]
        keyword_positions = {}
        
        for row_num in range(1, 150):  # Cerca in un'area più ampia
            for col_num in range(1, 42):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value and isinstance(cell.value, str):
                    cell_text = cell.value.strip().lower()
                    for keyword in keywords:
                        if keyword.lower() in cell_text:
                            col_letter = get_column_letter(col_num)
                            if keyword not in keyword_positions:
                                keyword_positions[keyword] = []
                            keyword_positions[keyword].append(f"{col_letter}{row_num}: '{cell.value}'")
        
        for keyword, positions in keyword_positions.items():
            if positions:
                print(f"\n'{keyword}' trovato in:")
                for pos in positions[:5]:  # Mostra prime 5 occorrenze
                    print(f"  - {pos}")
                if len(positions) > 5:
                    print(f"  ... e altre {len(positions)-5} occorrenze")
        
        # 4. Analisi formule nella zona riga 99
        print("\n4. FORMULE PRESENTI NELL'AREA RIGA 99")
        print("-" * 50)
        
        formule_trovate = []
        for row_num in range(95, 105):
            for col_num in range(2, 42):  # B a AO
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    col_letter = get_column_letter(col_num)
                    formule_trovate.append(f"{col_letter}{row_num}: {cell.value}")
        
        if formule_trovate:
            for formula in formule_trovate[:10]:  # Mostra prime 10 formule
                print(f"  {formula}")
            if len(formule_trovate) > 10:
                print(f"  ... e altre {len(formule_trovate)-10} formule")
        else:
            print("  Nessuna formula trovata nell'area specificata")
        
        # 5. Struttura temporale - cerca date/anni nelle intestazioni
        print("\n5. STRUTTURA TEMPORALE")
        print("-" * 50)
        
        date_pattern = []
        for row_num in range(1, 15):  # Prime righe per intestazioni temporali
            for col_num in range(2, 42):
                cell = sheet.cell(row=row_num, column=col_num)
                col_letter = get_column_letter(col_num)
                if cell.value:
                    cell_str = str(cell.value).strip()
                    # Cerca pattern di date/anni
                    if any(pattern in cell_str for pattern in ['2024', '2025', '2026', '2027', '2028', 'gen', 'feb', 'mar', 'apr', 'mag', 'giu', 'lug', 'ago', 'set', 'ott', 'nov', 'dic']):
                        date_pattern.append(f"{col_letter}{row_num}: '{cell_str}'")
        
        if date_pattern:
            print("Pattern temporali trovati:")
            for pattern in date_pattern[:15]:
                print(f"  {pattern}")
        else:
            print("Nessun pattern temporale identificato nelle prime righe")
        
        # 6. Riepilogo struttura dati
        print("\n6. RIEPILOGO STRUTTURA DATI")
        print("-" * 50)
        
        # Conta righe e colonne con dati
        max_row_with_data = 0
        max_col_with_data = 0
        
        for row_num in range(1, 200):
            for col_num in range(1, 50):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    max_row_with_data = max(max_row_with_data, row_num)
                    max_col_with_data = max(max_col_with_data, col_num)
        
        print(f"Ultima riga con dati: {max_row_with_data}")
        print(f"Ultima colonna con dati: {get_column_letter(max_col_with_data)}")
        
        workbook.close()
        
    except PermissionError:
        print("ERRORE: Il file Excel è aperto in un'altra applicazione.")
        print("Chiudi il file Excel e riprova.")
    except FileNotFoundError:
        print("ERRORE: File 'modello.xlsx' non trovato nella directory corrente.")
    except Exception as e:
        print(f"ERRORE: {str(e)}")

if __name__ == "__main__":
    analizza_foglio_calcoli()
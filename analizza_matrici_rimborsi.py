#!/usr/bin/env python3
"""
Script per analizzare le matrici rimborsi nel foglio Calcoli
e identificare le posizioni esatte di ogni prodotto
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys

def analizza_matrici_rimborsi():
    """Analizza le posizioni delle matrici rimborsi per ogni prodotto"""
    
    try:
        # Carica il workbook
        wb = openpyxl.load_workbook('modello.xlsx', data_only=False)
        
        # Verifica che esista il foglio Calcoli
        if 'Calcoli' not in wb.sheetnames:
            print("ERRORE: Foglio 'Calcoli' non trovato!")
            print(f"Fogli disponibili: {wb.sheetnames}")
            return
        
        ws_calcoli = wb['Calcoli']
        
        print("🔍 ANALISI MATRICI RIMBORSI - FOGLIO CALCOLI")
        print("=" * 60)
        
        # Scansiona il foglio per trovare le intestazioni dei prodotti
        matrici_prodotti = []
        
        # Cerca dalla riga 40 alla 200 per trovare le matrici
        for row in range(40, 201):
            for col in range(1, 10):  # Colonne A-I
                cell = ws_calcoli.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str):
                    # Cerca pattern "Prodotto X" o "Rimborsi Prodotto X"
                    if "prodotto" in cell.value.lower() and any(char.isdigit() for char in cell.value):
                        # Estrai il numero del prodotto
                        import re
                        match = re.search(r'prodotto\s*(\d+)', cell.value.lower())
                        if match:
                            prodotto_num = int(match.group(1))
                            matrici_prodotti.append({
                                'prodotto': prodotto_num,
                                'cella_intestazione': f"{get_column_letter(col)}{row}",
                                'riga': row,
                                'colonna': col,
                                'testo': cell.value
                            })
        
        # Ordina per numero prodotto
        matrici_prodotti.sort(key=lambda x: x['prodotto'])
        
        print(f"📊 TROVATE {len(matrici_prodotti)} MATRICI PRODOTTO:")
        print()
        
        for matrice in matrici_prodotti:
            print(f"🎯 PRODOTTO {matrice['prodotto']}:")
            print(f"   Intestazione in cella: {matrice['cella_intestazione']}")
            print(f"   Testo: '{matrice['testo']}'")
            
            # Cerca la prima cella della matrice (di solito qualche riga sotto)
            prima_cella_matrice = None
            formula_attuale = None
            
            # Cerca nelle 10 righe successive all'intestazione
            for offset in range(1, 11):
                test_row = matrice['riga'] + offset
                test_cell = ws_calcoli.cell(row=test_row, column=matrice['colonna'])
                
                # Se troviamo una formula o un valore numerico, questa potrebbe essere la prima cella
                if test_cell.value is not None:
                    if hasattr(test_cell, 'value') and (
                        isinstance(test_cell.value, (int, float)) or 
                        (isinstance(test_cell.value, str) and test_cell.value.startswith('='))
                    ):
                        prima_cella_matrice = f"{get_column_letter(matrice['colonna'])}{test_row}"
                        formula_attuale = test_cell.value
                        break
            
            if prima_cella_matrice:
                print(f"   Prima cella matrice: {prima_cella_matrice}")
                print(f"   Formula/valore attuale: {formula_attuale}")
            else:
                print(f"   Prima cella matrice: NON IDENTIFICATA (verificare manualmente)")
            
            print()
        
        # Analizza anche il foglio Input per vedere la struttura
        if 'Input' in wb.sheetnames:
            ws_input = wb['Input']
            print("📋 ANALISI FOGLIO INPUT:")
            print("=" * 40)
            
            # Cerca i dati di input intorno alla riga 67
            for row in range(60, 80):
                print(f"Riga {row}:")
                for col in range(4, 11):  # Colonne D-J
                    cell = ws_input.cell(row=row, column=col)
                    col_letter = get_column_letter(col)
                    print(f"   {col_letter}{row}: {cell.value}")
                print()
        
        # Genera le formule corrette
        print("🔧 FORMULE CORRETTE SUGGERITE:")
        print("=" * 50)
        
        for i, matrice in enumerate(matrici_prodotti):
            prodotto_num = matrice['prodotto']
            col_input = get_column_letter(4 + i)  # D=4, E=5, F=6, etc.
            
            print(f"📌 PRODOTTO {prodotto_num}:")
            print(f"   Colonna Input di riferimento: {col_input}")
            
            if matrice.get('prima_cella_matrice'):
                # Formula base suggerita (da adattare in base alla logica specifica)
                formula_suggerita = f"=Input.{col_input}67"  # Assumendo riga 67 come punto di partenza
                print(f"   Cella da modificare: {matrice['prima_cella_matrice']}")
                print(f"   Formula suggerita: {formula_suggerita}")
            
            print()
        
        # Template formula parametrizzata
        print("🎯 TEMPLATE FORMULA PARAMETRIZZATA:")
        print("=" * 45)
        print("Per facilitare la modifica di tutte le matrici, usa questo template:")
        print()
        print("=Input.[COLONNA_PRODOTTO][RIGA_RIFERIMENTO]")
        print()
        print("Dove:")
        print("- [COLONNA_PRODOTTO]: D per Prodotto 1, E per Prodotto 2, F per Prodotto 3, etc.")
        print("- [RIGA_RIFERIMENTO]: La riga specifica nel foglio Input (es. 67, 68, 69)")
        print()
        print("Esempi:")
        print("- Prodotto 1: =Input.D67")
        print("- Prodotto 2: =Input.E67") 
        print("- Prodotto 3: =Input.F67")
        print("- Prodotto 4: =Input.G67")
        print("- etc.")
        
    except FileNotFoundError:
        print("ERRORE: File 'modello.xlsx' non trovato nella directory corrente!")
    except Exception as e:
        print(f"ERRORE durante l'analisi: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analizza_matrici_rimborsi()
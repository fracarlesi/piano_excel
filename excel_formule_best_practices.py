#!/usr/bin/env python3
"""
🏦 BEST PRACTICES PER FORMULE EXCEL CON OPENPYXL
Script definitivo per evitare problemi con i riferimenti alle celle
"""

import openpyxl
from typing import List, Dict, Any
import re

class ExcelFormulaManager:
    """Gestore per formule Excel con best practices integrate"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        
    def __enter__(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.save(self.file_path)
            self.workbook.close()
    
    def write_formula(self, sheet_name: str, cell: str, formula: str) -> bool:
        """
        Scrive una formula Excel utilizzando le best practices
        
        Args:
            sheet_name: Nome del foglio
            cell: Indirizzo della cella (es. "B53")
            formula: Formula Excel in formato corretto
            
        Returns:
            bool: True se successo, False se errore
        """
        try:
            # Validazioni
            if not formula.startswith('='):
                raise ValueError("La formula deve iniziare con =")
            
            # Verifica sintassi italiana (punto e virgola)
            if ',' in formula and ';' not in formula:
                print(f"⚠️  ATTENZIONE: La formula contiene virgole. In Excel italiano usa ';'")
            
            # Ottieni o crea il foglio
            if sheet_name not in self.workbook.sheetnames:
                ws = self.workbook.create_sheet(sheet_name)
                print(f"✅ Creato foglio: {sheet_name}")
            else:
                ws = self.workbook[sheet_name]
            
            # Scrivi la formula
            ws[cell] = formula
            
            # Verifica immediata
            saved_formula = ws[cell].value
            dollar_count_original = formula.count('$')
            dollar_count_saved = saved_formula.count('$') if saved_formula else 0
            
            if dollar_count_original != dollar_count_saved:
                print(f"❌ ERRORE: Riferimenti $ non corrispondenti!")
                print(f"   Originale: {dollar_count_original} caratteri $")
                print(f"   Salvata: {dollar_count_saved} caratteri $")
                return False
            
            print(f"✅ Formula salvata in {sheet_name}!{cell}")
            print(f"   {formula}")
            
            return True
            
        except Exception as e:
            print(f"❌ Errore scrivendo formula in {sheet_name}!{cell}: {e}")
            return False
    
    def write_banking_formulas(self) -> Dict[str, bool]:
        """
        Scrive un set di formule comuni per piani industriali bancari
        """
        
        formulas = [
            # Formule per NPL e Recovery
            {
                'sheet': 'Calcoli',
                'cell': 'B10',
                'formula': '=SE(Input!$D$67="amortizing";B7/Input!$D$68;0)',
                'desc': 'Calcolo ammortamento condizionale'
            },
            {
                'sheet': 'Calcoli', 
                'cell': 'B11',
                'formula': '=Input!$E$20*(1-Input!$F$20)',
                'desc': 'NPL con recovery rate'
            },
            {
                'sheet': 'Calcoli',
                'cell': 'B12', 
                'formula': '=SE(Input!$G$15="linear";B11/Input!$H$15;B11*POTENZA(1+Input!$I$15;1/Input!$H$15)-B11)',
                'desc': 'Ammortamento lineare vs composto'
            },
            {
                'sheet': 'Calcoli',
                'cell': 'B13',
                'formula': '=CERCA.VERT(A13;Input!$A$50:$D$100;4;FALSO)',
                'desc': 'Lookup parametri'
            },
            {
                'sheet': 'Calcoli',
                'cell': 'B14',
                'formula': '=SOMMA.SE(Input!$B$10:$B$50;">="&Input!$E$5;Input!$C$10:$C$50)',
                'desc': 'Somma condizionale con criterio dinamico'
            },
            {
                'sheet': 'Calcoli',
                'cell': 'B15',
                'formula': '=SE(E(Input!$A$1>0;Input!$B$1<100;Input!$C$1="Si");Input!$A$1*Input!$B$1;0)',
                'desc': 'Formula complessa con funzione E()'
            }
        ]
        
        results = {}
        
        print("🏦 SCRIVENDO FORMULE BANCARIE...")
        print("=" * 35)
        
        for formula_info in formulas:
            print(f"\n📝 {formula_info['desc']}")
            
            success = self.write_formula(
                formula_info['sheet'],
                formula_info['cell'], 
                formula_info['formula']
            )
            
            key = f"{formula_info['sheet']}!{formula_info['cell']}"
            results[key] = success
            
        return results

def validate_formula_syntax(formula: str) -> Dict[str, Any]:
    """
    Valida la sintassi di una formula Excel
    """
    
    validation = {
        'valid': True,
        'errors': [],
        'warnings': [],
        'info': {}
    }
    
    # Check basic structure
    if not formula.startswith('='):
        validation['valid'] = False
        validation['errors'].append("Formula deve iniziare con =")
    
    # Check for Italian syntax
    if ',' in formula and ';' not in formula:
        validation['warnings'].append("Usa ';' invece di ',' per Excel italiano")
    
    # Count absolute references
    dollar_count = formula.count('$')
    validation['info']['absolute_refs'] = dollar_count // 2
    
    # Check quote consistency  
    single_quotes = formula.count("'")
    double_quotes = formula.count('"')
    validation['info']['single_quotes'] = single_quotes
    validation['info']['double_quotes'] = double_quotes
    
    # Pattern analysis
    sheet_refs = re.findall(r'[A-Za-z_][A-Za-z0-9_]*!', formula)
    validation['info']['sheet_references'] = len(sheet_refs)
    
    return validation

def demo_completo():
    """
    Demo completo delle best practices
    """
    
    print("🎯 DEMO COMPLETO FORMULE EXCEL")
    print("=" * 35)
    
    file_path = 'modello.xlsx'
    
    # Test sintassi diverse
    test_formulas = [
        {
            'name': 'FORMULA CORRETTA',
            'formula': '=SE(Input!$D$67="amortizing";B7/Input!$D$68;0)',
            'should_work': True
        },
        {
            'name': 'FORMULA CON VIRGOLE (warning)',
            'formula': '=SE(Input!$D$67="amortizing",B7/Input!$D$68,0)',
            'should_work': True
        },
        {
            'name': 'FORMULA SENZA =',
            'formula': 'SE(Input!$D$67="amortizing";B7/Input!$D$68;0)',
            'should_work': False
        }
    ]
    
    print("\n🧪 TEST VALIDAZIONE SINTASSI:")
    print("-" * 30)
    
    for test in test_formulas:
        print(f"\n🔍 {test['name']}:")
        print(f"   Formula: {test['formula']}")
        
        validation = validate_formula_syntax(test['formula'])
        
        if validation['valid']:
            print(f"   ✅ Sintassi valida")
        else:
            print(f"   ❌ Errori: {'; '.join(validation['errors'])}")
        
        if validation['warnings']:
            print(f"   ⚠️  Warning: {'; '.join(validation['warnings'])}")
        
        info = validation['info']
        print(f"   📊 Ref. assoluti: {info['absolute_refs']}, Sheet refs: {info['sheet_references']}")
    
    # Test scrittura completa
    print(f"\n💾 TEST SCRITTURA COMPLETA:")
    print("-" * 25)
    
    with ExcelFormulaManager(file_path) as manager:
        results = manager.write_banking_formulas()
        
        success_count = sum(1 for success in results.values() if success)
        total_count = len(results)
        
        print(f"\n📊 RISULTATI FINALI:")
        print(f"   ✅ Successi: {success_count}/{total_count}")
        
        if success_count == total_count:
            print(f"   🎉 TUTTE LE FORMULE SALVATE CORRETTAMENTE!")
        else:
            print(f"   ⚠️  Alcune formule hanno avuto problemi")

# Template per formule comuni
TEMPLATES = {
    'conditional_sum': '=SE(Input!${col}${row}="{condition}";SOMMA(${range});0)',
    'vlookup_absolute': '=CERCA.VERT(A{row};Input!$A$1:$Z$100;{col_index};FALSO)',
    'recovery_calculation': '=Input!${gbv_cell}*(1-Input!${recovery_cell})', 
    'amortization': '=SE(Input!${type_cell}="linear";{amount}/Input!${periods_cell};{amount}*POTENZA(1+Input!${rate_cell};1/Input!${periods_cell})-{amount})',
    'cumulative_conditional': '=SE(Input!${condition_cell}="Si";SOMMA($B${start_row}:B{current_row});0)'
}

if __name__ == "__main__":
    demo_completo()
    
    print(f"\n📚 TEMPLATE DISPONIBILI:")
    print("-" * 25)
    for name, template in TEMPLATES.items():
        print(f"   {name}: {template}")
    
    print(f"\n🎯 RIEPILOGO BEST PRACTICES:")
    print("=" * 30)
    print("✅ USA virgolette SINGOLE per delimitare la formula")
    print("✅ USA virgolette DOPPIE per le stringhe nella formula")  
    print("✅ USA ';' come separatore (Excel italiano)")
    print("✅ USA $col$row per riferimenti assoluti")
    print("✅ Valida sempre le formule prima di salvare")
    print("✅ Usa context manager per gestire il file")
    print("❌ NON usare virgolette doppie esterne") 
    print("❌ NON fare escape manuale dei $")
    print("❌ NON usare ',' come separatore in Excel italiano")
#!/usr/bin/env python3
"""
Analisi concisa e strutturata del file Excel modello.xlsx - Sheet Input e Calcoli
"""

import openpyxl
from collections import defaultdict
import re

def analyze_excel_structure():
    """Analizza la struttura del file Excel e produce un report conciso"""
    
    file_path = "/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello.xlsx"
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    
    # Analizza i due sheet principali
    input_sheet = workbook["Input"]
    calcoli_sheet = workbook["Calcoli"]
    
    print("=" * 80)
    print("📊 ANALISI STRUTTURALE MODELLO EXCEL - PIANO INDUSTRIALE BANCARIO")
    print("=" * 80)
    
    # ANALISI SHEET INPUT
    print("\n🔹 SHEET 'INPUT'")
    print("-" * 50)
    
    input_sections = analyze_input_structure(input_sheet)
    for section, details in input_sections.items():
        print(f"\n📋 {section}")
        print(f"   Range: {details['range']}")
        print(f"   Celle: {details['cells_count']}")
        if details['description']:
            print(f"   Contenuto: {details['description']}")
        if details['key_cells']:
            print(f"   Celle chiave: {', '.join(details['key_cells'][:5])}")
    
    # ANALISI SHEET CALCOLI  
    print("\n\n🔹 SHEET 'CALCOLI'")
    print("-" * 50)
    
    calcoli_sections = analyze_calcoli_structure(calcoli_sheet)
    for section, details in calcoli_sections.items():
        print(f"\n🔢 {section}")
        print(f"   Range: {details['range']}")
        print(f"   Celle: {details['cells_count']}")
        if details['description']:
            print(f"   Contenuto: {details['description']}")
        if details['formulas_count'] > 0:
            print(f"   Formule: {details['formulas_count']}")
        if details['key_cells']:
            print(f"   Celle chiave: {', '.join(details['key_cells'][:5])}")
    
    # MAPPA DIPENDENZE PRINCIPALI
    print("\n\n🔗 DIPENDENZE PRINCIPALI TRA SHEET")
    print("-" * 50)
    
    dependencies = find_cross_sheet_dependencies(calcoli_sheet)
    if dependencies:
        for calc_cell, input_refs in dependencies.items():
            print(f"   {calc_cell} ← {', '.join(input_refs)}")
    else:
        print("   Nessuna dipendenza cross-sheet identificata nelle prime righe")
    
    workbook.close()


def analyze_input_structure(sheet):
    """Analizza la struttura del sheet Input identificando le sezioni logiche"""
    
    sections = {}
    current_section = None
    section_start = None
    
    for row in range(1, min(sheet.max_row + 1, 250)):  # Limita l'analisi per performance
        cell_b = sheet[f"B{row}"]
        
        # Identifica intestazioni di sezione (celle che iniziano con numeri tipo "1.1", "1.2")
        if cell_b.value and isinstance(cell_b.value, str):
            if re.match(r'^\d+\.\d+', cell_b.value):
                # Chiude la sezione precedente
                if current_section and section_start:
                    sections[current_section] = analyze_section_content(sheet, section_start, row-1)
                
                # Inizia nuova sezione
                current_section = cell_b.value
                section_start = row
    
    # Chiude l'ultima sezione
    if current_section and section_start:
        sections[current_section] = analyze_section_content(sheet, section_start, min(sheet.max_row, 250))
    
    return sections


def analyze_calcoli_structure(sheet):
    """Analizza la struttura del sheet Calcoli"""
    
    sections = {}
    
    # Analizza per blocchi di righe per identificare sezioni logiche
    ranges = [
        ("Sezione A - Righe 1-50", 1, 50),
        ("Sezione B - Righe 51-100", 51, 100), 
        ("Sezione C - Righe 101-150", 101, 150),
        ("Sezione D - Righe 151-200", 151, 200),
        ("Sezione E - Righe 201-250", 201, 250),
        ("Sezione F - Righe 251-300", 251, 300),
    ]
    
    for section_name, start_row, end_row in ranges:
        if start_row <= sheet.max_row:
            sections[section_name] = analyze_section_content(sheet, start_row, min(end_row, sheet.max_row))
    
    return sections


def analyze_section_content(sheet, start_row, end_row):
    """Analizza il contenuto di una sezione specifica"""
    
    populated_cells = []
    formulas_count = 0
    key_cells = []
    
    for row in range(start_row, end_row + 1):
        for col in range(1, min(sheet.max_column + 1, 25)):  # Limita a 25 colonne per performance
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                cell_addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
                populated_cells.append(cell_addr)
                
                if cell.data_type == 'f':  # Formula
                    formulas_count += 1
                
                # Identifica celle chiave (con testo descrittivo in colonna B o C)
                if col <= 3 and isinstance(cell.value, str) and len(cell.value) > 5:
                    key_cells.append(f"{cell_addr}:{cell.value[:30]}")
    
    # Determina range effettivo
    if populated_cells:
        range_desc = f"{populated_cells[0]} - {populated_cells[-1]}"
    else:
        range_desc = "Vuoto"
    
    # Genera descrizione del contenuto
    description = ""
    if key_cells:
        # Prende i primi elementi descrittivi per capire il contenuto
        first_descriptions = [k.split(':')[1] for k in key_cells[:3]]
        description = "; ".join(first_descriptions)
    
    return {
        'range': range_desc,
        'cells_count': len(populated_cells),
        'formulas_count': formulas_count,
        'key_cells': [k.split(':')[0] for k in key_cells],
        'description': description
    }


def find_cross_sheet_dependencies(calcoli_sheet):
    """Trova dipendenze tra sheet Calcoli e Input"""
    
    dependencies = {}
    
    # Analizza le prime 100 righe per trovare riferimenti a Input
    for row in range(1, min(101, calcoli_sheet.max_row + 1)):
        for col in range(1, min(26, calcoli_sheet.max_column + 1)):
            cell = calcoli_sheet.cell(row=row, column=col)
            
            if cell.data_type == 'f' and cell.value:
                formula = str(cell.value)
                # Cerca riferimenti al sheet Input
                input_refs = re.findall(r'Input![A-Z]+\d+', formula)
                if input_refs:
                    cell_addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    dependencies[cell_addr] = input_refs
    
    return dependencies


if __name__ == "__main__":
    analyze_excel_structure()
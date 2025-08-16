#!/usr/bin/env python3
"""
Analisi dettagliata del file Excel modello.xlsx - Sheet Input e Calcoli
"""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import re


def get_cell_type(cell) -> str:
    """Determina il tipo di contenuto della cella"""
    if cell.value is None:
        return "vuoto"
    elif cell.data_type == 'f':  # formula
        return "formula"
    elif isinstance(cell.value, (int, float)):
        return "numero"
    elif isinstance(cell.value, str):
        return "testo"
    elif hasattr(cell.value, 'date'):  # datetime
        return "data"
    else:
        return "altro"


def analyze_formula(formula: str) -> Dict[str, Any]:
    """Analizza una formula Excel per identificare riferimenti e funzioni"""
    if not formula or not formula.startswith('='):
        return {}
    
    # Trova tutti i riferimenti di cella (es. A1, B2:C5)
    cell_refs = re.findall(r'[A-Z]+\d+(?::[A-Z]+\d+)?', formula)
    
    # Trova tutte le funzioni Excel
    functions = re.findall(r'([A-Z]+)\s*\(', formula)
    
    return {
        "formula_text": formula,
        "cell_references": list(set(cell_refs)),
        "functions_used": list(set(functions)),
        "complexity": len(cell_refs) + len(functions)
    }


def analyze_sheet(worksheet, sheet_name: str) -> Dict[str, Any]:
    """Analizza un singolo sheet del file Excel"""
    
    populated_cells = []
    cell_groups = {}
    dependencies = {}
    
    # Determina l'area utilizzata del foglio
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    print(f"\nAnalizzando sheet '{sheet_name}' - Area: {get_column_letter(max_col)}{max_row}")
    
    # Analizza ogni cella nell'area utilizzata
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            
            if cell.value is not None:  # Cella popolata
                cell_address = f"{get_column_letter(col)}{row}"
                cell_type = get_cell_type(cell)
                
                cell_info = {
                    "address": cell_address,
                    "row": row,
                    "column": col,
                    "value": cell.value,
                    "type": cell_type,
                    "display_value": str(cell.value)[:100] if cell.value else "",
                }
                
                # Se è una formula, analizzala
                if cell_type == "formula":
                    formula_analysis = analyze_formula(str(cell.value))
                    cell_info.update(formula_analysis)
                    
                    # Mappa le dipendenze
                    if formula_analysis.get("cell_references"):
                        dependencies[cell_address] = formula_analysis["cell_references"]
                
                # Aggiungi informazioni di formattazione se rilevanti
                if cell.font and cell.font.bold:
                    cell_info["formatting"] = "bold"
                
                populated_cells.append(cell_info)
    
    # Identifica raggruppamenti logici basati su posizione e contenuto
    cell_groups = identify_logical_groups(populated_cells)
    
    return {
        "sheet_name": sheet_name,
        "total_populated_cells": len(populated_cells),
        "max_row": max_row,
        "max_column": max_col,
        "populated_cells": populated_cells,
        "cell_groups": cell_groups,
        "dependencies": dependencies,
        "formulas_count": len([c for c in populated_cells if c["type"] == "formula"]),
        "numbers_count": len([c for c in populated_cells if c["type"] == "numero"]),
        "text_count": len([c for c in populated_cells if c["type"] == "testo"])
    }


def identify_logical_groups(populated_cells: List[Dict]) -> Dict[str, List[str]]:
    """Identifica raggruppamenti logici di celle basati su posizione e contenuto"""
    groups = {}
    
    # Raggruppa per righe consecutive
    row_groups = {}
    for cell in populated_cells:
        row = cell["row"]
        if row not in row_groups:
            row_groups[row] = []
        row_groups[row].append(cell["address"])
    
    # Identifica sezioni basate su celle di intestazione (testo in grassetto)
    header_cells = [c for c in populated_cells if c.get("formatting") == "bold"]
    
    # Raggruppa celle per tipo
    formula_cells = [c["address"] for c in populated_cells if c["type"] == "formula"]
    input_cells = [c["address"] for c in populated_cells if c["type"] in ["numero", "testo"] and c["type"] != "formula"]
    
    groups["formule"] = formula_cells
    groups["input_values"] = input_cells
    groups["headers"] = [c["address"] for c in header_cells]
    groups["by_rows"] = row_groups
    
    return groups


def create_dependency_map(dependencies: Dict[str, List[str]]) -> Dict[str, Any]:
    """Crea una mappa delle dipendenze tra celle"""
    # Inverte le dipendenze per trovare quali celle dipendono da una data cella
    dependent_on = {}
    depends_on = dependencies.copy()
    
    for formula_cell, referenced_cells in dependencies.items():
        for ref_cell in referenced_cells:
            if ref_cell not in dependent_on:
                dependent_on[ref_cell] = []
            dependent_on[ref_cell].append(formula_cell)
    
    return {
        "depends_on": depends_on,  # Cosa usa ogni cella
        "dependent_on": dependent_on  # Cosa dipende da ogni cella
    }


def generate_report(input_analysis: Dict, calcoli_analysis: Dict) -> str:
    """Genera un report dettagliato dell'analisi"""
    
    report = "# ANALISI DETTAGLIATA MODELLO EXCEL - PIANO INDUSTRIALE BANCARIO\n\n"
    
    # Sezione Input
    report += "## 📋 SHEET 'INPUT'\n\n"
    report += f"**Dimensioni**: {input_analysis['max_column']} colonne × {input_analysis['max_row']} righe\n"
    report += f"**Celle popolate**: {input_analysis['total_populated_cells']}\n"
    report += f"**Composizione**: {input_analysis['numbers_count']} numeri, {input_analysis['text_count']} testi, {input_analysis['formulas_count']} formule\n\n"
    
    report += "### 📊 CELLE POPOLATE DETTAGLIO:\n\n"
    for cell in input_analysis['populated_cells']:
        report += f"**{cell['address']}** ({cell['type']}) = {cell['display_value']}\n"
        if cell['type'] == 'formula' and 'functions_used' in cell:
            report += f"  - Funzioni: {', '.join(cell['functions_used'])}\n"
            report += f"  - Riferimenti: {', '.join(cell['cell_references'])}\n"
        report += "\n"
    
    # Sezione Calcoli
    report += "\n## 🔢 SHEET 'CALCOLI'\n\n"
    report += f"**Dimensioni**: {calcoli_analysis['max_column']} colonne × {calcoli_analysis['max_row']} righe\n"
    report += f"**Celle popolate**: {calcoli_analysis['total_populated_cells']}\n"
    report += f"**Composizione**: {calcoli_analysis['numbers_count']} numeri, {calcoli_analysis['text_count']} testi, {calcoli_analysis['formulas_count']} formule\n\n"
    
    report += "### 📊 CELLE POPOLATE DETTAGLIO:\n\n"
    for cell in calcoli_analysis['populated_cells']:
        report += f"**{cell['address']}** ({cell['type']}) = {cell['display_value']}\n"
        if cell['type'] == 'formula' and 'functions_used' in cell:
            report += f"  - Funzioni: {', '.join(cell['functions_used'])}\n"
            report += f"  - Riferimenti: {', '.join(cell['cell_references'])}\n"
        report += "\n"
    
    # Analisi dipendenze
    if calcoli_analysis['dependencies']:
        report += "\n### 🔗 MAPPA DIPENDENZE FORMULE:\n\n"
        dep_map = create_dependency_map(calcoli_analysis['dependencies'])
        
        for formula_cell, referenced in calcoli_analysis['dependencies'].items():
            report += f"**{formula_cell}** dipende da: {', '.join(referenced)}\n"
    
    # Raggruppamenti logici
    report += "\n### 📁 RAGGRUPPAMENTI LOGICI:\n\n"
    
    if input_analysis['cell_groups']['headers']:
        report += f"**Headers Input**: {', '.join(input_analysis['cell_groups']['headers'])}\n"
    if input_analysis['cell_groups']['formule']:
        report += f"**Formule Input**: {', '.join(input_analysis['cell_groups']['formule'])}\n"
    if input_analysis['cell_groups']['input_values']:
        report += f"**Valori Input**: {', '.join(input_analysis['cell_groups']['input_values'])}\n"
    
    report += "\n"
    
    if calcoli_analysis['cell_groups']['headers']:
        report += f"**Headers Calcoli**: {', '.join(calcoli_analysis['cell_groups']['headers'])}\n"
    if calcoli_analysis['cell_groups']['formule']:
        report += f"**Formule Calcoli**: {', '.join(calcoli_analysis['cell_groups']['formule'])}\n"
    if calcoli_analysis['cell_groups']['input_values']:
        report += f"**Valori Calcoli**: {', '.join(calcoli_analysis['cell_groups']['input_values'])}\n"
    
    return report


def main():
    """Funzione principale per l'analisi del file Excel"""
    
    file_path = "/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello.xlsx"
    
    try:
        # Carica il workbook
        print(f"Caricamento del file: {file_path}")
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        # Verifica che esistano i sheet richiesti
        available_sheets = workbook.sheetnames
        print(f"Sheet disponibili: {available_sheets}")
        
        # Analizza i sheet richiesti
        analyses = {}
        
        if "Input" in available_sheets:
            print("\nAnalizzando sheet 'Input'...")
            analyses["Input"] = analyze_sheet(workbook["Input"], "Input")
        else:
            print("Sheet 'Input' non trovato!")
            return
        
        if "Calcoli" in available_sheets:
            print("\nAnalizzando sheet 'Calcoli'...")
            analyses["Calcoli"] = analyze_sheet(workbook["Calcoli"], "Calcoli")
        else:
            print("Sheet 'Calcoli' non trovato!")
            return
        
        # Genera il report
        report = generate_report(analyses["Input"], analyses["Calcoli"])
        
        # Salva il report
        report_path = "/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/analisi_report.md"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report)
        
        print(f"\n✅ Analisi completata! Report salvato in: {report_path}")
        print("\n" + "="*60)
        print(report)
        
        workbook.close()
        
    except FileNotFoundError:
        print(f"❌ File non trovato: {file_path}")
    except Exception as e:
        print(f"❌ Errore durante l'analisi: {str(e)}")


if __name__ == "__main__":
    main()
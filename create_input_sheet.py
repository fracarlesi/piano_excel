#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script per creare il foglio Input con tutte le assumptions del modello bancario.
Versione aggiornata con:
- Nomi completi dei prodotti invece delle sigle
- Sezione dettagliata per FTE per divisione e funzione
- Sezione RAL per seniority e divisione
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_input_sheet():
    """Crea il file Excel con il foglio Input popolato con tutte le assumptions"""
    
    # Crea nuovo workbook
    wb = Workbook()
    
    # Rimuovi il foglio default e crea il foglio Input
    wb.remove(wb.active)
    ws = wb.create_sheet("Input")
    
    # Imposta larghezza colonne
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 35
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 40
    
    # Font di default
    default_font = Font(name='Calibri', size=11)
    
    # Riga corrente
    current_row = 2
    
    # Titolo principale
    title_cell = ws.cell(row=current_row, column=2, value="ASSUMPTIONS - PARAMETRI DI INPUT")
    ws.merge_cells(f'B{current_row}:I{current_row}')
    title_cell.font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 3
    
    # Funzione helper per inserire titolo sezione
    def insert_section_title(title, row):
        cell = ws.cell(row=row, column=2, value=title)
        ws.merge_cells(f'B{row}:I{row}')
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        # Bordi
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col in range(2, 10):  # B to I
            ws.cell(row=row, column=col).border = thin_border
        return row + 2
    
    # Funzione helper per formattare headers
    def format_headers(row, headers, columns):
        for col, header in zip(columns, headers):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # ====================
    # SEZIONE 1.1: Parametri Macro e Tassi di Mercato
    # ====================
    current_row = insert_section_title("1.1 Parametri Macro e Tassi di Mercato", current_row)
    
    # Headers
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Dati ECB Rate
    ws.cell(row=current_row, column=2, value="ECB Rate")
    for col, val in enumerate([0.025, 0.0275, 0.03, 0.03, 0.03], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Base per calcolo costo funding e depositi vincolati")
    current_row += 1
    
    # Dati Euribor 6M
    ws.cell(row=current_row, column=2, value="Euribor 6M")
    for col, val in enumerate([0.0325, 0.035, 0.0375, 0.0375, 0.0375], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Base per pricing prodotti credito a tasso variabile")
    current_row += 3
    
    # ====================
    # SEZIONE 1.2: Bilancio di Partenza (Anno 0)
    # ====================
    current_row = insert_section_title("1.2 Bilancio di Partenza (Anno 0)", current_row)
    
    # Headers
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    # Dati bilancio
    bilancio_data = [
        ("Cash (Anno 0)", 200, "Cassa iniziale per calcolo cash & central banks"),
        ("Stock Crediti RE (Anno 0)", 0, "Stock iniziale crediti RE per roll-forward"),
        ("Stock Crediti SME (Anno 0)", 0, "Stock iniziale crediti SME per roll-forward"),
        ("Stock Crediti PG (Anno 0)", 0, "Stock iniziale crediti PG per roll-forward"),
        ("Stock Titoli (Anno 0)", 0, "Stock iniziale titoli per calcolo interessi treasury"),
        ("Stock Depositi (Anno 0)", 0, "Stock iniziale depositi per calcolo interessi passivi"),
        ("Patrimonio Netto (Anno 0)", 200, "Base equity per calcolo CET1 e dividendi"),
        ("Ricavi Totali (Anno 0)", 0, "Base per calcolo RWA operativi anno 1")
    ]
    
    for param, value, desc in bilancio_data:
        ws.cell(row=current_row, column=2, value=param)
        cell = ws.cell(row=current_row, column=3, value=value)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 2
    
    # ====================
    # SEZIONE 1.3: Parametri di Rischio e Capitale
    # ====================
    current_row = insert_section_title("1.3 Parametri di Rischio e Capitale", current_row)
    
    # Headers
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    # Dati rischio
    rischio_data = [
        ("Pillar 1 Requirement (CET1)", 0.045, "Requisito minimo di capitale CET1", True),
        ("Pillar 2 Requirement (P2R)", 0.02, "Requisito aggiuntivo specifico per la banca", True),
        ("Capital Conservation Buffer", 0.025, "Buffer di conservazione del capitale", True),
        ("Counter-cyclical Buffer", 0.005, "Buffer anticiclico macroprudenziale", True),
        ("RWA Operativi (% Ricavi Y-1)", 0.15, "Moltiplicatore per calcolare il rischio operativo", True),
        ("RWA Mercato (% Portafoglio)", 0.10, "Ponderazione media del rischio di mercato", True),
        ("Deducations da CET1 (% Equity)", 0.01, "Percentuale di deduzioni dal CET1", True)
    ]
    
    for param, value, desc, is_pct in rischio_data:
        ws.cell(row=current_row, column=2, value=param)
        cell = ws.cell(row=current_row, column=3, value=value)
        if is_pct:
            cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 2
    
    # ====================
    # SEZIONE 1.4: Costi Generali, Dividendi e Tasse
    # ====================
    current_row = insert_section_title("1.4 Costi Generali, Dividendi e Tasse", current_row)
    
    # Headers per dati multi-anno
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Spese di Marketing
    ws.cell(row=current_row, column=2, value="Spese di Marketing")
    for col, val in enumerate([2.0, 2.2, 2.5, 2.8, 3.0], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Costo assoluto per marketing e comunicazione")
    current_row += 1
    
    # Consulenze e Spese Legali
    ws.cell(row=current_row, column=2, value="Consulenze e Spese Legali")
    for col, val in enumerate([2.0, 2.1, 2.1, 2.2, 2.2], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Costi per consulenti esterni, legali e revisori")
    current_row += 1
    
    # Costi Immobiliari
    ws.cell(row=current_row, column=2, value="Costi Immobiliari")
    for col, val in enumerate([3.0, 3.0, 3.1, 3.1, 3.2], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Costi per affitti, utenze e manutenzione")
    current_row += 1
    
    # Spese Generali Amministrative
    ws.cell(row=current_row, column=2, value="Spese Generali Amministrative")
    for col, val in enumerate([1.5, 1.5, 1.6, 1.6, 1.7], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Altre spese di funzionamento")
    current_row += 2
    
    # Parametri singoli
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    # Contributo FITD
    ws.cell(row=current_row, column=2, value="Contributo FITD (% Depositi)")
    cell = ws.cell(row=current_row, column=3, value=0.0015)
    cell.number_format = '0.00%'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Contributo obbligatorio al Fondo Interbancario")
    current_row += 1
    
    # Aliquota Fiscale
    ws.cell(row=current_row, column=2, value="Aliquota Fiscale")
    cell = ws.cell(row=current_row, column=3, value=0.28)
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Aliquota fiscale media applicabile")
    current_row += 2
    
    # Dividend Payout
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Dividend Payout")
    for col, val in enumerate([0.30, 0.30, 0.35, 0.35, 0.40], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Percentuale di utile netto distribuita come dividendi")
    current_row += 3
    
    # ====================
    # SEZIONE 1.5: Erogazioni per Divisione
    # ====================
    current_row = insert_section_title("1.5 Erogazioni per Divisione", current_row)
    
    # Real Estate - Valori assoluti per tutti gli anni
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5"],
                  [2, 3, 4, 5, 6, 7])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Nuove Erogazioni RE (€ mln)")
    for col, val in enumerate([200, 210, 220, 227, 234], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Volumi di nuovi finanziamenti divisione Real Estate")
    current_row += 2
    
    # SME - Valori assoluti per tutti gli anni
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5"],
                  [2, 3, 4, 5, 6, 7])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Nuove Erogazioni SME (€ mln)")
    for col, val in enumerate([250, 270, 292, 306, 321], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Volumi di nuovi finanziamenti divisione SME")
    current_row += 2
    
    # Public Guarantee - Valori assoluti per tutti gli anni
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5"],
                  [2, 3, 4, 5, 6, 7])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Nuove Erogazioni PG (€ mln)")
    for col, val in enumerate([50, 55, 61, 66, 69], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Volumi di nuovi finanziamenti divisione Public Guarantee")
    current_row += 3
    
    # ====================
    # SEZIONE 1.6: Parametri Prodotti Real Estate (con nomi completi)
    # ====================
    current_row = insert_section_title("1.6 Parametri Prodotti - Real Estate Division", current_row)
    
    # Headers prodotti con nomi completi - tutti sulla stessa riga
    format_headers(current_row, 
                  ["Parametro", "Construction Bridge Loan", "Mezzanine Loan Amortizing", "Mezzanine Loan Asset Mgmt", 
                   "Mezzanine Loan Pre-Amort", "Mezzanine Pre-Amort Asset Mgmt", "Non-Recourse Estate", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Mix Prodotti
    ws.cell(row=current_row, column=2, value="Mix Prodotti")
    for col, val in enumerate([0.20, 0.50, 0, 0.25, 0, 0.05], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% allocazione delle nuove erogazioni")
    current_row += 1
    
    # Amortizing Type
    ws.cell(row=current_row, column=2, value="Amortizing Type")
    for col, val in enumerate(["bullet", "amortizing", "amortizing", "amortizing", "amortizing", "bullet"], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tipologia di rimborso")
    current_row += 1
    
    # Loan Maturity
    ws.cell(row=current_row, column=2, value="Loan Maturity (Anni)")
    for col, val in enumerate([2, 7, 7, 7, 7, 4], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Durata contrattuale del finanziamento")
    current_row += 1
    
    # Pre-amortizing Period
    ws.cell(row=current_row, column=2, value="Pre-amortizing Period")
    for col, val in enumerate([0, 0, 0, 1, 1, 0], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Periodo iniziale di soli interessi (anni)")
    current_row += 1
    
    # RWA
    ws.cell(row=current_row, column=2, value="RWA (Bonis)")
    for col, val in enumerate([0.75, 0.60, 0.80, 0.75, 0.75, 0.75], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Ponderazione del rischio per crediti in bonis")
    current_row += 1
    
    # Danger Rate
    ws.cell(row=current_row, column=2, value="Danger Rate")
    for col, val in enumerate([0.05, 0.05, 0.05, 0.05, 0.05, 0.05], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di passaggio a default annuale")
    current_row += 1
    
    # LGD
    ws.cell(row=current_row, column=2, value="LGD")
    for col, val in enumerate([0.50, 0.40, 0.30, 0.40, 0.30, 0.006], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Loss Given Default")
    current_row += 1
    
    # Interest Rate
    ws.cell(row=current_row, column=2, value="Interest Rate")
    for col, val in enumerate([0.08, 0.08, 0.06, 0.08, 0.06, 0.06], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di interesse applicato al cliente")
    current_row += 1
    
    # Up-front Fees
    ws.cell(row=current_row, column=2, value="Up-front Fees")
    for col, val in enumerate([0.01, 0.01, 0.01, 0.01, 0.01, 0.01], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Commissioni iniziali sulle erogazioni")
    current_row += 1
    
    # PARAMETRI DEFAULT E RECOVERY per Real Estate
    # Default Timing (trimestri)
    ws.cell(row=current_row, column=2, value="Default Timing (Q)")
    for col in range(3, 9):
        cell = ws.cell(row=current_row, column=col, value=12)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Trimestre medio di manifestazione default")
    current_row += 1
    
    # Recovery Timing (trimestri)
    ws.cell(row=current_row, column=2, value="Recovery Timing (Q)")
    for col in range(3, 9):
        cell = ws.cell(row=current_row, column=col, value=8)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Trimestre medio di recupero post-default")
    current_row += 1
    
    # Recovery Rate Garanzia Immobiliare
    ws.cell(row=current_row, column=2, value="Recovery Rate Immobili")
    for col, val in enumerate([0.60, 0.65, 0.70, 0.65, 0.70, 0.60], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% recupero da garanzie immobiliari")
    current_row += 1
    
    # Recovery Rate Garanzia MCC
    ws.cell(row=current_row, column=2, value="Recovery Rate MCC")
    for col in range(3, 9):
        cell = ws.cell(row=current_row, column=col, value=0.80)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% recupero da garanzie MCC")
    current_row += 1
    
    # PARAMETRI ECL per Real Estate
    # ECL Horizon (trimestri)
    ws.cell(row=current_row, column=2, value="ECL Horizon (Q)")
    for col in range(3, 9):
        cell = ws.cell(row=current_row, column=col, value=4)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Orizzonte temporale ECL (trimestri)")
    current_row += 1
    
    # PD Multiplier
    ws.cell(row=current_row, column=2, value="PD Multiplier")
    for col in range(3, 9):
        cell = ws.cell(row=current_row, column=col, value=1.5)
        cell.number_format = '0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Moltiplicatore PD per stress")
    current_row += 3
    
    # ====================
    # SEZIONE 1.7: Parametri Prodotti SME (con nomi completi)
    # ====================
    current_row = insert_section_title("1.7 Parametri Prodotti - SME Division", current_row)
    
    # Headers prodotti SME con nomi completi
    format_headers(current_row, 
                  ["Parametro", "Business Loan", "Refinancing", "State Support", "New Finance", "Restructuring", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Mix Prodotti SME
    ws.cell(row=current_row, column=2, value="Mix Prodotti")
    for col, val in enumerate([0.15, 0.30, 0.10, 0.40, 0.05], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% allocazione delle nuove erogazioni")
    current_row += 1
    
    # Amortizing Type SME
    ws.cell(row=current_row, column=2, value="Amortizing Type")
    for col, val in enumerate(["bullet", "amortizing", "amortizing", "amortizing", "amortizing"], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tipologia di rimborso")
    current_row += 1
    
    # Loan Maturity SME
    ws.cell(row=current_row, column=2, value="Loan Maturity (Anni)")
    for col, val in enumerate([2, 5, 4, 4, 5], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Durata contrattuale del finanziamento")
    current_row += 1
    
    # Pre-amortizing Period SME
    ws.cell(row=current_row, column=2, value="Pre-amortizing Period")
    for col, val in enumerate([0, 2, 0, 0, 0], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Periodo iniziale di soli interessi (anni)")
    current_row += 1
    
    # RWA SME
    ws.cell(row=current_row, column=2, value="RWA (Bonis)")
    for col, val in enumerate([0.80, 0.80, 1.00, 1.35, 1.00], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Ponderazione del rischio per crediti in bonis")
    current_row += 1
    
    # Danger Rate SME
    ws.cell(row=current_row, column=2, value="Danger Rate")
    for col, val in enumerate([0.10, 0.10, 0.10, 0.10, 0.10], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di passaggio a default annuale")
    current_row += 1
    
    # LGD SME
    ws.cell(row=current_row, column=2, value="LGD")
    for col, val in enumerate([0.50, 0.40, 0.30, 0.40, 0.60], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Loss Given Default")
    current_row += 1
    
    # Interest Rate SME
    ws.cell(row=current_row, column=2, value="Interest Rate")
    for col, val in enumerate([0.08, 0.08, 0.08, 0.08, 0.08], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di interesse applicato al cliente")
    current_row += 1
    
    # Up-front Fees SME
    ws.cell(row=current_row, column=2, value="Up-front Fees")
    for col, val in enumerate([0.01, 0.01, 0.01, 0.01, 0.01], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Commissioni iniziali sulle erogazioni")
    current_row += 1
    
    # PARAMETRI DEFAULT E RECOVERY per SME
    # Default Timing (trimestri)
    ws.cell(row=current_row, column=2, value="Default Timing (Q)")
    for col in range(3, 8):
        cell = ws.cell(row=current_row, column=col, value=8)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Trimestre medio di manifestazione default")
    current_row += 1
    
    # Recovery Timing (trimestri)
    ws.cell(row=current_row, column=2, value="Recovery Timing (Q)")
    for col in range(3, 8):
        cell = ws.cell(row=current_row, column=col, value=6)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Trimestre medio di recupero post-default")
    current_row += 1
    
    # Recovery Rate Garanzia Immobiliare
    ws.cell(row=current_row, column=2, value="Recovery Rate Immobili")
    for col, val in enumerate([0.50, 0.55, 0.40, 0.45, 0.40], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% recupero da garanzie immobiliari")
    current_row += 1
    
    # Recovery Rate Garanzia MCC
    ws.cell(row=current_row, column=2, value="Recovery Rate MCC")
    for col, val in enumerate([0.70, 0.75, 0.80, 0.75, 0.70], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% recupero da garanzie MCC")
    current_row += 1
    
    # PARAMETRI ECL per SME
    # ECL Horizon (trimestri)
    ws.cell(row=current_row, column=2, value="ECL Horizon (Q)")
    for col in range(3, 8):
        cell = ws.cell(row=current_row, column=col, value=4)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Orizzonte temporale ECL (trimestri)")
    current_row += 1
    
    # PD Multiplier
    ws.cell(row=current_row, column=2, value="PD Multiplier")
    for col in range(3, 8):
        cell = ws.cell(row=current_row, column=col, value=2.0)
        cell.number_format = '0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Moltiplicatore PD per stress")
    current_row += 3
    
    # ====================
    # SEZIONE 1.8: Parametri Prodotti Public Guarantee (con nomi completi)
    # ====================
    current_row = insert_section_title("1.8 Parametri Prodotti - Public Guarantee Division", current_row)
    
    # Headers prodotti PG con nomi completi
    format_headers(current_row, 
                  ["Parametro", "Anticipo Contratti PA", "Fondo Garanzia Amortizing", "Fondo Garanzia Pre-Amortizing", "", "", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Mix Prodotti PG
    ws.cell(row=current_row, column=2, value="Mix Prodotti")
    for col, val in enumerate([0.40, 0.30, 0.30], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% allocazione delle nuove erogazioni")
    current_row += 1
    
    # Amortizing Type PG
    ws.cell(row=current_row, column=2, value="Amortizing Type")
    for col, val in enumerate(["bullet", "amortizing", "amortizing"], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tipologia di rimborso")
    current_row += 1
    
    # Loan Maturity PG
    ws.cell(row=current_row, column=2, value="Loan Maturity (Anni)")
    for col, val in enumerate([1, 5, 7], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Durata contrattuale del finanziamento")
    current_row += 1
    
    # Pre-amortizing Period PG
    ws.cell(row=current_row, column=2, value="Pre-amortizing Period")
    for col, val in enumerate([0, 0, 2], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Periodo iniziale di soli interessi (anni)")
    current_row += 1
    
    # RWA PG
    ws.cell(row=current_row, column=2, value="RWA (Bonis)")
    for col, val in enumerate([0.40, 0, 0], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Ponderazione del rischio per crediti in bonis")
    current_row += 1
    
    # Danger Rate PG
    ws.cell(row=current_row, column=2, value="Danger Rate")
    for col, val in enumerate([0.05, 0.05, 0.05], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di passaggio a default annuale")
    current_row += 1
    
    # LGD PG
    ws.cell(row=current_row, column=2, value="LGD")
    for col, val in enumerate([0.50, 0.50, 0.50], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Loss Given Default")
    current_row += 1
    
    # Interest Rate PG
    ws.cell(row=current_row, column=2, value="Interest Rate")
    for col, val in enumerate([0.05, 0.05, 0.05], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di interesse applicato al cliente")
    current_row += 1
    
    # Up-front Fees PG
    ws.cell(row=current_row, column=2, value="Up-front Fees")
    for col, val in enumerate([0.02, 0.02, 0.02], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Commissioni iniziali sulle erogazioni")
    current_row += 1
    
    # PARAMETRI DEFAULT E RECOVERY per Public Guarantee
    # Default Timing (trimestri)
    ws.cell(row=current_row, column=2, value="Default Timing (Q)")
    for col in range(3, 6):
        cell = ws.cell(row=current_row, column=col, value=10)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Trimestre medio di manifestazione default")
    current_row += 1
    
    # Recovery Timing (trimestri)
    ws.cell(row=current_row, column=2, value="Recovery Timing (Q)")
    for col in range(3, 6):
        cell = ws.cell(row=current_row, column=col, value=4)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Trimestre medio di recupero post-default")
    current_row += 1
    
    # Recovery Rate Garanzia Immobiliare
    ws.cell(row=current_row, column=2, value="Recovery Rate Immobili")
    for col in range(3, 6):
        cell = ws.cell(row=current_row, column=col, value=0.40)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% recupero da garanzie immobiliari")
    current_row += 1
    
    # Recovery Rate Garanzia MCC
    ws.cell(row=current_row, column=2, value="Recovery Rate MCC")
    for col in range(3, 6):
        cell = ws.cell(row=current_row, column=col, value=0.90)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% recupero da garanzie MCC")
    current_row += 1
    
    # PARAMETRI ECL per Public Guarantee
    # ECL Horizon (trimestri)
    ws.cell(row=current_row, column=2, value="ECL Horizon (Q)")
    for col in range(3, 6):
        cell = ws.cell(row=current_row, column=col, value=4)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Orizzonte temporale ECL (trimestri)")
    current_row += 1
    
    # PD Multiplier
    ws.cell(row=current_row, column=2, value="PD Multiplier")
    for col in range(3, 6):
        cell = ws.cell(row=current_row, column=col, value=1.2)
        cell.number_format = '0.0'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Moltiplicatore PD per stress")
    current_row += 3
    
    # ====================
    # NUOVA SEZIONE 1.9: Parametri Personale - FTE per Divisione e Funzione
    # ====================
    current_row = insert_section_title("1.9 Parametri Personale - FTE per Divisione e Funzione", current_row)
    
    # Sottosezione FTE per Divisione Business
    ws.cell(row=current_row, column=2, value="FTE per Divisione Business")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, 
                  ["Divisione", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # FTE per divisioni business
    divisioni_business = [
        ("Real Estate", [15, 18, 22, 25, 28], "Team dedicato al settore immobiliare"),
        ("SME", [15, 20, 25, 30, 35], "Team dedicato alle piccole e medie imprese"),
        ("Public Guarantee", [15, 16, 18, 20, 22], "Team per finanziamenti con garanzia pubblica"),
        ("Digital Banking", [16, 20, 25, 28, 30], "Team per servizi bancari digitali"),
        ("Wealth Management", [13, 15, 18, 20, 22], "Team gestione patrimoniale"),
        ("Tech Platform", [25, 30, 35, 40, 45], "Team sviluppo e gestione piattaforma tecnologica")
    ]
    
    for divisione, valori, desc in divisioni_business:
        ws.cell(row=current_row, column=2, value=divisione)
        for col, val in enumerate(valori, start=3):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 1
    
    # Sottosezione FTE per Funzioni Centrali
    ws.cell(row=current_row, column=2, value="FTE per Funzioni Centrali")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, 
                  ["Funzione", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # FTE per funzioni centrali
    funzioni_centrali = [
        ("CEO Office", [3, 3, 3, 4, 4], "Ufficio del CEO e supporto strategico"),
        ("CFO & Finance", [8, 9, 10, 11, 12], "Controllo di gestione e reporting finanziario"),
        ("Risk Management", [10, 11, 12, 13, 14], "Gestione del rischio e compliance"),
        ("Legal & Compliance", [6, 7, 8, 9, 10], "Affari legali e conformità normativa"),
        ("HR & Organization", [5, 6, 7, 8, 9], "Risorse umane e sviluppo organizzativo"),
        ("Operations", [8, 9, 10, 11, 12], "Operazioni bancarie e back-office"),
        ("Marketing & Communication", [4, 5, 5, 6, 6], "Marketing e comunicazione istituzionale"),
        ("Internal Audit", [4, 4, 5, 5, 6], "Revisione interna"),
        ("Treasury", [3, 3, 4, 4, 4], "Tesoreria e gestione liquidità")
    ]
    
    for funzione, valori, desc in funzioni_centrali:
        ws.cell(row=current_row, column=2, value=funzione)
        for col, val in enumerate(valori, start=3):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 1
    
    # Totale FTE
    ws.cell(row=current_row, column=2, value="TOTALE FTE BANCA")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    totali_fte = [150, 175, 200, 220, 240]
    for col, val in enumerate(totali_fte, start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    ws.cell(row=current_row, column=9, value="Totale organico della banca")
    current_row += 3
    
    # ====================
    # NUOVA SEZIONE 1.10: Parametri RAL per Seniority e Divisione
    # ====================
    current_row = insert_section_title("1.10 Parametri RAL per Seniority e Divisione", current_row)
    
    # Sottosezione RAL Media per Seniority
    ws.cell(row=current_row, column=2, value="RAL Media per Seniority (€/anno)")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, 
                  ["Seniority", "RAL Base", "Bonus Target %", "RAL Total", "", "", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # RAL per seniority
    seniority_data = [
        ("Junior (0-3 anni)", 35000, 0.10, 38500, "Profili junior in ingresso"),
        ("Professional (3-7 anni)", 50000, 0.15, 57500, "Professionisti con esperienza"),
        ("Senior (7-12 anni)", 70000, 0.20, 84000, "Senior professional e specialist"),
        ("Manager", 95000, 0.30, 123500, "Responsabili di team e funzioni"),
        ("Senior Manager", 130000, 0.40, 182000, "Responsabili di divisione"),
        ("Director", 180000, 0.50, 270000, "Direttori e C-level")
    ]
    
    for seniority, ral_base, bonus, ral_total, desc in seniority_data:
        ws.cell(row=current_row, column=2, value=seniority)
        cell = ws.cell(row=current_row, column=3, value=ral_base)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=current_row, column=4, value=bonus)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=current_row, column=5, value=ral_total)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 2
    
    # Sottosezione Mix Seniority per Divisione
    ws.cell(row=current_row, column=2, value="Mix Seniority per Divisione (%)")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, 
                  ["Divisione", "Junior", "Professional", "Senior", "Manager", "Sr Manager", "Director", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Mix seniority per divisione
    mix_seniority = [
        ("Real Estate", [0.20, 0.30, 0.25, 0.15, 0.08, 0.02], "Mix seniority RE"),
        ("SME", [0.25, 0.35, 0.20, 0.12, 0.06, 0.02], "Mix seniority SME"),
        ("Public Guarantee", [0.30, 0.35, 0.20, 0.10, 0.04, 0.01], "Mix seniority PG"),
        ("Digital Banking", [0.35, 0.30, 0.20, 0.10, 0.04, 0.01], "Mix seniority Digital"),
        ("Wealth Management", [0.15, 0.25, 0.30, 0.20, 0.08, 0.02], "Mix seniority Wealth"),
        ("Tech Platform", [0.40, 0.30, 0.15, 0.10, 0.04, 0.01], "Mix seniority Tech"),
        ("Funzioni Centrali", [0.25, 0.30, 0.25, 0.12, 0.06, 0.02], "Mix seniority Central")
    ]
    
    for divisione, percentuali, desc in mix_seniority:
        ws.cell(row=current_row, column=2, value=divisione)
        for col, val in enumerate(percentuali, start=3):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.number_format = '0%'
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 2
    
    # Sottosezione Incrementi RAL Annuali
    ws.cell(row=current_row, column=2, value="Incrementi RAL Annuali")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # Incrementi RAL
    ws.cell(row=current_row, column=2, value="Incremento RAL Base")
    for col, val in enumerate([0.02, 0.025, 0.025, 0.03, 0.03], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Incremento annuale delle retribuzioni")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Inflazione Attesa")
    for col, val in enumerate([0.02, 0.02, 0.02, 0.02, 0.02], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di inflazione atteso")
    current_row += 2
    
    # Sottosezione Altri Costi del Personale
    ws.cell(row=current_row, column=2, value="Altri Costi del Personale")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    # Altri costi personale
    altri_costi = [
        ("Contributi Previdenziali (%RAL)", 0.30, "Contributi INPS e previdenza complementare"),
        ("TFR (%RAL)", 0.07, "Accantonamento TFR annuale"),
        ("Welfare Aziendale (€/FTE)", 2000, "Benefit e welfare per dipendente"),
        ("Formazione (€/FTE)", 1500, "Budget formazione per dipendente"),
        ("Trasferte e Rimborsi (€/FTE)", 3000, "Rimborsi spese e trasferte medie"),
        ("Turnover Rate Annuo", 0.10, "Tasso di turnover del personale"),
        ("Costo Recruiting (€/nuovo FTE)", 5000, "Costo medio per nuova assunzione")
    ]
    
    for param, value, desc in altri_costi:
        ws.cell(row=current_row, column=2, value=param)
        cell = ws.cell(row=current_row, column=3, value=value)
        if "%" in param:
            cell.number_format = '0.0%'
        elif "€" in param:
            cell.number_format = '#,##0'
        elif param == "Turnover Rate Annuo":
            cell.number_format = '0.0%'
        else:
            cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 3
    
    # ====================
    # SEZIONE 1.11: Altri Parametri (Digital, Wealth, Treasury, IT)
    # ====================
    current_row = insert_section_title("1.11 Altri Parametri Divisioni", current_row)
    
    # Digital Banking parameters (mantenuti come prima)
    ws.cell(row=current_row, column=2, value="Digital Banking")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    digital_data = [
        ("Clienti Base (Anno 0)", 50000, "Numero di clienti con conto base all'inizio"),
        ("Clienti Premium (Anno 0)", 5000, "Numero di clienti con conto premium all'inizio"),
        ("Deposito Medio Cliente Base (€)", 2000, "Giacenza media per cliente base"),
        ("Deposito Medio Cliente Premium (€)", 15000, "Giacenza media per cliente premium"),
        ("Canone Mensile Base (€)", 3, "Canone mensile conto base"),
        ("Canone Mensile Premium (€)", 8, "Canone mensile conto premium")
    ]
    
    for param, value, desc in digital_data:
        ws.cell(row=current_row, column=2, value=param)
        cell = ws.cell(row=current_row, column=3, value=value)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 1
    
    # Crescita clienti Digital
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Crescita Clienti Base")
    for col, val in enumerate([0.20, 0.18, 0.15, 0.12, 0.10], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di crescita clienti base")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Crescita Clienti Premium")
    for col, val in enumerate([0.30, 0.28, 0.25, 0.25, 0.25], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Tasso di crescita clienti premium")
    current_row += 2
    
    # Wealth Management parameters
    ws.cell(row=current_row, column=2, value="Wealth Management")
    ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=11, bold=True, italic=True)
    current_row += 1
    
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="AUM Growth")
    for col, val in enumerate([0.15, 0.15, 0.12, 0.12, 0.10], start=3):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Crescita Asset Under Management")
    current_row += 1
    
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Management Fee (% AUM)")
    cell = ws.cell(row=current_row, column=3, value=0.008)
    cell.number_format = '0.00%'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Commissione di gestione media ricorrente")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Performance Fee (% AUM Perf)")
    cell = ws.cell(row=current_row, column=3, value=0.15)
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Commissione di performance")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="AUM Performance (% AUM Totali)")
    cell = ws.cell(row=current_row, column=3, value=0.30)
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="% AUM che genera performance fee")
    current_row += 3
    
    # ====================
    # SEZIONE 1.12: IT, Telefonia e CAPEX
    # ====================
    current_row = insert_section_title("1.12 IT, Telefonia e CAPEX", current_row)
    
    format_headers(current_row, 
                  ["Parametro", "Y1", "Y2", "Y3", "Y4", "Y5", "", "Descrizione"],
                  [2, 3, 4, 5, 6, 7, 8, 9])
    current_row += 1
    
    # IT costs
    it_costs = [
        ("Licenza Temenos", [2.0, 2.1, 2.2, 2.3, 2.4], "Sistema di core banking"),
        ("Costi Cloud", [1.5, 2.0, 2.5, 2.8, 3.0], "Infrastruttura cloud"),
        ("Costi Infoprovider", [0.5, 0.5, 0.6, 0.6, 0.7], "Dati e servizi informativi"),
        ("Canone Internet", [0.2, 0.2, 0.25, 0.25, 0.3], "Connettività aziendale"),
        ("Licenze Software", [0.3, 0.35, 0.4, 0.45, 0.5], "Altre licenze software"),
        ("Sviluppo Software (CAPEX)", [5.0, 4.0, 3.0, 2.0, 1.0], "Investimenti capitalizzati")
    ]
    
    for param, values, desc in it_costs:
        ws.cell(row=current_row, column=2, value=param)
        for col, val in enumerate(values, start=3):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.number_format = '#,##0.0'
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=9, value=desc)
        current_row += 1
    
    current_row += 1
    
    format_headers(current_row, ["Parametro", "Valore"], [2, 3])
    format_headers(current_row, ["Descrizione"], [9])
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Noleggio Dispositivi per FTE (€)")
    cell = ws.cell(row=current_row, column=3, value=800)
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Costo annuo noleggio PC/laptop per dipendente")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Telefonia per FTE (€)")
    cell = ws.cell(row=current_row, column=3, value=400)
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Costo annuo telefonia mobile per dipendente")
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Vita Utile Software (Anni)")
    cell = ws.cell(row=current_row, column=3, value=5)
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=9, value="Periodo di ammortamento software")
    current_row += 3
    
    # Aggiungi bordi finali a tutto il contenuto
    last_row = current_row - 1
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(2, last_row + 1):
        for col in range(2, 10):  # B to I
            if ws.cell(row=row, column=col).value is not None:
                ws.cell(row=row, column=col).border = thin_border
    
    # Non salva più il file qui, restituisce il workbook
    return wb

def main():
    """Funzione principale per creare il modello bancario con foglio Input aggiornato"""
    print("📊 Creazione del modello bancario - STEP 1: Foglio Input (versione aggiornata)...")
    
    # Crea il workbook con il foglio Input
    wb = create_input_sheet()
    
    # Salva il file con il nome definitivo come da CLAUDE.md
    output_file = 'modello_bancario_completo.xlsx'
    wb.save(output_file)
    print(f"✅ File Excel '{output_file}' creato con successo!")
    print("📊 STEP 1 COMPLETATO: Il foglio Input contiene:")
    print("   - Nomi completi dei prodotti invece delle sigle")
    print("   - Sezione dettagliata FTE per divisione e funzione")
    print("   - Sezione RAL per seniority e divisione")
    print("   - Mix seniority per divisione")
    print("   - Altri costi del personale")

if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Analisi specifica della riga 99 e identificazione del pattern colonne
per le matrici GBV DEFAULTED
"""

import openpyxl
from openpyxl.utils import get_column_letter

def analizza_riga_99_pattern():
    file_path = "modello.xlsx"
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws_calcoli = wb['Calcoli']
        
        print("🎯 ANALISI SPECIFICA RIGA 99 - MATRICE GBV DEFAULTED")
        print("=" * 60)
        
        # 1. Identifica la struttura della matrice GBV DEFAULTED
        print("\n📊 STRUTTURA MATRICE GBV DEFAULTED (Riga 95-110)")
        print("-" * 50)
        
        # Trova tutte le colonne che hanno un header "Anno X" o "TX"
        colonne_temporali = {}
        
        # Analizza le righe 97-98 per mappare la struttura temporale
        for col in range(1, 100):  # Espande la ricerca
            col_letter = get_column_letter(col)
            
            # Riga 97: Anno
            cell_97 = ws_calcoli.cell(row=97, column=col)
            # Riga 98: Trimestre
            cell_98 = ws_calcoli.cell(row=98, column=col)
            
            if cell_97.value or cell_98.value:
                anno = str(cell_97.value) if cell_97.value else ""
                trimestre = str(cell_98.value) if cell_98.value else ""
                
                colonne_temporali[col] = {
                    'lettera': col_letter,
                    'anno': anno,
                    'trimestre': trimestre,
                    'posizione': col
                }
        
        # Stampa la mappatura temporale
        print("🗓️ MAPPATURA TEMPORALE COLONNE:")
        anni_ordinati = {}
        for col, info in colonne_temporali.items():
            if 'Anno' in info['anno']:
                anno_num = info['anno'].replace('Anno ', '')
                if anno_num not in anni_ordinati:
                    anni_ordinati[anno_num] = []
                anni_ordinati[anno_num].append(info)
        
        for anno in sorted(anni_ordinati.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            print(f"\n  📅 {anno}:")
            trimestri = sorted(anni_ordinati[anno], key=lambda x: x['posizione'])
            for trim in trimestri:
                print(f"    {trim['lettera']}: {trim['trimestre']}")
        
        # 2. Identifica i valori esistenti nella riga 99
        print(f"\n📊 VALORI ESISTENTI RIGA 99:")
        print("-" * 30)
        
        valori_riga_99 = {}
        for col in range(1, 100):
            cell = ws_calcoli.cell(row=99, column=col)
            if cell.value is not None:
                col_letter = get_column_letter(col)
                if cell.data_type == 'f':
                    valore = f"FORMULA: {str(cell.value)[:50]}..."
                else:
                    valore = str(cell.value)
                valori_riga_99[col] = {
                    'lettera': col_letter,
                    'valore': valore,
                    'tipo': 'formula' if cell.data_type == 'f' else 'valore'
                }
        
        for col, info in valori_riga_99.items():
            print(f"  {info['lettera']}99: {info['valore']} ({info['tipo']})")
        
        # 3. Cerca pattern nelle righe circostanti per riferimenti
        print(f"\n📊 PATTERN DELLE FORMULE NELLE RIGHE ADIACENTI:")
        print("-" * 50)
        
        formule_pattern = {}
        for row in range(95, 110):
            for col in range(1, 50):  # Prime 50 colonne
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.data_type == 'f':
                    col_letter = get_column_letter(col)
                    formula = str(cell.value)
                    
                    # Cerca pattern comuni
                    if row not in formule_pattern:
                        formule_pattern[row] = []
                    formule_pattern[row].append({
                        'posizione': f"{col_letter}{row}",
                        'formula': formula[:80] + "..." if len(formula) > 80 else formula
                    })
        
        for row in sorted(formule_pattern.keys()):
            print(f"\n  Riga {row}:")
            for formula_info in formule_pattern[row][:3]:  # Prime 3 formule per riga
                print(f"    {formula_info['posizione']}: {formula_info['formula']}")
        
        # 4. Identifica la logica di calcolo per la riga 99
        print(f"\n📊 SUGGERIMENTI PER FORMULE RIGA 99:")
        print("-" * 40)
        
        print("Basandomi sulla struttura identificata:")
        print("1. La riga 99 rappresenta l'anno 1 della matrice GBV DEFAULTED")
        print("2. Le colonne B-E dovrebbero contenere i valori per T1-T4 dell'Anno 1")
        print("3. Le colonne F-I dovrebbero contenere i valori per T1-T4 dell'Anno 2")
        print("4. E così via per gli anni successivi...")
        
        # Mappa le colonne per i primi anni
        mapping_colonne = []
        for anno in ['1', '2', '3', '4', '5']:
            trimestri_anno = [info for info in colonne_temporali.values() 
                            if f'Anno {anno}' in info['anno']]
            if trimestri_anno:
                trimestri_ordinati = sorted(trimestri_anno, key=lambda x: x['posizione'])
                for i, trim in enumerate(trimestri_ordinati):
                    mapping_colonne.append({
                        'colonna': trim['lettera'],
                        'anno': anno,
                        'trimestre': i + 1,
                        'descrizione': f"Anno {anno} - T{i + 1}"
                    })
        
        print(f"\n🗺️ MAPPING COLONNE SUGGERITO PER RIGA 99:")
        for mapping in mapping_colonne[:20]:  # Prime 20 colonne
            print(f"  {mapping['colonna']}99: {mapping['descrizione']}")
        
        # 5. Cerca riferimenti alle celle Input per capire la logica
        print(f"\n📊 ANALISI RIFERIMENTI INPUT NELLE MATRICI:")
        print("-" * 45)
        
        riferimenti_input = set()
        for row in range(95, 110):
            for col in range(1, 50):
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.data_type == 'f' and 'Input!' in str(cell.value):
                    formula = str(cell.value)
                    # Estrae i riferimenti Input
                    import re
                    matches = re.findall(r'Input![A-Z]+\d+', formula)
                    riferimenti_input.update(matches)
        
        print("📋 Riferimenti Input trovati nelle vicinanze:")
        for ref in sorted(list(riferimenti_input))[:10]:
            print(f"  {ref}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Errore durante l'analisi: {str(e)}")

if __name__ == "__main__":
    analizza_riga_99_pattern()
#!/usr/bin/env python3
"""
Script per correggere le formule dei rimborsi nel file Excel
Usa la sintassi corretta con virgolette singole per evitare problemi
"""

import openpyxl
from openpyxl.utils import get_column_letter

def main():
    # Apri il file
    wb = openpyxl.load_workbook("modello.xlsx")
    ws_calc = wb["Calcoli"]

    print("=== CORREZIONE FORMULE RIMBORSI CON SINTASSI CORRETTA ===\n")

    # Per ogni prodotto
    for prodotto in range(20):
        base_col_calc = 2 + (prodotto * 43)
        col_input = 4 + prodotto
        col_letter_input = get_column_letter(col_input)
        
        # Matrice rimborsi (seconda matrice)
        base_row_rimborsi = 3 + 46
        header_row = base_row_rimborsi + 3
        data_start_row = header_row + 1
        
        if prodotto == 0:
            print(f"Prodotto 1 (colonna {col_letter_input}) - Esempi formule:\n")
        
        # Popola matrice
        for erog_row in range(40):
            for trim_col in range(40):
                row = data_start_row + erog_row
                col = base_col_calc + trim_col
                
                # Riferimento erogazione
                erog_row_ref = 7 + erog_row
                erog_col_ref = base_col_calc + erog_row
                erog_cell_ref = get_column_letter(erog_col_ref) + str(erog_row_ref)
                
                if trim_col >= erog_row:
                    trimestri_da_erog = trim_col - erog_row + 1
                    
                    # FORMULA CORRETTA con virgolette singole esterne
                    # Questo evita problemi con i caratteri $
                    formula = f'=SE(Input!${col_letter_input}$67="amortizing";{erog_cell_ref}/Input!${col_letter_input}$68;0)'
                    
                    # Se siamo oltre 40 trimestri, metti 0
                    if trimestri_da_erog > 40:
                        formula = '=0'
                    
                    cell = ws_calc.cell(row=row, column=col)
                    cell.value = formula
                    
                    # Mostra esempi per il primo prodotto
                    if prodotto == 0 and erog_row == 0 and trim_col < 5:
                        cell_ref = get_column_letter(col) + str(row)
                        print(f"  {cell_ref}: {formula}")
                else:
                    # Prima dell'erogazione
                    cell = ws_calc.cell(row=row, column=col)
                    cell.value = '=0'

    print("\n✅ Formule corrette inserite!")
    print("\nCaratteristiche delle formule:")
    print("- Virgolette singole esterne per evitare problemi con $")
    print("- Virgolette doppie interne per i valori stringa")
    print("- Punto e virgola come separatore (Excel italiano)")
    print("- Riferimenti assoluti completi (es. Input!$D$67)")

    # Salva il file
    wb.save("modello.xlsx")
    print("\nFile salvato con successo!")

if __name__ == "__main__":
    main()
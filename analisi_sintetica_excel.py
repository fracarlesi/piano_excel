#!/usr/bin/env python3
"""
Analisi sintetica ma completa del file modello.xlsx
Focus su strutture principali, formule significative e stato di completamento
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import TYPE_FORMULA, TYPE_NUMERIC, TYPE_STRING
from collections import defaultdict
import re
from typing import Dict, List, Any

class ExcelSyntheticAnalyzer:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        
    def load_workbook(self) -> bool:
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            return True
        except Exception as e:
            print(f"❌ ERRORE: {e}")
            return False
    
    def get_significant_formulas(self, worksheet, max_formulas=20) -> List[Dict]:
        """Estrae le formule più significative del foglio"""
        formulas = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == TYPE_FORMULA and cell.value:
                    formula_info = {
                        'address': cell.coordinate,
                        'formula': str(cell.value),
                        'complexity': self.assess_formula_complexity(str(cell.value))
                    }
                    formulas.append(formula_info)
                    
                    if len(formulas) >= max_formulas:
                        return formulas
        
        return formulas
    
    def assess_formula_complexity(self, formula: str) -> str:
        """Valuta la complessità di una formula"""
        function_count = len(re.findall(r'([A-Z][A-Z0-9_]*)\s*\(', formula))
        reference_count = len(re.findall(r'[A-Z]+[0-9]+', formula))
        
        if function_count > 5 or reference_count > 15 or len(formula) > 100:
            return "ALTA"
        elif function_count > 2 or reference_count > 5 or len(formula) > 50:
            return "MEDIA"
        else:
            return "BASSA"
    
    def identify_key_sections(self, worksheet) -> List[Dict]:
        """Identifica le sezioni chiave del foglio"""
        sections = []
        current_section = None
        
        for row_idx in range(1, min(worksheet.max_row + 1, 100)):  # Prime 100 righe
            row_cells = [worksheet.cell(row=row_idx, column=col).value 
                        for col in range(1, min(worksheet.max_column + 1, 10))]
            
            # Cerca intestazioni (celle di testo significative nella colonna B)
            if row_idx <= worksheet.max_row:
                cell_b = worksheet.cell(row=row_idx, column=2).value
                if cell_b and isinstance(cell_b, str) and len(cell_b.strip()) > 5:
                    if any(keyword in cell_b.lower() for keyword in 
                          ['parametr', 'input', 'calcol', 'output', 'ricav', 'cost', 'patrimon', 'capital']):
                        sections.append({
                            'row': row_idx,
                            'title': cell_b.strip(),
                            'type': 'HEADER'
                        })
        
        return sections
    
    def get_temporal_structure(self, worksheet) -> Dict:
        """Analizza la struttura temporale"""
        temporal_info = {'years': [], 'periods': []}
        
        # Cerca nelle prime 20 righe
        for row_idx in range(1, min(21, worksheet.max_row + 1)):
            for col_idx in range(1, min(21, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value and isinstance(cell_value, str):
                    # Cerca pattern di anni
                    years = re.findall(r'20\d{2}', cell_value)
                    if years:
                        temporal_info['years'].extend(years)
                    
                    # Cerca pattern di periodi
                    if any(period in cell_value.upper() for period in ['Y1', 'Y2', 'Y3', 'Q1', 'Q2', 'T1', 'T2']):
                        temporal_info['periods'].append(cell_value)
        
        temporal_info['years'] = sorted(list(set(temporal_info['years'])))
        temporal_info['periods'] = list(set(temporal_info['periods']))
        
        return temporal_info
    
    def analyze_data_density(self, worksheet) -> Dict:
        """Analizza la densità dei dati nel foglio"""
        total_cells = worksheet.max_row * worksheet.max_column
        cells_with_data = 0
        formulas_count = 0
        numeric_count = 0
        text_count = 0
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells_with_data += 1
                    if cell.data_type == TYPE_FORMULA:
                        formulas_count += 1
                    elif cell.data_type == TYPE_NUMERIC:
                        numeric_count += 1
                    elif cell.data_type == TYPE_STRING:
                        text_count += 1
        
        return {
            'total_cells': total_cells,
            'cells_with_data': cells_with_data,
            'formulas': formulas_count,
            'numeric': numeric_count,
            'text': text_count,
            'density_percentage': (cells_with_data / total_cells * 100) if total_cells > 0 else 0
        }
    
    def find_interconnections(self) -> Dict:
        """Trova le interconnessioni tra fogli"""
        connections = defaultdict(list)
        
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            
            # Cerca formule che referenziano altri fogli
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == TYPE_FORMULA and cell.value:
                        formula = str(cell.value)
                        # Pattern per riferimenti esterni: 'Foglio'!Cella o Foglio!Cella
                        external_refs = re.findall(r"'?([^'!]+)'?!([A-Z]+[0-9]+)", formula)
                        for ref_sheet, ref_cell in external_refs:
                            if ref_sheet in self.workbook.sheetnames:
                                connections[sheet_name].append({
                                    'from_cell': cell.coordinate,
                                    'to_sheet': ref_sheet,
                                    'to_cell': ref_cell,
                                    'formula_preview': formula[:50] + "..."
                                })
        
        return dict(connections)
    
    def analyze_sheet(self, sheet_name: str) -> Dict:
        """Analizza un singolo foglio"""
        worksheet = self.workbook[sheet_name]
        
        print(f"\n🔍 Analisi {sheet_name}...")
        
        # Statistiche base
        density = self.analyze_data_density(worksheet)
        
        # Sezioni chiave
        sections = self.identify_key_sections(worksheet)
        
        # Struttura temporale
        temporal = self.get_temporal_structure(worksheet)
        
        # Formule significative
        formulas = self.get_significant_formulas(worksheet)
        
        return {
            'density': density,
            'sections': sections,
            'temporal': temporal,
            'significant_formulas': formulas,
            'dimensions': f"{worksheet.max_column}x{worksheet.max_row}"
        }
    
    def print_summary_report(self, analysis: Dict):
        """Stampa report riassuntivo"""
        print("\n" + "="*80)
        print("📊 REPORT SINTETICO ANALISI EXCEL - modello.xlsx")
        print("="*80)
        
        # Panoramica generale
        print(f"\n🗂️  STRUTTURA GENERALE:")
        print(f"   • Fogli: {len(self.workbook.sheetnames)}")
        print(f"   • Nomi fogli: {', '.join(self.workbook.sheetnames)}")
        
        # Analisi per foglio
        for sheet_name, sheet_analysis in analysis.items():
            if sheet_name == 'interconnections':
                continue
                
            print(f"\n📋 FOGLIO: {sheet_name}")
            print("-" * 50)
            
            density = sheet_analysis['density']
            print(f"   📐 Dimensioni: {sheet_analysis['dimensions']}")
            print(f"   📊 Densità dati: {density['density_percentage']:.1f}%")
            print(f"   🔢 Celle con dati: {density['cells_with_data']:,}")
            print(f"   🧮 Formule: {density['formulas']:,}")
            print(f"   💰 Valori numerici: {density['numeric']:,}")
            print(f"   📝 Testi: {density['text']:,}")
            
            # Struttura temporale
            temporal = sheet_analysis['temporal']
            if temporal['years'] or temporal['periods']:
                print(f"   📅 Struttura temporale:")
                if temporal['years']:
                    print(f"      • Anni trovati: {', '.join(temporal['years'])}")
                if temporal['periods']:
                    print(f"      • Periodi: {', '.join(temporal['periods'][:10])}")
            
            # Sezioni principali
            sections = sheet_analysis['sections']
            if sections:
                print(f"   📑 Sezioni principali:")
                for section in sections[:10]:  # Prime 10 sezioni
                    print(f"      • Riga {section['row']}: {section['title'][:60]}")
                if len(sections) > 10:
                    print(f"      ... e altre {len(sections) - 10} sezioni")
            
            # Formule significative
            formulas = sheet_analysis['significant_formulas']
            if formulas:
                print(f"   🧮 Formule significative:")
                high_complexity = [f for f in formulas if f['complexity'] == 'ALTA']
                medium_complexity = [f for f in formulas if f['complexity'] == 'MEDIA']
                
                print(f"      • Complessità ALTA: {len(high_complexity)}")
                print(f"      • Complessità MEDIA: {len(medium_complexity)}")
                print(f"      • Complessità BASSA: {len(formulas) - len(high_complexity) - len(medium_complexity)}")
                
                # Mostra esempi di formule complesse
                if high_complexity:
                    print(f"      📌 Esempi formule complesse:")
                    for formula in high_complexity[:3]:
                        preview = formula['formula'][:60] + "..." if len(formula['formula']) > 60 else formula['formula']
                        print(f"         {formula['address']}: {preview}")
        
        # Interconnessioni
        interconnections = analysis.get('interconnections', {})
        if interconnections:
            print(f"\n🔗 INTERCONNESSIONI TRA FOGLI:")
            print("-" * 40)
            for source_sheet, connections in interconnections.items():
                if connections:
                    print(f"   📊 {source_sheet} → Collegamenti: {len(connections)}")
                    # Raggruppa per foglio di destinazione
                    target_sheets = defaultdict(int)
                    for conn in connections:
                        target_sheets[conn['to_sheet']] += 1
                    
                    for target, count in target_sheets.items():
                        print(f"      → {target}: {count} riferimenti")
        
        # Stato di completamento
        print(f"\n📈 STATO DI COMPLETAMENTO:")
        print("-" * 30)
        
        total_cells_with_data = sum(
            analysis[sheet]['density']['cells_with_data'] 
            for sheet in self.workbook.sheetnames 
            if sheet in analysis
        )
        total_formulas = sum(
            analysis[sheet]['density']['formulas'] 
            for sheet in self.workbook.sheetnames 
            if sheet in analysis
        )
        
        print(f"   📊 Totale celle con dati: {total_cells_with_data:,}")
        print(f"   🧮 Totale formule: {total_formulas:,}")
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in analysis:
                density_pct = analysis[sheet_name]['density']['density_percentage']
                status = "🟢 DENSO" if density_pct > 10 else "🟡 MEDIO" if density_pct > 2 else "🔴 SPARSO"
                print(f"   {sheet_name}: {status} ({density_pct:.1f}% utilizzato)")
        
        # Raccomandazioni
        print(f"\n💡 RACCOMANDAZIONI:")
        print("-" * 20)
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in analysis:
                sheet_data = analysis[sheet_name]
                density_pct = sheet_data['density']['density_percentage']
                formulas_count = sheet_data['density']['formulas']
                
                if sheet_name.lower() == 'input':
                    if density_pct < 5:
                        print(f"   ⚠️  {sheet_name}: Bassa densità dati - verificare completamento parametri")
                elif sheet_name.lower() == 'calcoli':
                    if formulas_count < 100:
                        print(f"   ⚠️  {sheet_name}: Poche formule - motore di calcolo da sviluppare")
                elif sheet_name.lower() == 'output':
                    if formulas_count < 50:
                        print(f"   ⚠️  {sheet_name}: Output limitato - espandere reporting")
        
        print(f"\n✅ ANALISI COMPLETATA")
    
    def run_analysis(self):
        """Esegue l'analisi completa"""
        if not self.load_workbook():
            return
        
        print("🚀 AVVIO ANALISI SINTETICA EXCEL")
        
        analysis = {}
        
        # Analizza ogni foglio
        for sheet_name in self.workbook.sheetnames:
            analysis[sheet_name] = self.analyze_sheet(sheet_name)
        
        # Analizza interconnessioni
        print(f"\n🔗 Analisi interconnessioni...")
        analysis['interconnections'] = self.find_interconnections()
        
        # Stampa report
        self.print_summary_report(analysis)


def main():
    file_path = "/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello.xlsx"
    analyzer = ExcelSyntheticAnalyzer(file_path)
    analyzer.run_analysis()


if __name__ == "__main__":
    main()
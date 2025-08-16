#!/usr/bin/env python3
"""
Analisi finale e precisa della riga 99 con focus su colonne B-AO
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

def analizza_colonne_b_ao():
    file_path = "modello.xlsx"
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws_calcoli = wb['Calcoli']
        
        print("🎯 ANALISI FINALE COLONNE B-AO RIGA 99")
        print("=" * 50)
        
        # Range B-AO: B=2, AO=41
        col_start = 2  # B
        col_end = 41   # AO
        
        print(f"📊 ANALISI HEADER TEMPORALE (Righe 97-98)")
        print("-" * 40)
        
        # Mappa precisa delle colonne B-AO
        colonne_mapping = []
        
        for col in range(col_start, col_end + 1):
            col_letter = get_column_letter(col)
            
            # Header riga 97 (Anno)
            cell_97 = ws_calcoli.cell(row=97, column=col)
            # Header riga 98 (Trimestre) 
            cell_98 = ws_calcoli.cell(row=98, column=col)
            # Valore riga 99
            cell_99 = ws_calcoli.cell(row=99, column=col)
            
            anno_val = cell_97.value if cell_97.value else ""
            trim_val = cell_98.value if cell_98.value else ""
            val_99 = cell_99.value if cell_99.value else ""
            
            colonne_mapping.append({
                'colonna': col_letter,
                'indice': col,
                'anno': str(anno_val),
                'trimestre': str(trim_val), 
                'valore_99': str(val_99),
                'formula_99': cell_99.data_type == 'f'
            })
        
        # Raggruppa per anno
        print("🗓️ STRUTTURA TEMPORALE IDENTIFICATA:")
        anni_found = {}
        
        for col_info in colonne_mapping:
            if 'Anno' in col_info['anno']:
                anno_num = col_info['anno'].replace('Anno ', '').strip()
                if anno_num not in anni_found:
                    anni_found[anno_num] = []
                anni_found[anno_num].append(col_info)
        
        for anno in sorted(anni_found.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            print(f"\n  📅 Anno {anno}:")
            trimestri = sorted(anni_found[anno], key=lambda x: x['indice'])
            for trim in trimestri:
                status = "VUOTA" if not trim['valore_99'] or trim['valore_99'] == 'None' else f"VALORE: {trim['valore_99']}"
                if trim['formula_99']:
                    status = "FORMULA"
                print(f"    {trim['colonna']}: {trim['trimestre']} - {status}")
        
        # Identifica le colonne vuote per la riga 99
        print(f"\n📊 COLONNE VUOTE NELLA RIGA 99 (CANDIDATE PER FORMULE):")
        print("-" * 55)
        
        colonne_vuote = []
        for col_info in colonne_mapping:
            if not col_info['valore_99'] or col_info['valore_99'] == 'None':
                colonne_vuote.append(col_info)
        
        # Raggruppa le colonne vuote per anno
        vuote_per_anno = {}
        for col_info in colonne_vuote:
            if 'Anno' in col_info['anno']:
                anno_num = col_info['anno'].replace('Anno ', '').strip()
                if anno_num not in vuote_per_anno:
                    vuote_per_anno[anno_num] = []
                vuote_per_anno[anno_num].append(col_info)
        
        for anno in sorted(vuote_per_anno.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            print(f"\n  📅 Anno {anno} - Colonne vuote:")
            trimestri = sorted(vuote_per_anno[anno], key=lambda x: x['indice'])
            for trim in trimestri:
                print(f"    {trim['colonna']}99: {trim['trimestre']}")
        
        # Cerca formule esistenti nelle righe adiacenti per pattern
        print(f"\n📊 PATTERN FORMULE NELLE RIGHE ADIACENTI:")
        print("-" * 45)
        
        # Analizza righe 97-105 per trovare pattern
        pattern_formule = {}
        for row in range(97, 106):
            for col in range(col_start, col_end + 1):
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.data_type == 'f':
                    col_letter = get_column_letter(col)
                    if row not in pattern_formule:
                        pattern_formule[row] = []
                    pattern_formule[row].append({
                        'colonna': col_letter,
                        'formula': str(cell.value)
                    })
        
        for row in sorted(pattern_formule.keys()):
            if pattern_formule[row]:
                print(f"\n  Riga {row}:")
                for formula_info in pattern_formule[row][:5]:  # Prime 5 formule
                    formula_short = formula_info['formula'][:60] + "..." if len(formula_info['formula']) > 60 else formula_info['formula']
                    print(f"    {formula_info['colonna']}: {formula_short}")
        
        # Suggerimenti finali
        print(f"\n🎯 RACCOMANDAZIONI PER RIGA 99:")
        print("-" * 35)
        
        print("1. STRUTTURA IDENTIFICATA:")
        print("   - Le colonne sono organizzate per Anno e Trimestre")
        print("   - Ogni anno ha 4 trimestri consecutivi")
        print("   - La riga 99 rappresenta il primo anno di erogazione")
        
        print("\n2. COLONNE DA POPOLARE:")
        for anno in sorted(vuote_per_anno.keys(), key=lambda x: int(x) if x.isdigit() else 0)[:5]:
            trimestri = sorted(vuote_per_anno[anno], key=lambda x: x['indice'])
            if trimestri:
                print(f"   Anno {anno}: {', '.join([t['colonna'] for t in trimestri])}")
        
        print("\n3. LOGICA SUGGERITA:")
        print("   - La matrice GBV DEFAULTED dovrebbe calcolare i crediti in default")
        print("   - Potrebbe essere basata su tassi di default applicati al portfolio")
        print("   - Riferimenti probabili: parametri Input per tassi default")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Errore durante l'analisi: {str(e)}")

if __name__ == "__main__":
    analizza_colonne_b_ao()
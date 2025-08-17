#!/usr/bin/env python3
"""
Script per popolare le matrici erogazioni direttamente nel file Excel
Esegui questo script Python dal terminale per aggiornare il file Excel
"""

import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime

def calcola_erogazione_diagonale(anno, trimestre, erogazioni_anni, allocazioni_trim, mix_prodotto):
    """Calcola il valore dell'erogazione per un dato anno e trimestre"""
    if anno <= len(erogazioni_anni):
        return erogazioni_anni[anno-1] * allocazioni_trim[trimestre-1] * mix_prodotto
    return 0

def popola_matrice_prodotto(wb, prodotto_num):
    """
    Popola la matrice erogazioni per un prodotto specifico
    
    Args:
        wb: workbook openpyxl
        prodotto_num: numero del prodotto (1 o 2)
    """
    ws_input = wb['Input']
    ws_calcoli = wb['Calcoli']
    
    print(f"\n📊 Popolamento Matrice Prodotto {prodotto_num}")
    print("-" * 50)
    
    # Leggi parametri dal foglio Input
    # Erogazioni annuali (riga 55, colonne D-M)
    erogazioni_anni = []
    for col in range(4, 14):  # Colonne D-M (4-13)
        val = ws_input.cell(row=55, column=col).value
        erogazioni_anni.append(val if val else 0)
    print(f"Erogazioni annuali: {erogazioni_anni[:5]}...")
    
    # Allocazioni trimestrali (riga 58, colonne D-G)
    allocazioni_trim = []
    for col in range(4, 8):  # Colonne D-G (4-7)
        val = ws_input.cell(row=58, column=col).value
        allocazioni_trim.append(val if val else 0.25)
    print(f"Allocazioni trimestrali: {allocazioni_trim}")
    
    # Mix prodotto (riga 66, colonna dipende dal prodotto)
    col_mix = 3 + prodotto_num  # D=4 per P1, E=5 per P2
    mix_prodotto = ws_input.cell(row=66, column=col_mix).value or 0
    print(f"Mix prodotto {prodotto_num}: {mix_prodotto}")
    
    # Determina la colonna iniziale per il prodotto
    if prodotto_num == 1:
        col_start = 2  # Colonna B
        print("Posizione: Colonne B-AM (2-39)")
    elif prodotto_num == 2:
        col_start = 44  # Colonna AR 
        print("Posizione: Colonne AR-CC (44-81)")
    else:
        col_start = 2 + (prodotto_num - 1) * 43
        print(f"Posizione: Colonna {col_start}")
    
    # Pulisci prima la matrice (imposta tutto a 0 o vuoto)
    for row in range(9, 49):  # Righe 9-48
        for col in range(col_start, col_start + 40):
            ws_calcoli.cell(row=row, column=col).value = None
    
    # Popola la diagonale
    valori_inseriti = 0
    for anno in range(1, 11):  # 10 anni
        for trimestre in range(1, 5):  # 4 trimestri per anno
            # Calcola posizione sulla diagonale
            trimestre_assoluto = (anno - 1) * 4 + trimestre
            
            if trimestre_assoluto <= 40:  # Limite a 40 trimestri
                # Posizione nella matrice Excel
                row_excel = 8 + trimestre_assoluto  # Riga 9 per Q1, etc.
                col_excel = col_start + trimestre_assoluto - 1  # Diagonale
                
                # Calcola e inserisci il valore
                valore = calcola_erogazione_diagonale(
                    anno, trimestre, erogazioni_anni, allocazioni_trim, mix_prodotto
                )
                
                if valore != 0:
                    ws_calcoli.cell(row=row_excel, column=col_excel).value = valore
                    valori_inseriti += 1
                    
                    # Mostra primi valori per debug
                    if valori_inseriti <= 5:
                        col_letter = openpyxl.utils.get_column_letter(col_excel)
                        print(f"  Cella {col_letter}{row_excel}: {valore:.2f}")
    
    print(f"✅ Inseriti {valori_inseriti} valori nella diagonale")
    return valori_inseriti

def main():
    """Funzione principale"""
    
    # Path del file Excel
    file_path = '/Users/francescocarlesi/Downloads/Progetti Python/piano industriale excel/modello.xlsx'
    
    if not os.path.exists(file_path):
        print(f"❌ File non trovato: {file_path}")
        return
    
    print("=" * 60)
    print("🚀 POPOLAMENTO MATRICI EROGAZIONI CREDITI")
    print("=" * 60)
    
    try:
        # Carica il workbook
        print(f"\n📂 Apertura file: {os.path.basename(file_path)}")
        wb = load_workbook(file_path, data_only=False)
        
        # Crea backup prima di modificare
        backup_path = file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(backup_path)
        print(f"💾 Backup creato: {os.path.basename(backup_path)}")
        
        # Popola le matrici per i prodotti 1 e 2
        totale_valori = 0
        totale_valori += popola_matrice_prodotto(wb, 1)
        totale_valori += popola_matrice_prodotto(wb, 2)
        
        # Salva il file modificato
        print(f"\n💾 Salvataggio modifiche...")
        wb.save(file_path)
        wb.close()
        
        print("\n" + "=" * 60)
        print(f"✅ COMPLETATO! Inseriti {totale_valori} valori totali")
        print(f"📊 File aggiornato: {os.path.basename(file_path)}")
        print("=" * 60)
        
    except PermissionError:
        print("\n❌ ERRORE: Il file Excel è aperto!")
        print("   Chiudi Excel e riprova.")
    except Exception as e:
        print(f"\n❌ ERRORE: {str(e)}")

if __name__ == "__main__":
    main()
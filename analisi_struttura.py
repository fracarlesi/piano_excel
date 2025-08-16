#!/usr/bin/env python3
"""
Analisi struttura del file modello.xlsx
Focus su:
- Struttura del foglio Calcoli, area riga 99
- Matrici Stock GBV
- Celle con Erogazioni e Rimborsi
- Struttura colonne B-AO nella riga 99
- Formule esistenti come riferimento
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def analizza_modello_excel():
    file_path = "modello.xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ File {file_path} non trovato!")
        return
    
    try:
        # Carica il workbook
        wb = openpyxl.load_workbook(file_path, data_only=False)
        print(f"✅ File {file_path} caricato con successo")
        print(f"📋 Fogli disponibili: {wb.sheetnames}")
        
        # Analizza il foglio Calcoli
        if 'Calcoli' in wb.sheetnames:
            ws_calcoli = wb['Calcoli']
            print(f"\n🔍 ANALISI FOGLIO 'Calcoli'")
            print(f"Dimensioni: {ws_calcoli.max_row} righe x {ws_calcoli.max_column} colonne")
            
            # Analizza area intorno alla riga 99
            print(f"\n📊 ANALISI AREA RIGA 99:")
            
            # Esamina le righe 95-105 per contesto
            for row in range(95, 106):
                row_data = []
                for col in range(1, min(42, ws_calcoli.max_column + 1)):  # A-AO = 41 colonne
                    cell = ws_calcoli.cell(row=row, column=col)
                    if cell.value is not None:
                        col_letter = get_column_letter(col)
                        if isinstance(cell.value, str) and len(cell.value) > 50:
                            value_preview = cell.value[:50] + "..."
                        else:
                            value_preview = str(cell.value)
                        row_data.append(f"{col_letter}{row}: {value_preview}")
                
                if row_data:
                    print(f"Riga {row}: {' | '.join(row_data[:5])}")  # Mostra prime 5 celle non vuote
            
            # Cerca matrici Stock GBV
            print(f"\n🔍 RICERCA MATRICI STOCK GBV:")
            stock_gbv_found = []
            for row in range(1, ws_calcoli.max_row + 1):
                for col in range(1, ws_calcoli.max_column + 1):
                    cell = ws_calcoli.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if 'stock' in cell.value.lower() and 'gbv' in cell.value.lower():
                            col_letter = get_column_letter(col)
                            stock_gbv_found.append(f"{col_letter}{row}: {cell.value}")
            
            if stock_gbv_found:
                for item in stock_gbv_found[:10]:  # Mostra prime 10 occorrenze
                    print(f"  📍 {item}")
            else:
                print("  ❌ Nessuna matrice Stock GBV trovata con questa ricerca")
            
            # Cerca Erogazioni e Rimborsi
            print(f"\n🔍 RICERCA EROGAZIONI E RIMBORSI:")
            erogazioni_rimborsi = []
            for row in range(1, ws_calcoli.max_row + 1):
                for col in range(1, ws_calcoli.max_column + 1):
                    cell = ws_calcoli.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        value_lower = cell.value.lower()
                        if 'erogazioni' in value_lower or 'rimborsi' in value_lower:
                            col_letter = get_column_letter(col)
                            erogazioni_rimborsi.append(f"{col_letter}{row}: {cell.value}")
            
            if erogazioni_rimborsi:
                for item in erogazioni_rimborsi[:15]:  # Mostra prime 15 occorrenze
                    print(f"  📍 {item}")
            else:
                print("  ❌ Nessuna cella con Erogazioni/Rimborsi trovata")
            
            # Analizza struttura colonne B-AO nella riga 99
            print(f"\n📊 STRUTTURA COLONNE B-AO RIGA 99:")
            row_99_structure = []
            for col in range(2, 42):  # B=2, AO=41
                cell = ws_calcoli.cell(row=99, column=col)
                col_letter = get_column_letter(col)
                if cell.value is not None:
                    if isinstance(cell.value, str) and len(cell.value) > 30:
                        value_preview = cell.value[:30] + "..."
                    else:
                        value_preview = str(cell.value)
                    row_99_structure.append(f"{col_letter}: {value_preview}")
                else:
                    row_99_structure.append(f"{col_letter}: [vuota]")
            
            # Mostra la struttura a gruppi di 5 colonne
            for i in range(0, len(row_99_structure), 5):
                group = row_99_structure[i:i+5]
                print(f"  {' | '.join(group)}")
            
            # Cerca formule esistenti come riferimento
            print(f"\n🔍 FORMULE ESISTENTI COME RIFERIMENTO:")
            formule_trovate = []
            for row in range(1, min(200, ws_calcoli.max_row + 1)):  # Limita a prime 200 righe per performance
                for col in range(1, min(50, ws_calcoli.max_column + 1)):  # Limita a prime 50 colonne
                    cell = ws_calcoli.cell(row=row, column=col)
                    if cell.data_type == 'f':  # Formula
                        col_letter = get_column_letter(col)
                        formula = str(cell.value)
                        if len(formula) > 100:
                            formula = formula[:100] + "..."
                        formule_trovate.append(f"{col_letter}{row}: {formula}")
            
            if formule_trovate:
                print(f"  📍 Trovate {len(formule_trovate)} formule. Prime 10:")
                for formula in formule_trovate[:10]:
                    print(f"    {formula}")
            else:
                print("  ❌ Nessuna formula trovata nell'area analizzata")
        
        else:
            print(f"❌ Foglio 'Calcoli' non trovato!")
        
        # Analizza altri fogli per completezza
        print(f"\n📋 ANALISI ALTRI FOGLI:")
        for sheet_name in wb.sheetnames:
            if sheet_name != 'Calcoli':
                ws = wb[sheet_name]
                print(f"  📄 {sheet_name}: {ws.max_row} righe x {ws.max_column} colonne")
                
                # Cerca celle con contenuto interessante
                interesting_cells = []
                for row in range(1, min(50, ws.max_row + 1)):
                    for col in range(1, min(20, ws.max_column + 1)):
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            value_lower = cell.value.lower()
                            if any(keyword in value_lower for keyword in ['stock', 'gbv', 'erogazioni', 'rimborsi']):
                                col_letter = get_column_letter(col)
                                interesting_cells.append(f"{col_letter}{row}: {cell.value}")
                
                if interesting_cells:
                    print(f"    🔍 Celle interessanti: {interesting_cells[:3]}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Errore durante l'analisi: {str(e)}")

if __name__ == "__main__":
    analizza_modello_excel()
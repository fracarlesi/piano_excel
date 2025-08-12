#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STEP 2: Script per aggiungere il foglio Modello TRIMESTRALE con tutte le tabelle di calcolo
al file modello_bancario_completo.xlsx esistente.
Supporta 20 trimestri (Q1-Q20) invece di 5 anni.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_model_sheet():
    """Aggiunge il foglio Modello trimestrale al file Excel esistente con tutte le tabelle"""
    
    # Carica il workbook esistente
    filename = 'modello_bancario_completo.xlsx'
    print(f"📂 Caricamento file esistente: {filename}")
    wb = load_workbook(filename)
    
    # Crea il foglio Modello
    if 'Modello' in wb.sheetnames:
        wb.remove(wb['Modello'])
    ws = wb.create_sheet('Modello')
    print("📊 Creazione foglio 'Modello' trimestrale...")
    
    # Imposta larghezza colonne per 20 trimestri
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 45
    # Colonne C-V per i 20 trimestri (Q1-Q20)
    for col_idx in range(3, 23):  # C to V (20 colonne)
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    ws.column_dimensions['W'].width = 5  # Spazio
    ws.column_dimensions['X'].width = 50  # Descrizione
    
    # Stili
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Stili per distinzione INPUT/FORMULA
    formula_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Grigio chiaro per FORMULA
    formula_font = Font(name='Calibri', size=11, color="000000")  # Nero per testo FORMULA
    formula_border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='dashed'),
        bottom=Side(style='dashed')
    )
    
    # Funzione helper per applicare formattazione FORMULA
    def format_formula_cell(cell, number_format=None):
        """Applica la formattazione delle celle con FORMULA"""
        cell.fill = formula_fill
        cell.font = formula_font
        cell.border = formula_border
        cell.alignment = Alignment(horizontal='center')
        if number_format:
            cell.number_format = number_format
    
    # Funzione per creare headers trimestrali
    def create_quarterly_headers(row):
        """Crea headers per i 20 trimestri (Q1-Q20)"""
        # Header "Prodotto/Voce"
        cell = ws.cell(row=row, column=2, value="Prodotto")
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = thin_border
        
        # Headers Q1-Q20
        for i in range(20):
            col_idx = 3 + i
            quarter = i + 1
            year = (i // 4) + 1
            quarter_in_year = (i % 4) + 1
            header_text = f"Q{quarter}\n(Y{year}Q{quarter_in_year})"
            
            cell = ws.cell(row=row, column=col_idx, value=header_text)
            cell.font = Font(name='Calibri', size=10, bold=True)
            cell.fill = section_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Header "Descrizione"
        cell = ws.cell(row=row, column=24, value="Descrizione")  # Colonna X
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = thin_border
    
    current_row = 2
    
    # ===== SEZIONE 1: CALCOLI DI APPOGGIO =====
    print("  📈 Creazione sezione: Calcoli di Appoggio...")
    
    # Titolo principale
    ws.cell(row=current_row, column=2, value="2. CALCOLI DI APPOGGIO - MOTORE DI CALCOLO TRIMESTRALE")
    ws.merge_cells(f'B{current_row}:X{current_row}')
    ws.cell(row=current_row, column=2).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill = header_fill
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Definizione prodotti
    re_products = ["CBL", "MLA", "MLAM", "MLPA", "MLPAM", "NRE"]
    sme_products = ["BL", "REFI", "SS", "NF", "RES"]
    pg_products = ["ACFP", "FGA", "FGPA"]
    
    # RIMOSSE TABELLE 1.1 E 1.2 - PARTIAMO DIRETTAMENTE CON LE MATRICI VINTAGE
    # Le tabelle aggregate saranno somme delle matrici vintage
    
    # 1.1 MATRICI VINTAGE - BLOCCO A: EROGAZIONI E VOLUMI
    ws.cell(row=current_row, column=2, value="1.1 MATRICI VINTAGE - BLOCCO A: EROGAZIONI E VOLUMI (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Matrici 20x20 per ogni prodotto: Erogazioni, Rimborsi, Stock Lordo (GBV)")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # BLOCCO A - Per ogni prodotto, creiamo matrici vintage 20x20
    all_products = [
        ("RE", ["CBL", "MLA", "MLAM", "MLPA", "MLPAM", "NRE"]),
        ("SME", ["BL", "REFI", "SS", "NF", "RES"]),
        ("PG", ["ACFP", "FGA", "FGPA"])
    ]
    
    # Helper function per creare matrice vintage 20x20
    def create_vintage_matrix(title, product_name, color_hex):
        nonlocal current_row
        
        ws.cell(row=current_row, column=2, value=f"{title} - {product_name}")
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=10)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        current_row += 1
        
        # Headers vintage
        ws.cell(row=current_row, column=2, value="Vintage\\Time").border = thin_border
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=9)
        ws.cell(row=current_row, column=23, value="Total").border = thin_border  # Colonna totali
        ws.cell(row=current_row, column=23).font = Font(bold=True, size=9)
        
        for q in range(1, 21):
            cell = ws.cell(row=current_row, column=2+q, value=f"Q{q}")
            cell.border = thin_border
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Righe per ogni vintage V1(Q1) - V20(Q20)
        for v in range(1, 21):
            year = ((v-1) // 4) + 1
            quarter_in_year = ((v-1) % 4) + 1
            vintage_label = f"V{v}(Q{v})"
            
            ws.cell(row=current_row, column=2, value=vintage_label).border = thin_border
            ws.cell(row=current_row, column=2).font = Font(bold=True, size=9)
            
            # Celle dati per Q1-Q20
            for q in range(1, 21):
                cell = ws.cell(row=current_row, column=2+q)
                format_formula_cell(cell, '#,##0.0')
            
            # Cella totale per vintage
            cell = ws.cell(row=current_row, column=23)
            format_formula_cell(cell, '#,##0.0')
            current_row += 1
        
        # Riga totali
        ws.cell(row=current_row, column=2, value="TOTAL").border = thin_border
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=9)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for q in range(1, 21):
            cell = ws.cell(row=current_row, column=2+q)
            format_formula_cell(cell, '#,##0.0')
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Totale generale
        cell = ws.cell(row=current_row, column=23)
        format_formula_cell(cell, '#,##0.0')
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 2  # Spazio tra matrici
    
    # A.1 - MATRICE EROGAZIONI PER VINTAGE
    ws.cell(row=current_row, column=2, value="A.1 - MATRICE EROGAZIONI PER VINTAGE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - EROGAZIONI", product, "E6F3FF")
    
    # A.2 - MATRICE RIMBORSI PER VINTAGE
    ws.cell(row=current_row, column=2, value="A.2 - MATRICE RIMBORSI PER VINTAGE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - RIMBORSI", product, "FFE6CC")
    
    # A.3 - MATRICE STOCK LORDO (GBV) PER VINTAGE
    ws.cell(row=current_row, column=2, value="A.3 - MATRICE STOCK LORDO (GBV) PER VINTAGE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - STOCK GBV", product, "E6FFE6")
    
    current_row += 2
    
    # 1.2 MATRICI VINTAGE - BLOCCO B: RISCHIO CREDITO
    ws.cell(row=current_row, column=2, value="1.2 MATRICI VINTAGE - BLOCCO B: RISCHIO CREDITO (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Matrici 20x20 per ogni prodotto: Default, Stock Performing, Stock NPL")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # B.1 - MATRICE DEFAULT PER VINTAGE (CON TIMING DISTRIBUTION)
    ws.cell(row=current_row, column=2, value="B.1 - MATRICE DEFAULT PER VINTAGE (CON TIMING DISTRIBUTION)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - DEFAULT", product, "FFE6E6")
    
    # B.2 - MATRICE STOCK PERFORMING
    ws.cell(row=current_row, column=2, value="B.2 - MATRICE STOCK PERFORMING")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - PERFORMING", product, "E6F3FF")
    
    # B.3 - MATRICE STOCK NPL
    ws.cell(row=current_row, column=2, value="B.3 - MATRICE STOCK NPL")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - NPL", product, "FFCCCC")
    
    current_row += 2
    
    # 1.3 MATRICI VINTAGE - BLOCCO C: ECL E VALUTAZIONI
    ws.cell(row=current_row, column=2, value="1.3 MATRICI VINTAGE - BLOCCO C: ECL E VALUTAZIONI (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Matrici 20x20: ECL Stage 1, NBV Performing, NPV Recuperi, NBV Non-Performing")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # C.1 - MATRICE ECL STAGE 1 (FORMULA: Stock × Danger Rate × LGD)
    ws.cell(row=current_row, column=2, value="C.1 - MATRICE ECL STAGE 1 (Stock × Danger Rate × LGD)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - ECL STAGE 1", product, "FFE6CC")
    
    # C.2 - MATRICE NBV PERFORMING (Stock - ECL)
    ws.cell(row=current_row, column=2, value="C.2 - MATRICE NBV PERFORMING (Stock - ECL)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - NBV PERFORMING", product, "CCE6FF")
    
    # C.3 - MATRICE NPV RECUPERI (ATTUALIZZATI CON EURIBOR + SPREAD)
    ws.cell(row=current_row, column=2, value="C.3 - MATRICE NPV RECUPERI (Attualizzati con Euribor + Spread)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - NPV RECUPERI", product, "E6FFE6")
    
    # C.4 - MATRICE NBV NON-PERFORMING
    ws.cell(row=current_row, column=2, value="C.4 - MATRICE NBV NON-PERFORMING")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - NBV NPL", product, "FFCCCC")
    
    current_row += 2
    
    # 1.4 MATRICI VINTAGE - BLOCCO D: RICAVI
    ws.cell(row=current_row, column=2, value="1.4 MATRICI VINTAGE - BLOCCO D: RICAVI (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Matrici 20x20: Interessi Attivi, Interessi di Mora, Commissioni Up-front")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # D.1 - MATRICE INTERESSI ATTIVI SU PERFORMING
    ws.cell(row=current_row, column=2, value="D.1 - MATRICE INTERESSI ATTIVI SU PERFORMING")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - INT ATTIVI", product, "E6F3FF")
    
    # D.2 - MATRICE INTERESSI DI MORA SU NPL
    ws.cell(row=current_row, column=2, value="D.2 - MATRICE INTERESSI DI MORA SU NPL")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - INT MORA", product, "FFE6E6")
    
    # D.3 - MATRICE COMMISSIONI UP-FRONT
    ws.cell(row=current_row, column=2, value="D.3 - MATRICE COMMISSIONI UP-FRONT")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    current_row += 1
    
    for division, products in all_products:
        for product in products:
            create_vintage_matrix(f"{division} - COMM UP-FRONT", product, "E6FFE6")
    
    current_row += 3
    
    # 1.5 MATRICI VINTAGE - BLOCCO E: COSTI E ACCANTONAMENTI
    ws.cell(row=current_row, column=2, value="1.5 MATRICI VINTAGE - BLOCCO E: COSTI E ACCANTONAMENTI (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Tabelle LLP (Loan Loss Provisions) e Write-off per trimestre")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # E.1 - TABELLA LLP (LOAN LOSS PROVISIONS) PER TRIMESTRE
    ws.cell(row=current_row, column=2, value="E.1 - TABELLA LLP (LOAN LOSS PROVISIONS) PER TRIMESTRE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    current_row += 1
    
    # Headers per LLP
    create_quarterly_headers(current_row)
    current_row += 1
    
    # LLP per prodotto
    for division, products in all_products:
        for product in products:
            ws.cell(row=current_row, column=2, value=f"{division} - {product} - LLP").border = thin_border
            for col in range(3, 23):  # C to V (Q1-Q20)
                cell = ws.cell(row=current_row, column=col)
                format_formula_cell(cell, '#,##0.0')
            current_row += 1
    
    # Totale LLP
    ws.cell(row=current_row, column=2, value="TOTALE LLP").border = thin_border
    ws.cell(row=current_row, column=2).font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col in range(3, 23):  # C to V (Q1-Q20)
        cell = ws.cell(row=current_row, column=col)
        format_formula_cell(cell, '#,##0.0')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    current_row += 2
    
    # E.2 - TABELLA WRITE-OFF PER TRIMESTRE
    ws.cell(row=current_row, column=2, value="E.2 - TABELLA WRITE-OFF PER TRIMESTRE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    current_row += 1
    
    # Headers per Write-off
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Write-off per prodotto
    for division, products in all_products:
        for product in products:
            ws.cell(row=current_row, column=2, value=f"{division} - {product} - Write-off").border = thin_border
            for col in range(3, 23):  # C to V (Q1-Q20)
                cell = ws.cell(row=current_row, column=col)
                format_formula_cell(cell, '#,##0.0')
            current_row += 1
    
    # Totale Write-off
    ws.cell(row=current_row, column=2, value="TOTALE WRITE-OFF").border = thin_border
    ws.cell(row=current_row, column=2).font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col in range(3, 23):  # C to V (Q1-Q20)
        cell = ws.cell(row=current_row, column=col)
        format_formula_cell(cell, '#,##0.0')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    current_row += 2
    
    # 1.6 MATRICI VINTAGE - BLOCCO F: AGGREGAZIONI DIVISIONALI
    ws.cell(row=current_row, column=2, value="1.6 MATRICI VINTAGE - BLOCCO F: AGGREGAZIONI DIVISIONALI (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Riepiloghi divisionali: Real Estate (6 prodotti), SME (5 prodotti), Public Guarantee (3 prodotti)")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # F.1 - RIEPILOGO REAL ESTATE (SOMMA 6 PRODOTTI)
    ws.cell(row=current_row, column=2, value="F.1 - RIEPILOGO REAL ESTATE DIVISION (SOMMA 6 PRODOTTI)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Voci di riepilogo Real Estate
    re_summary_items = [
        "RE - TOTALE EROGAZIONI",
        "RE - TOTALE STOCK CREDITI",
        "RE - TOTALE INTERESSI ATTIVI",
        "RE - TOTALE COMMISSIONI",
        "RE - TOTALE LLP",
        "RE - TOTALE WRITE-OFF"
    ]
    
    for item in re_summary_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
        current_row += 1
    
    current_row += 1
    
    # F.2 - RIEPILOGO SME (SOMMA 5 PRODOTTI)
    ws.cell(row=current_row, column=2, value="F.2 - RIEPILOGO SME DIVISION (SOMMA 5 PRODOTTI)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Voci di riepilogo SME
    sme_summary_items = [
        "SME - TOTALE EROGAZIONI",
        "SME - TOTALE STOCK CREDITI",
        "SME - TOTALE INTERESSI ATTIVI",
        "SME - TOTALE COMMISSIONI",
        "SME - TOTALE LLP",
        "SME - TOTALE WRITE-OFF"
    ]
    
    for item in sme_summary_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
        current_row += 1
    
    current_row += 1
    
    # F.3 - RIEPILOGO PUBLIC GUARANTEE (SOMMA 3 PRODOTTI)
    ws.cell(row=current_row, column=2, value="F.3 - RIEPILOGO PUBLIC GUARANTEE DIVISION (SOMMA 3 PRODOTTI)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Voci di riepilogo Public Guarantee
    pg_summary_items = [
        "PG - TOTALE EROGAZIONI",
        "PG - TOTALE STOCK CREDITI",
        "PG - TOTALE INTERESSI ATTIVI",
        "PG - TOTALE COMMISSIONI",
        "PG - TOTALE LLP",
        "PG - TOTALE WRITE-OFF"
    ]
    
    for item in pg_summary_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
        current_row += 1
    
    current_row += 2
    
    # 1.7 MATRICI VINTAGE - BLOCCO G: FUNDING E LIQUIDITÀ
    ws.cell(row=current_row, column=2, value="1.7 MATRICI VINTAGE - BLOCCO G: FUNDING E LIQUIDITÀ (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Evoluzione trimestrale depositi, funding wholesale e liquidità BCE")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # G.1 - TABELLA DEPOSITI CLIENTELA (EVOLUZIONE TRIMESTRALE)
    ws.cell(row=current_row, column=2, value="G.1 - TABELLA DEPOSITI CLIENTELA (EVOLUZIONE TRIMESTRALE)")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Depositi per segmento
    deposit_items = [
        "Digital Banking - Depositi a Vista",
        "Digital Banking - Depositi Vincolati",
        "Wealth Management - Depositi Gestiti",
        "Corporate - Depositi Operativi",
        "Corporate - Depositi Vincolati",
        "TOTALE DEPOSITI CLIENTELA"
    ]
    
    for item in deposit_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        is_total = "TOTALE" in item
        if is_total:
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
    
    current_row += 1
    
    # G.2 - TABELLA FUNDING WHOLESALE
    ws.cell(row=current_row, column=2, value="G.2 - TABELLA FUNDING WHOLESALE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Funding wholesale
    funding_items = [
        "Funding BCE - TLTRO",
        "Funding BCE - Operazioni Principali",
        "Senior Debt Securities",
        "Subordinated Debt",
        "Interbank Funding",
        "TOTALE FUNDING WHOLESALE"
    ]
    
    for item in funding_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        is_total = "TOTALE" in item
        if is_total:
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
    
    current_row += 1
    
    # G.3 - TABELLA LIQUIDITÀ E RISERVE BCE
    ws.cell(row=current_row, column=2, value="G.3 - TABELLA LIQUIDITÀ E RISERVE BCE")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Liquidità
    liquidity_items = [
        "Cassa e Disponibilità Liquide",
        "Riserva Obbligatoria BCE",
        "Depositi presso BCE",
        "Titoli Eligible per Operazioni BCE",
        "LCR Buffer",
        "TOTALE LIQUIDITÀ DISPONIBILE"
    ]
    
    for item in liquidity_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        is_total = "TOTALE" in item
        if is_total:
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
    
    current_row += 2
    
    # 1.8 MATRICI VINTAGE - BLOCCO H: ALTRI PATRIMONIALI
    ws.cell(row=current_row, column=2, value="1.8 MATRICI VINTAGE - BLOCCO H: ALTRI PATRIMONIALI (€ mln)")
    ws.cell(row=current_row, column=2).font = Font(size=12, bold=True)
    ws.cell(row=current_row, column=2).fill = subheader_fill
    ws.cell(row=current_row, column=2).font = Font(color="FFFFFF", bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="Roll-forward immobilizzazioni, portafoglio titoli, evoluzione patrimonio netto")
    ws.cell(row=current_row, column=2).font = Font(italic=True, size=10)
    current_row += 1
    
    # H.1 - ROLL-FORWARD IMMOBILIZZAZIONI
    ws.cell(row=current_row, column=2, value="H.1 - ROLL-FORWARD IMMOBILIZZAZIONI")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Roll-forward immobilizzazioni
    capex_items = [
        "Stock Iniziale Immobilizzazioni",
        "Nuovi Investimenti IT/Software",
        "Nuovi Investimenti Immobili",
        "Ammortamenti Periodo",
        "Dismissioni/Vendite",
        "Stock Finale Immobilizzazioni Nette"
    ]
    
    for item in capex_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        is_total = "Stock Finale" in item
        if is_total:
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
    
    current_row += 1
    
    # H.2 - PORTAFOGLIO TITOLI
    ws.cell(row=current_row, column=2, value="H.2 - PORTAFOGLIO TITOLI")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Portafoglio titoli
    securities_items = [
        "Titoli di Stato Italiani",
        "Titoli di Stato Esteri",
        "Titoli Corporate Investment Grade",
        "Titoli Corporate High Yield",
        "Altri Strumenti Finanziari",
        "TOTALE PORTAFOGLIO TITOLI"
    ]
    
    for item in securities_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        is_total = "TOTALE" in item
        if is_total:
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
    
    current_row += 1
    
    # H.3 - EVOLUZIONE PATRIMONIO NETTO
    ws.cell(row=current_row, column=2, value="H.3 - EVOLUZIONE PATRIMONIO NETTO")
    ws.cell(row=current_row, column=2).font = Font(size=11, bold=True)
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    current_row += 1
    
    # Headers
    create_quarterly_headers(current_row)
    current_row += 1
    
    # Patrimonio netto
    equity_items = [
        "Patrimonio Netto Iniziale",
        "Utile/Perdita Netto Periodo",
        "Aumenti di Capitale",
        "Distribuzione Dividendi",
        "Altri Movimenti",
        "Patrimonio Netto Finale"
    ]
    
    for item in equity_items:
        ws.cell(row=current_row, column=2, value=item).border = thin_border
        is_total = "Finale" in item
        if is_total:
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(3, 23):  # C to V (Q1-Q20)
            cell = ws.cell(row=current_row, column=col)
            format_formula_cell(cell, '#,##0.0')
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
    
    current_row += 3
    
    # ===== SEZIONE 2: CONTO ECONOMICO =====
    print("  💰 Creazione sezione: Conto Economico Consolidato...")
    
    ws.cell(row=current_row, column=2, value="3. CONTO ECONOMICO CONSOLIDATO TRIMESTRALE (€ mln)")
    ws.merge_cells(f'B{current_row}:X{current_row}')
    ws.cell(row=current_row, column=2).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill = header_fill
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers trimestrali per CE
    ws.cell(row=current_row, column=2, value="Voce").font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = section_fill
    ws.cell(row=current_row, column=2).border = thin_border
    
    for i in range(20):
        col_idx = 3 + i
        quarter = i + 1
        year = (i // 4) + 1
        quarter_in_year = (i % 4) + 1
        header_text = f"Q{quarter}\n(Y{year}Q{quarter_in_year})"
        
        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.cell(row=current_row, column=24, value="Descrizione").font = Font(bold=True)
    ws.cell(row=current_row, column=24).fill = section_fill
    ws.cell(row=current_row, column=24).border = thin_border
    current_row += 1
    
    # Voci del Conto Economico con indentazione
    ce_items = [
        ("1. MARGINE DI INTERESSE", True, 0),
        ("1.1. Interest Income", False, 1),
        ("    1.1.1. da Real Estate Division", False, 2),
        ("    1.1.2. da SME Division", False, 2),
        ("    1.1.3. da Public Guarantee Division", False, 2),
        ("    1.1.4. da Digital Banking Division", False, 2),
        ("    1.1.5. da Treasury Division", False, 2),
        ("1.2. Interest Expenses", False, 1),
        ("1.3. NET INTEREST INCOME", True, 1),
        ("", False, 0),
        ("2. COMMISSIONI", True, 0),
        ("2.1. Commission Income", False, 1),
        ("    2.1.1. da Real Estate Division", False, 2),
        ("    2.1.2. da SME Division", False, 2),
        ("    2.1.3. da Public Guarantee Division", False, 2),
        ("    2.1.4. da Wealth Management Division", False, 2),
        ("    2.1.5. da Digital Banking Division", False, 2),
        ("2.2. Commission Expenses", False, 1),
        ("2.3. NET COMMISSION INCOME", True, 1),
        ("", False, 0),
        ("3. TOTAL REVENUES", True, 0),
        ("", False, 0),
        ("4. COSTI OPERATIVI", True, 0),
        ("4.1. Staff Costs", False, 1),
        ("    4.1.1. da Real Estate Division", False, 2),
        ("    4.1.2. da SME Division", False, 2),
        ("    4.1.3. da Public Guarantee Division", False, 2),
        ("    4.1.4. da Digital Banking Division", False, 2),
        ("    4.1.5. da Wealth Management Division", False, 2),
        ("    4.1.6. da Technology Division", False, 2),
        ("    4.1.7. da Central Functions", False, 2),
        ("4.2. Other Operating Costs", False, 1),
        ("4.3. TOTAL OPERATING COSTS", True, 1),
        ("", False, 0),
        ("5. GROSS OPERATING PROFIT", True, 0),
        ("", False, 0),
        ("6. Ammortamenti", False, 0),
        ("7. Loan Loss Provisions", False, 0),
        ("    7.1. da Real Estate Division", False, 1),
        ("    7.2. da SME Division", False, 1),
        ("    7.3. da Public Guarantee Division", False, 1),
        ("", False, 0),
        ("8. OPERATING PROFIT", True, 0),
        ("", False, 0),
        ("9. Taxes", False, 0),
        ("10. NET PROFIT", True, 0),
    ]
    
    for item_name, is_bold, indent_level in ce_items:
        if item_name:  # Skip empty rows
            cell = ws.cell(row=current_row, column=2, value=item_name)
            cell.border = thin_border
            if is_bold:
                cell.font = Font(bold=True)
                if indent_level == 0:
                    cell.fill = section_fill
            
            # Add value cells for all 20 quarters
            for col in range(3, 23):  # C to V (Q1-Q20)
                cell = ws.cell(row=current_row, column=col, value=0)
                if is_bold:
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                else:
                    format_formula_cell(cell, '#,##0.0')
            
            # Description column
            ws.cell(row=current_row, column=24, value="").border = thin_border
        
        current_row += 1
    
    current_row += 3
    
    # ===== SEZIONE 3: STATO PATRIMONIALE =====
    print("  🏦 Creazione sezione: Stato Patrimoniale Consolidato...")
    
    ws.cell(row=current_row, column=2, value="4. STATO PATRIMONIALE CONSOLIDATO TRIMESTRALE (€ mln)")
    ws.merge_cells(f'B{current_row}:X{current_row}')
    ws.cell(row=current_row, column=2).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill = header_fill
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers trimestrali per SP
    ws.cell(row=current_row, column=2, value="Voce").font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = section_fill
    ws.cell(row=current_row, column=2).border = thin_border
    
    for i in range(20):
        col_idx = 3 + i
        quarter = i + 1
        year = (i // 4) + 1
        quarter_in_year = (i % 4) + 1
        header_text = f"Q{quarter}\n(Y{year}Q{quarter_in_year})"
        
        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.cell(row=current_row, column=24, value="Descrizione").font = Font(bold=True)
    ws.cell(row=current_row, column=24).fill = section_fill
    ws.cell(row=current_row, column=24).border = thin_border
    current_row += 1
    
    # Voci dello Stato Patrimoniale
    sp_items = [
        ("1. ATTIVO", True, 0),
        ("1.1. Loans to Customers (Gross)", False, 1),
        ("    1.1.1. Real Estate Division loans", False, 2),
        ("    1.1.2. SME Division loans", False, 2),
        ("    1.1.3. Public Guarantee Division loans", False, 2),
        ("1.2. Loan Loss Reserves", False, 1),
        ("1.3. Loans to Customers (Net)", True, 1),
        ("1.4. Securities Portfolio", False, 1),
        ("1.5. Immobilizzazioni Immateriali (Net)", False, 1),
        ("1.6. Cash & Central Banks", False, 1),
        ("1.7. Other Assets", False, 1),
        ("1.8. TOTAL ASSETS", True, 1),
        ("", False, 0),
        ("2. PASSIVO E PATRIMONIO NETTO", True, 0),
        ("2.1. Customer Deposits", False, 1),
        ("    2.1.1. Digital Banking Division deposits", False, 2),
        ("    2.1.2. Wealth Management Division deposits", False, 2),
        ("    2.1.3. Corporate deposits", False, 2),
        ("2.2. Debt Securities & Funding", False, 1),
        ("2.3. Other Liabilities", False, 1),
        ("2.4. TOTAL LIABILITIES", True, 1),
        ("", False, 0),
        ("2.5. Share Capital & Reserves", False, 1),
        ("2.6. Retained Earnings", False, 1),
        ("2.7. TOTAL EQUITY", True, 1),
        ("2.8. TOTAL LIABILITIES & EQUITY", True, 1),
        ("", False, 0),
        ("2.9. CHECK DI QUADRATURA", True, 0),
    ]
    
    for item_name, is_bold, indent_level in sp_items:
        if item_name:
            cell = ws.cell(row=current_row, column=2, value=item_name)
            cell.border = thin_border
            if is_bold:
                cell.font = Font(bold=True)
                if indent_level == 0:
                    cell.fill = section_fill
            
            # Add value cells for all 20 quarters
            for col in range(3, 23):  # C to V (Q1-Q20)
                cell = ws.cell(row=current_row, column=col, value=0)
                if is_bold:
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                else:
                    format_formula_cell(cell, '#,##0.0')
            
            ws.cell(row=current_row, column=24, value="").border = thin_border
        
        current_row += 1
    
    current_row += 3
    
    # ===== SEZIONE 4: CAPITAL REQUIREMENTS =====
    print("  🏛️ Creazione sezione: Capital Requirements...")
    
    ws.cell(row=current_row, column=2, value="5. CAPITAL REQUIREMENTS TRIMESTRALI (€ mln)")
    ws.merge_cells(f'B{current_row}:X{current_row}')
    ws.cell(row=current_row, column=2).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill = header_fill
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers trimestrali per Capital Requirements
    ws.cell(row=current_row, column=2, value="Voce").font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = section_fill
    ws.cell(row=current_row, column=2).border = thin_border
    
    for i in range(20):
        col_idx = 3 + i
        quarter = i + 1
        year = (i // 4) + 1
        quarter_in_year = (i % 4) + 1
        header_text = f"Q{quarter}\n(Y{year}Q{quarter_in_year})"
        
        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.cell(row=current_row, column=24, value="Descrizione").font = Font(bold=True)
    ws.cell(row=current_row, column=24).fill = section_fill
    ws.cell(row=current_row, column=24).border = thin_border
    current_row += 1
    
    # Voci Capital Requirements
    cap_items = [
        ("1. RISK WEIGHTED ASSETS (RWA)", True, 0),
        ("1.1. Credit Risk RWA", False, 1),
        ("    1.1.1. da Real Estate Division", False, 2),
        ("    1.1.2. da SME Division", False, 2),
        ("    1.1.3. da Public Guarantee Division", False, 2),
        ("1.2. Market Risk RWA", False, 1),
        ("1.3. Operational Risk RWA", False, 1),
        ("1.4. TOTAL RWA", True, 1),
        ("", False, 0),
        ("2. CAPITAL", True, 0),
        ("2.1. CET1 Capital", False, 1),
        ("2.2. AT1 Capital", False, 1),
        ("2.3. Tier 2 Capital", False, 1),
        ("2.4. TOTAL CAPITAL", True, 1),
        ("", False, 0),
        ("3. REQUIREMENTS & RATIOS", True, 0),
        ("3.1. CET1 Ratio", False, 1),
        ("3.2. Total Capital Ratio", False, 1),
        ("3.3. Total Requirement %", False, 1),
        ("3.4. Excess Capital (vs CET1)", False, 1),
    ]
    
    for item_name, is_bold, indent_level in cap_items:
        if item_name:
            cell = ws.cell(row=current_row, column=2, value=item_name)
            cell.border = thin_border
            if is_bold:
                cell.font = Font(bold=True)
                if indent_level == 0:
                    cell.fill = section_fill
            
            # Add value cells for all 20 quarters
            for col in range(3, 23):  # C to V (Q1-Q20)
                cell = ws.cell(row=current_row, column=col, value=0)
                if is_bold:
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                else:
                    format_formula_cell(cell, '#,##0.0')
            
            ws.cell(row=current_row, column=24, value="").border = thin_border
        
        current_row += 1
    
    current_row += 3
    
    # ===== SEZIONE 5: KEY PERFORMANCE INDICATORS =====
    print("  📊 Creazione sezione: Key Performance Indicators...")
    
    ws.cell(row=current_row, column=2, value="6. KEY PERFORMANCE INDICATORS TRIMESTRALI (KPI)")
    ws.merge_cells(f'B{current_row}:X{current_row}')
    ws.cell(row=current_row, column=2).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill = header_fill
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers trimestrali per KPI
    ws.cell(row=current_row, column=2, value="KPI").font = Font(bold=True)
    ws.cell(row=current_row, column=2).fill = section_fill
    ws.cell(row=current_row, column=2).border = thin_border
    
    for i in range(20):
        col_idx = 3 + i
        quarter = i + 1
        year = (i // 4) + 1
        quarter_in_year = (i % 4) + 1
        header_text = f"Q{quarter}\n(Y{year}Q{quarter_in_year})"
        
        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.cell(row=current_row, column=24, value="Descrizione").font = Font(bold=True)
    ws.cell(row=current_row, column=24).fill = section_fill
    ws.cell(row=current_row, column=24).border = thin_border
    current_row += 1
    
    # KPI items
    kpi_items = [
        ("1. PROFITABILITY - CONSOLIDATO", True, 0),
        ("1.1. ROE % (Annualizzato)", False, 1),
        ("1.2. ROA % (Annualizzato)", False, 1),
        ("1.3. NIM (Net Interest Margin) % (Annualizzato)", False, 1),
        ("1.4. Cost/Income % (Rolling 4Q)", False, 1),
        ("", False, 0),
        ("2. PROFITABILITY - PER DIVISIONE", True, 0),
        ("2.1. ROE % - Real Estate Division (Annualizzato)", False, 1),
        ("2.2. ROE % - SME Division (Annualizzato)", False, 1),
        ("2.3. ROE % - Public Guarantee Div. (Annualizzato)", False, 1),
        ("", False, 0),
        ("3. ASSET QUALITY", True, 0),
        ("3.1. NPL Ratio (Net) %", False, 1),
        ("3.2. Coverage Ratio %", False, 1),
        ("3.3. Cost of Risk (bps) (Annualizzato)", False, 1),
        ("", False, 0),
        ("4. LIQUIDITY & LEVERAGE", True, 0),
        ("4.1. LCR %", False, 1),
        ("4.2. NSFR %", False, 1),
        ("4.3. Loan/Deposit %", False, 1),
        ("4.4. Leverage Ratio %", False, 1),
    ]
    
    for item_name, is_bold, indent_level in kpi_items:
        if item_name:
            cell = ws.cell(row=current_row, column=2, value=item_name)
            cell.border = thin_border
            if is_bold:
                cell.font = Font(bold=True)
                if indent_level == 0:
                    cell.fill = section_fill
            
            # Add value cells for all 20 quarters
            for col in range(3, 23):  # C to V (Q1-Q20)
                cell = ws.cell(row=current_row, column=col, value=0)
                if is_bold:
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    # Formato percentuale per i KPI
                    if "%" in item_name or "bps" in item_name:
                        cell.number_format = '0.00%' if "%" in item_name else '0'
                else:
                    # Formato per celle FORMULA dei KPI
                    kpi_format = '0.00%' if "%" in item_name else ('0' if "bps" in item_name else '#,##0.0')
                    format_formula_cell(cell, kpi_format)
            
            ws.cell(row=current_row, column=24, value="").border = thin_border
        
        current_row += 1
    
    return wb

def main():
    """Funzione principale per STEP 2 - Modello Trimestrale"""
    print("📊 STEP 2: Aggiunta foglio Modello TRIMESTRALE con tabelle di calcolo...")
    print("⏱️ Supporta 20 trimestri (Q1-Q20) invece di 5 anni")
    print("=" * 70)
    
    try:
        # Aggiungi il foglio Modello trimestrale
        wb = add_model_sheet()
        
        # Salva il file aggiornato
        filename = 'modello_bancario_completo.xlsx'
        wb.save(filename)
        
        print("=" * 70)
        print(f"✅ STEP 2 COMPLETATO!")
        print(f"📁 File '{filename}' aggiornato con successo")
        print("📊 Il foglio 'Modello' contiene (VERSIONE TRIMESTRALE):")
        print("   - Calcoli di Appoggio (20 trimestri)")
        print("   - Conto Economico Consolidato (20 trimestri)")
        print("   - Stato Patrimoniale Consolidato (20 trimestri)")
        print("   - Capital Requirements (20 trimestri)")
        print("   - Key Performance Indicators (20 trimestri)")
        print("   - Headers: Q1, Q2, Q3... Q20 (con indicazione anno)")
        print("\n⏭️  Pronto per STEP 3: Inserimento formule Excel trimestrali")
        
    except FileNotFoundError:
        print("❌ ERRORE: File 'modello_bancario_completo.xlsx' non trovato!")
        print("   Assicurati di aver eseguito prima lo STEP 1 (create_input_sheet.py)")
    except Exception as e:
        print(f"❌ ERRORE: {str(e)}")

if __name__ == "__main__":
    main()
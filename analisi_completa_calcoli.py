#!/usr/bin/env python3
"""
Script per analisi completa del foglio Calcoli
Cerca tutte le celle con contenuto per identificare le matrici rimborsi
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re

def analisi_completa_calcoli():
    """Analizza tutto il contenuto del foglio Calcoli per trovare le matrici"""
    
    try:
        # Carica il workbook
        wb = openpyxl.load_workbook('modello.xlsx', data_only=False)
        
        if 'Calcoli' not in wb.sheetnames:
            print("ERRORE: Foglio 'Calcoli' non trovato!")
            return
        
        ws_calcoli = wb['Calcoli']
        
        print("🔍 SCANSIONE COMPLETA FOGLIO CALCOLI")
        print("=" * 50)
        
        # Trova tutte le celle con contenuto dalle righe 40 in poi
        celle_interessanti = []
        
        for row in range(40, 201):
            for col in range(1, 15):  # Colonne A-N
                cell = ws_calcoli.cell(row=row, column=col)
                
                if cell.value is not None:
                    celle_interessanti.append({
                        'cella': f"{get_column_letter(col)}{row}",
                        'riga': row,
                        'colonna': col,
                        'valore': cell.value,
                        'tipo': type(cell.value).__name__
                    })
        
        print(f"📊 Trovate {len(celle_interessanti)} celle con contenuto")
        print()
        
        # Raggruppa per righe per vedere la struttura
        righe_con_contenuto = {}
        for cella in celle_interessanti:
            riga = cella['riga']
            if riga not in righe_con_contenuto:
                righe_con_contenuto[riga] = []
            righe_con_contenuto[riga].append(cella)
        
        # Mostra solo le righe con contenuto significativo
        for riga in sorted(righe_con_contenuto.keys()):
            celle_riga = righe_con_contenuto[riga]
            
            # Cerca pattern che indicano matrici o prodotti
            testi_riga = [str(c['valore']) for c in celle_riga if isinstance(c['valore'], str)]
            
            if any('rimborsi' in testo.lower() or 'prodotto' in testo.lower() or 
                   'matrice' in testo.lower() or 'ammortamento' in testo.lower()
                   for testo in testi_riga):
                
                print(f"🎯 RIGA {riga} (POTENZIALE MATRICE):")
                for cella in celle_riga:
                    print(f"   {cella['cella']}: {cella['valore']} ({cella['tipo']})")
                print()
            
            # Mostra anche righe con molte formule (potrebbero essere matrici di calcolo)
            elif len([c for c in celle_riga if isinstance(c['valore'], str) and c['valore'].startswith('=')]) >= 3:
                print(f"📈 RIGA {riga} (FORMULE MULTIPLE):")
                for cella in celle_riga:
                    valore_mostrato = str(cella['valore'])[:50] + "..." if len(str(cella['valore'])) > 50 else str(cella['valore'])
                    print(f"   {cella['cella']}: {valore_mostrato}")
                print()
        
        # Cerca pattern specifici per "rimborsi"
        print("🔍 RICERCA SPECIFICA 'RIMBORSI':")
        print("=" * 40)
        
        for riga in sorted(righe_con_contenuto.keys()):
            celle_riga = righe_con_contenuto[riga]
            
            for cella in celle_riga:
                if isinstance(cella['valore'], str) and 'rimborsi' in cella['valore'].lower():
                    print(f"📌 TROVATO 'RIMBORSI' in {cella['cella']}: '{cella['valore']}'")
                    
                    # Cerca matrici nelle righe successive
                    print(f"   Analizzando righe successive...")
                    for offset in range(1, 11):
                        test_row = riga + offset
                        if test_row in righe_con_contenuto:
                            print(f"   Riga {test_row}: {len(righe_con_contenuto[test_row])} celle")
                            # Mostra le prime 3 celle della riga
                            for i, c in enumerate(righe_con_contenuto[test_row][:3]):
                                val_short = str(c['valore'])[:30] + "..." if len(str(c['valore'])) > 30 else str(c['valore'])
                                print(f"     {c['cella']}: {val_short}")
                    print()
        
        # Cerca anche pattern numerici che potrebbero essere inizio matrice
        print("🔍 RICERCA PATTERN NUMERICI (POSSIBILI MATRICI):")
        print("=" * 50)
        
        for riga in sorted(righe_con_contenuto.keys()):
            celle_riga = righe_con_contenuto[riga]
            
            # Se una riga ha molti valori numerici, potrebbe essere una matrice
            valori_numerici = [c for c in celle_riga if isinstance(c['valore'], (int, float))]
            formule = [c for c in celle_riga if isinstance(c['valore'], str) and c['valore'].startswith('=')]
            
            if len(valori_numerici) >= 5 or len(formule) >= 5:
                print(f"📊 RIGA {riga}: {len(valori_numerici)} numeri, {len(formule)} formule")
                
                # Mostra le prime celle
                for cella in celle_riga[:8]:
                    val_short = str(cella['valore'])[:40] + "..." if len(str(cella['valore'])) > 40 else str(cella['valore'])
                    print(f"   {cella['cella']}: {val_short}")
                print()
        
    except Exception as e:
        print(f"ERRORE: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analisi_completa_calcoli()
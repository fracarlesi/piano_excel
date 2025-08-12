#!/usr/bin/env python3
"""
Script per popolare tutte le matrici di erogazioni con formule Excel.
Inserisce formule sulla diagonale principale per tutti i 14 prodotti creditizi.
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

def format_formula_cell(cell):
    """Applica formattazione grigia per celle con formule"""
    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    cell.border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='dashed'),
        bottom=Side(style='dashed')
    )
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.number_format = '#,##0.0'

def find_matrix_row(ws, product_name, matrix_type="EROGAZIONI"):
    """Trova la riga di inizio di una matrice vintage"""
    for row in range(1, 500):
        cell_b = ws.cell(row=row, column=2).value
        if cell_b and product_name in str(cell_b) and matrix_type in str(cell_b).upper():
            # La matrice inizia 2 righe sotto (header + vintage labels)
            return row + 2
    return None

def populate_erogazioni_matrix(ws, division, product, mix_row, erog_row, start_search_row=1):
    """Popola la matrice erogazioni per un prodotto specifico"""
    
    # Cerca la matrice con il nome esatto
    search_pattern = f"{division} - EROGAZIONI - {product}"
    
    matrix_row = None
    for row in range(1, 500):  # Cerca in tutto il foglio
        cell_value = ws.cell(row=row, column=2).value
        if cell_value and search_pattern == str(cell_value).strip():
            matrix_row = row + 2  # Skip header
            break
    
    if not matrix_row:
        print(f"  ⚠️ Matrice non trovata per {division} - {product}")
        return start_search_row
    
    print(f"  ✅ Trovata matrice {product} alla riga {matrix_row-2}")
    
    # Mappa trimestre -> colonna anno nelle erogazioni
    quarter_to_year_col = {
        1: 'C', 2: 'C', 3: 'C', 4: 'C',    # Q1-Q4: Anno 1 (colonna C)
        5: 'D', 6: 'D', 7: 'D', 8: 'D',    # Q5-Q8: Anno 2 (colonna D)
        9: 'E', 10: 'E', 11: 'E', 12: 'E', # Q9-Q12: Anno 3 (colonna E)
        13: 'F', 14: 'F', 15: 'F', 16: 'F', # Q13-Q16: Anno 4 (colonna F)
        17: 'G', 18: 'G', 19: 'G', 20: 'G'  # Q17-Q20: Anno 5 (colonna G)
    }
    
    # Inserisci formule sulla diagonale
    formulas_inserted = 0
    for vintage in range(1, 21):
        row = matrix_row + vintage - 1
        col = 2 + vintage  # Colonna C=3 per V1, D=4 per V2, etc.
        
        # Determina la colonna dell'anno per questo trimestre
        year_col = quarter_to_year_col.get(vintage, 'C')
        
        # Formula: =Input!$[year_col]$[erog_row]*Input!$C$[mix_row]/4
        formula = f"=Input!${year_col}${erog_row}*Input!$C${mix_row}/4"
        
        # Inserisci formula
        cell = ws.cell(row=row, column=col)
        cell.value = formula
        format_formula_cell(cell)
        formulas_inserted += 1
        
        # Popola anche celle fuori diagonale con 0
        for other_col in range(3, 23):  # Colonne C-V
            if other_col != col:
                other_cell = ws.cell(row=row, column=other_col)
                if not other_cell.value:
                    other_cell.value = 0
                    format_formula_cell(other_cell)
    
    print(f"    → Inserite {formulas_inserted} formule sulla diagonale")
    
    # Riga TOTAL con somme
    total_row = matrix_row + 20
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    for col in range(3, 23):  # C-V
        col_letter = get_column_letter(col)
        start_row = matrix_row
        end_row = matrix_row + 19
        formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        cell = ws.cell(row=total_row, column=col)
        cell.value = formula
        format_formula_cell(cell)
        cell.font = Font(bold=True)
    
    return matrix_row + 25  # Ritorna la prossima riga di ricerca

def main():
    print("🚀 POPOLAMENTO FORMULE MATRICI EROGAZIONI")
    print("=" * 50)
    
    # Carica il file Excel
    filename = "modello_bancario_completo.xlsx"
    print(f"📁 Caricamento file: {filename}")
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb["Modello"]
        print("✅ File caricato correttamente")
    except Exception as e:
        print(f"❌ Errore caricamento file: {e}")
        return
    
    # Definizione prodotti e riferimenti
    products_config = {
        "RE": {
            "products": {
                "CBL": 107,
                "MLA": 108,
                "MLAM": 109,
                "MLPA": 110,
                "MLPAM": 111,
                "NRE": 112
            },
            "erog_row": 89
        },
        "SME": {
            "products": {
                "BL": 127,
                "REFI": 128,
                "SS": 129,
                "NF": 130,
                "RES": 131
            },
            "erog_row": 90
        },
        "PG": {
            "products": {
                "ACFP": 146,
                "FGA": 147,
                "FGPA": 148
            },
            "erog_row": 91
        }
    }
    
    # Popola matrici per ogni divisione
    print("\n📊 Inserimento formule per divisione:\n")
    
    search_row = 1
    total_products = 0
    
    for division, config in products_config.items():
        print(f"🏢 {division} Division:")
        
        for product, mix_row in config["products"].items():
            search_row = populate_erogazioni_matrix(
                ws, division, product, 
                mix_row, config["erog_row"],
                search_row
            )
            total_products += 1
        
        print()
    
    # Salva il file
    print(f"💾 Salvataggio file: {filename}")
    wb.save(filename)
    print("✅ File salvato correttamente")
    
    print("\n" + "=" * 50)
    print("📋 REPORT FINALE")
    print("=" * 50)
    print(f"✅ Popolate matrici erogazioni per {total_products} prodotti")
    print(f"📁 File aggiornato: {filename}")
    print("\n📊 Formule inserite:")
    print("  • Diagonale principale: Erogazioni trimestralizzate")
    print("  • Fuori diagonale: 0")
    print("  • Riga TOTAL: Somma di tutti i vintage")
    print("\n✅ COMPLETATO - Aprire Excel per verificare i calcoli")
    print("=" * 50)

if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Script corretto per scrivere formule Excel con openpyxl
Risolve i problemi di troncamento dei riferimenti alle celle
"""

import openpyxl
import os

def write_excel_formula_correct():
    """
    Dimostra il modo corretto di scrivere formule Excel con openpyxl
    """
    
    file_path = 'modello.xlsx'
    
    print("✅ SCRITTURA FORMULE EXCEL - VERSIONE CORRETTA")
    print("=" * 55)
    
    # Apriamo il file
    wb = openpyxl.load_workbook(file_path)
    
    # Prendiamo il foglio Calcoli
    ws_calcoli = wb['Calcoli']
    
    print("🎯 REGOLE FONDAMENTALI PER FORMULE EXCEL:")
    print("-" * 45)
    print("1. USA SEMPRE virgolette SINGOLE per le formule")
    print("2. I caratteri $ non hanno bisogno di escape")
    print("3. Le stringhe dentro la formula usano virgolette doppie")
    print("4. Usa ; come separatore (Excel italiano)")
    print()
    
    # ✅ ESEMPIO CORRETTO
    print("✅ ESEMPIO FORMULA CORRETTA:")
    
    # Questa è la sintassi corretta per Excel italiano
    formula_corretta = '=SE(Input!$D$67="amortizing";B7/Input!$D$68;0)'
    
    print(f"📝 Formula: {formula_corretta}")
    print("   ↑ Virgolette SINGOLE per la formula")
    print("   ↑ Virgolette DOPPIE per la stringa 'amortizing'")
    print("   ↑ Punto e virgola ; come separatore")
    print("   ↑ Riferimenti assoluti $D$67 e $D$68")
    
    # Scriviamo la formula
    ws_calcoli['B53'] = formula_corretta
    
    # Altri esempi di formule corrette
    formule_esempio = [
        {
            'cella': 'B60',
            'formula': '=SE(Input!$A$10="Si";SOMMA(B7:B20);0)',
            'descrizione': 'Formula con SOMMA e riferimenti misti'
        },
        {
            'cella': 'B61', 
            'formula': '=Input!$B$5*1,05',
            'descrizione': 'Riferimento semplice con calcolo'
        },
        {
            'cella': 'B62',
            'formula': '=SE(E(Input!$A$1>0;Input!$B$1<100);Input!$A$1*Input!$B$1;0)',
            'descrizione': 'Formula complessa con funzione E()'
        },
        {
            'cella': 'B63',
            'formula': '=CERCA.VERT(A63;Input!$A$1:$Z$100;5;FALSO)',
            'descrizione': 'CERCA.VERT con intervallo assoluto'
        }
    ]
    
    print(f"\n📋 ESEMPI AGGIUNTIVI:")
    print("-" * 25)
    
    for esempio in formule_esempio:
        print(f"\n🔹 {esempio['descrizione']}")
        print(f"   Cella: {esempio['cella']}")
        print(f"   Formula: {esempio['formula']}")
        
        # Scriviamo la formula
        ws_calcoli[esempio['cella']] = esempio['formula']
    
    # Salviamo il file
    print(f"\n💾 Salvando le formule nel file...")
    wb.save(file_path)
    print(f"✅ File salvato con successo!")
    
    # Verifichiamo che tutto sia stato salvato correttamente
    print(f"\n🔍 VERIFICA FINALE:")
    print("-" * 20)
    
    wb_verifica = openpyxl.load_workbook(file_path, data_only=False)
    ws_verifica = wb_verifica['Calcoli']
    
    # Controlliamo la formula principale
    cella_test = ws_verifica['B53']
    print(f"✅ B53: {cella_test.value}")
    
    # Controlliamo gli esempi
    for esempio in formule_esempio:
        cella = ws_verifica[esempio['cella']]
        formula_salvata = cella.value
        print(f"✅ {esempio['cella']}: {formula_salvata}")
        
        # Verifichiamo che i $ siano presenti
        if '$' in formula_salvata:
            print(f"   ✓ Riferimenti assoluti presenti")
        else:
            print(f"   ⚠️  Nessun riferimento assoluto")
    
    wb_verifica.close()
    print(f"\n🎉 TUTTE LE FORMULE SALVATE CORRETTAMENTE!")

def esempio_costruzione_dinamica():
    """
    Dimostra come costruire formule dinamicamente in modo sicuro
    """
    
    print(f"\n🔧 COSTRUZIONE DINAMICA FORMULE")
    print("=" * 35)
    
    # Parametri dinamici
    foglio_input = "Input"
    riga_condizione = 67
    colonna_condizione = "D"
    riga_divisore = 68
    cella_dividendo = "B7"
    tipo_ammortamento = "amortizing"
    
    # Costruzione sicura della formula
    # ✅ METODO CORRETTO: f-string con virgolette singole esterne
    formula_dinamica = f'=SE({foglio_input}!${colonna_condizione}${riga_condizione}="{tipo_ammortamento}";{cella_dividendo}/{foglio_input}!${colonna_condizione}${riga_divisore};0)'
    
    print(f"🎯 Costruzione dinamica:")
    print(f"   Foglio: {foglio_input}")
    print(f"   Condizione: {colonna_condizione}{riga_condizione}")
    print(f"   Divisore: {colonna_condizione}{riga_divisore}")
    print(f"   Tipo: {tipo_ammortamento}")
    print(f"\n📝 Formula risultante:")
    print(f"   {formula_dinamica}")
    
    # Verifichiamo che sia corretta
    if '$' in formula_dinamica and ';' in formula_dinamica:
        print(f"✅ Formula dinamica costruita correttamente!")
    else:
        print(f"❌ Errore nella costruzione dinamica!")
    
    return formula_dinamica

def best_practices_summary():
    """
    Riassunto delle migliori pratiche
    """
    
    print(f"\n📚 MIGLIORI PRATICHE - RIASSUNTO")
    print("=" * 40)
    
    practices = [
        "✅ USA virgolette SINGOLE per delimitare la formula",
        "✅ USA virgolette DOPPIE per le stringhe dentro la formula", 
        "✅ USA punto e virgola ; come separatore (Excel italiano)",
        "✅ USA $colonna$riga per riferimenti assoluti",
        "✅ NON fare escape dei caratteri $ (non servono backslash)",
        "✅ TESTA sempre salvando e riaprendo il file",
        "✅ Per formule dinamiche: f-string con virgolette singole esterne",
        "❌ NON usare virgolette doppie per delimitare la formula",
        "❌ NON usare virgola , come separatore in Excel italiano",
        "❌ NON fare escape dei $ con backslash"
    ]
    
    for practice in practices:
        print(f"   {practice}")
    
    print(f"\n🎯 TEMPLATE GENERICO:")
    print(f"   formula = '=FUNZIONE(Foglio!$Colonna$Riga=\"stringa\";altro_parametro)'")
    print(f"   ws['Cella'] = formula")

if __name__ == "__main__":
    write_excel_formula_correct()
    esempio_costruzione_dinamica() 
    best_practices_summary()
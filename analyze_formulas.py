#!/usr/bin/env python3
"""
Analizza le formule nel file Excel per identificare problemi
"""

from openpyxl import load_workbook
import re

def analyze_formulas():
    """Analizza tutte le formule nel file Excel"""
    
    print("🔍 Analisi dettagliata delle formule...")
    print("="*70)
    
    # Carica il workbook
    wb = load_workbook('modello_bancario_completo.xlsx', data_only=False)
    
    # Analizza foglio Modello (sheet2)
    ws_modello = wb['Modello']
    
    print(f"📊 Analisi foglio 'Modello' (sheet2.xml)")
    print(f"   Dimensioni: {ws_modello.max_row} righe x {ws_modello.max_column} colonne\n")
    
    # Controlla formule riga per riga
    problematic_formulas = []
    
    for row_idx in range(1, min(20, ws_modello.max_row + 1)):
        row_formulas = []
        
        for col_idx in range(1, min(25, ws_modello.max_column + 1)):
            cell = ws_modello.cell(row=row_idx, column=col_idx)
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                # Analizza la formula
                formula = cell.value
                
                # Cerca riferimenti a Input
                if 'Input!' in formula:
                    # Estrai tutti i riferimenti
                    refs = re.findall(r'Input!\$?([A-Z]+)\$?(\d+)', formula)
                    
                    for col_ref, row_ref in refs:
                        row_num = int(row_ref)
                        
                        # Controlla se il riferimento è valido
                        try:
                            ref_value = wb['Input'][f'{col_ref}{row_num}'].value
                            if ref_value is None:
                                row_formulas.append({
                                    'cell': f"{cell.column_letter}{cell.row}",
                                    'formula': formula[:100],
                                    'issue': f"Riferimento a cella vuota Input!{col_ref}{row_num}"
                                })
                        except:
                            row_formulas.append({
                                'cell': f"{cell.column_letter}{cell.row}",
                                'formula': formula[:100],
                                'issue': f"Riferimento non valido Input!{col_ref}{row_num}"
                            })
        
        if row_formulas:
            print(f"❌ Riga {row_idx}: {len(row_formulas)} formule con potenziali problemi")
            for issue in row_formulas[:3]:  # Mostra max 3 per riga
                print(f"   - Cella {issue['cell']}: {issue['issue']}")
                print(f"     Formula: {issue['formula'][:60]}...")
            problematic_formulas.extend(row_formulas)
    
    print("\n" + "="*70)
    
    # Analisi specifica riga 6
    print("\n🔎 Analisi dettagliata RIGA 6:")
    print("-"*50)
    
    for col_idx in range(1, min(25, ws_modello.max_column + 1)):
        cell = ws_modello.cell(row=6, column=col_idx)
        
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"\nCella {cell.column_letter}6:")
                print(f"  Formula: {cell.value}")
                
                # Verifica ogni riferimento nella formula
                if 'Input!' in cell.value:
                    refs = re.findall(r'Input!\$?([A-Z]+)\$?(\d+)', cell.value)
                    for col_ref, row_ref in refs:
                        try:
                            check_cell = wb['Input'][f'{col_ref}{row_ref}']
                            check_value = check_cell.value
                            if check_value is None:
                                print(f"  ⚠️ PROBLEMA: Input!{col_ref}{row_ref} è VUOTA")
                            else:
                                print(f"  ✓ Input!{col_ref}{row_ref} = {check_value}")
                        except Exception as e:
                            print(f"  ❌ ERRORE: Input!{col_ref}{row_ref} - {str(e)}")
            elif cell.value:
                print(f"Cella {cell.column_letter}6: {cell.value}")
    
    # Verifica riferimenti nelle prime righe
    print("\n" + "="*70)
    print("\n📋 Riepilogo problemi trovati:")
    
    if problematic_formulas:
        # Raggruppa per tipo di problema
        issues_by_type = {}
        for issue in problematic_formulas:
            issue_type = issue['issue'].split('Input!')[0] if 'Input!' in issue['issue'] else 'Altro'
            if issue_type not in issues_by_type:
                issues_by_type[issue_type] = []
            issues_by_type[issue_type].append(issue)
        
        for issue_type, issues in issues_by_type.items():
            print(f"\n{issue_type}: {len(issues)} occorrenze")
            for issue in issues[:5]:  # Mostra max 5 esempi
                print(f"  - {issue['cell']}: {issue['issue']}")
    else:
        print("\n✅ Nessun problema evidente trovato nelle formule analizzate")
    
    # Verifica struttura Input
    print("\n" + "="*70)
    print("\n📊 Verifica foglio Input:")
    ws_input = wb['Input']
    
    # Cerca righe chiave
    key_rows = {}
    for row in range(1, min(200, ws_input.max_row + 1)):
        cell_b = ws_input[f'B{row}'].value
        if cell_b:
            cell_str = str(cell_b).upper()
            if "NUOVE EROGAZIONI RE" in cell_str:
                key_rows['erog_re'] = row
                # Verifica valori
                print(f"\n✓ Trovato 'Nuove Erogazioni RE' alla riga {row}")
                for col in ['C', 'D', 'E', 'F', 'G']:
                    val = ws_input[f'{col}{row}'].value
                    print(f"  {col}{row} = {val}")
            elif "MIX PRODOTTI" in cell_str:
                print(f"\n✓ Trovato 'Mix Prodotti' alla riga {row}")
                # Verifica contesto
                context = ws_input[f'B{row-2}'].value if row > 2 else None
                print(f"  Contesto: {context}")
                for col in ['C', 'D', 'E', 'F', 'G', 'H']:
                    val = ws_input[f'{col}{row}'].value
                    if val:
                        print(f"  {col}{row} = {val}")

def main():
    """Funzione principale"""
    print("\n" + "="*70)
    print("🔍 ANALISI FORMULE EXCEL")
    print("="*70 + "\n")
    
    try:
        analyze_formulas()
        
        print("\n" + "="*70)
        print("✅ ANALISI COMPLETATA")
        print("="*70)
        
    except Exception as e:
        print(f"\n❌ Errore durante l'analisi: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
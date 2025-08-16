#!/usr/bin/env python3
"""
Analisi dettagliata delle matrici e struttura del modello.xlsx
Focus specifico su:
- Struttura matrici Stock GBV
- Layout Erogazioni e Rimborsi
- Pattern delle formule esistenti
- Struttura temporale delle colonne
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re

def analizza_dettaglio_modello():
    file_path = "modello.xlsx"
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws_calcoli = wb['Calcoli']
        
        print("🔍 ANALISI DETTAGLIATA STRUTTURA MODELLO")
        print("=" * 60)
        
        # 1. Analisi struttura matrici Stock GBV
        print("\n📊 1. STRUTTURA MATRICI STOCK GBV (Riga 233)")
        print("-" * 40)
        
        # Analizza l'header e la struttura intorno alla riga 233
        for row in range(231, 240):
            row_content = []
            for col in range(1, 50):  # Prime 50 colonne
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.value is not None:
                    col_letter = get_column_letter(col)
                    value = str(cell.value)
                    if len(value) > 30:
                        value = value[:30] + "..."
                    row_content.append(f"{col_letter}{row}: {value}")
            
            if row_content:
                print(f"Riga {row}: {' | '.join(row_content[:5])}")
        
        # 2. Analisi matrici Erogazioni (riga 3 e dintorni)
        print("\n📊 2. STRUTTURA MATRICI EROGAZIONI (Riga 3)")
        print("-" * 40)
        
        for row in range(1, 10):
            row_content = []
            for col in range(1, 50):
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.value is not None:
                    col_letter = get_column_letter(col)
                    value = str(cell.value)
                    if len(value) > 30:
                        value = value[:30] + "..."
                    row_content.append(f"{col_letter}{row}: {value}")
            
            if row_content:
                print(f"Riga {row}: {' | '.join(row_content[:5])}")
        
        # 3. Analisi struttura temporale delle colonne
        print("\n📊 3. STRUTTURA TEMPORALE COLONNE (Riga 97-98)")
        print("-" * 40)
        
        # Esamina le intestazioni temporali
        header_row_97 = []
        header_row_98 = []
        
        for col in range(1, 50):  # Prime 50 colonne
            cell_97 = ws_calcoli.cell(row=97, column=col)
            cell_98 = ws_calcoli.cell(row=98, column=col)
            col_letter = get_column_letter(col)
            
            if cell_97.value is not None or cell_98.value is not None:
                val_97 = str(cell_97.value) if cell_97.value else "[vuota]"
                val_98 = str(cell_98.value) if cell_98.value else "[vuota]"
                header_row_97.append(f"{col_letter}: {val_97}")
                header_row_98.append(f"{col_letter}: {val_98}")
        
        print("Riga 97 (Anni):")
        for i in range(0, len(header_row_97), 4):
            print(f"  {' | '.join(header_row_97[i:i+4])}")
        
        print("Riga 98 (Trimestri):")
        for i in range(0, len(header_row_98), 4):
            print(f"  {' | '.join(header_row_98[i:i+4])}")
        
        # 4. Analisi pattern delle formule
        print("\n📊 4. PATTERN FORMULE ESISTENTI")
        print("-" * 40)
        
        # Raggruppa le formule per pattern
        formula_patterns = {}
        
        for row in range(1, 100):  # Prime 100 righe
            for col in range(1, 100):  # Prime 100 colonne
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    
                    # Identifica pattern comuni
                    if 'Input!' in formula:
                        pattern_key = "Riferimenti Input"
                    elif 'SUM(' in formula:
                        pattern_key = "Somme"
                    elif 'IF(' in formula:
                        pattern_key = "Condizionali"
                    elif '*' in formula:
                        pattern_key = "Moltiplicazioni"
                    else:
                        pattern_key = "Altri"
                    
                    if pattern_key not in formula_patterns:
                        formula_patterns[pattern_key] = []
                    
                    col_letter = get_column_letter(col)
                    if len(formula_patterns[pattern_key]) < 3:  # Massimo 3 esempi per pattern
                        formula_patterns[pattern_key].append(f"{col_letter}{row}: {formula[:60]}...")
        
        for pattern, examples in formula_patterns.items():
            print(f"\n  🔧 {pattern}:")
            for example in examples:
                print(f"    {example}")
        
        # 5. Identificazione matrici e loro posizioni
        print("\n📊 5. POSIZIONI MATRICI IDENTIFICATE")
        print("-" * 40)
        
        matrici_info = []
        
        # Cerca tutte le celle che contengono "Matrice"
        for row in range(1, 500):
            for col in range(1, 100):
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and 'Matrice' in cell.value:
                    col_letter = get_column_letter(col)
                    matrici_info.append({
                        'posizione': f"{col_letter}{row}",
                        'nome': cell.value,
                        'riga': row,
                        'colonna': col
                    })
        
        # Raggruppa per tipo di matrice
        matrici_per_tipo = {}
        for matrice in matrici_info:
            nome = matrice['nome']
            if 'EROGAZIONI' in nome:
                tipo = 'EROGAZIONI'
            elif 'RIMBORSI' in nome:
                tipo = 'RIMBORSI'
            elif 'STOCK GBV' in nome:
                tipo = 'STOCK GBV'
            elif 'DEFAULTED' in nome:
                tipo = 'DEFAULTED'
            else:
                tipo = 'ALTRE'
            
            if tipo not in matrici_per_tipo:
                matrici_per_tipo[tipo] = []
            matrici_per_tipo[tipo].append(matrice)
        
        for tipo, matrici in matrici_per_tipo.items():
            print(f"\n  📋 {tipo} ({len(matrici)} matrici):")
            for matrice in matrici[:5]:  # Prime 5 per tipo
                print(f"    {matrice['posizione']}: {matrice['nome']}")
        
        # 6. Analisi area riga 99 dettagliata
        print("\n📊 6. ANALISI DETTAGLIATA RIGA 99")
        print("-" * 40)
        
        # Cerca contenuto nelle righe circostanti alla 99
        context_area = {}
        for row in range(95, 105):
            for col in range(1, 50):
                cell = ws_calcoli.cell(row=row, column=col)
                if cell.value is not None:
                    if row not in context_area:
                        context_area[row] = {}
                    context_area[row][col] = cell.value
        
        for row in sorted(context_area.keys()):
            print(f"\nRiga {row}:")
            for col in sorted(context_area[row].keys())[:10]:  # Prime 10 colonne con contenuto
                col_letter = get_column_letter(col)
                value = str(context_area[row][col])
                if len(value) > 20:
                    value = value[:20] + "..."
                print(f"  {col_letter}: {value}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Errore durante l'analisi: {str(e)}")

if __name__ == "__main__":
    analizza_dettaglio_modello()